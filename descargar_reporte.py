#!/usr/bin/env python3
"""
Script para automatizar la descarga y conversión del Reporte de Vendedores de Cargolink a Excel.
Uso: python3 descargar_reporte.py [YYYY-MM-DD_inicio] [YYYY-MM-DD_fin]
"""

import sys
import os
import csv
import shutil
import requests
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# Credenciales y URLs (definidas en variables de entorno, ver .env.example)
LOGIN_URL = 'https://fwd.cargolink.mx/seguridad/control.php?loginfrom=usuario'
USUARIO = os.environ.get('CARGOLINK_USUARIO')
PASSWORD = os.environ.get('CARGOLINK_PASSWORD')

def main():
    if not USUARIO or not PASSWORD:
        print("Error: faltan las variables de entorno CARGOLINK_USUARIO y/o CARGOLINK_PASSWORD.")
        sys.exit(1)

    # Fechas por defecto: todo el año actual a la fecha
    now = datetime.now()
    fecha_inicio = sys.argv[1] if len(sys.argv) > 1 else f"{now.year}-01-01"
    fecha_fin = sys.argv[2] if len(sys.argv) > 2 else now.strftime('%Y-%m-%d')
    if fecha_fin < '2026-07-30' and now.year == 2026:
        fecha_fin = '2026-07-30'

    print(f"--> Iniciando sesión en Cargolink ({USUARIO})...")
    session = requests.Session()
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)'}
    
    r_login = session.post(LOGIN_URL, data={'usuario': USUARIO, 'password': PASSWORD}, headers=headers)
    if r_login.status_code != 200 or '"activo"' not in r_login.text:
        print("Error al iniciar sesión:", r_login.text)
        sys.exit(1)
    print("✓ Sesión iniciada correctamente.")

    report_url = (
        f"https://fwd.cargolink.mx/templates/pdfs/excel_vendedores.php?"
        f"fecha_inicio={fecha_inicio}&fecha_fin={fecha_fin}&"
        f"ejecutivo=undefined&vendedor=undefined&id_cliente=undefined&"
        f"sucursal=undefined&id_cliente_factura=undefined&status_booking=1,2"
    )

    print(f"--> Descargando reporte del {fecha_inicio} al {fecha_fin}...")
    res = session.get(report_url, headers=headers)
    if res.status_code != 200 or len(res.content) == 0:
        print("Error al descargar reporte del servidor.")
        sys.exit(1)
    print(f"✓ Reporte descargado ({len(res.content) / (1024*1024):.2f} MB).")

    # Extraer filas HTML
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(res.content, 'html.parser')
    rows_data = []
    for tr in soup.find_all('tr'):
        row = [td.get_text(strip=True) for td in tr.find_all(['th', 'td'])]
        if row:
            rows_data.append(row)

    output_xlsx = f"Reporte_Vendedores_{fecha_inicio}_al_{fecha_fin}.xlsx"
    workspace_dir = os.environ.get("REPORTES_DIR") or os.path.dirname(os.path.abspath(__file__))
    os.makedirs(workspace_dir, exist_ok=True)
    dest_path = os.path.join(workspace_dir, output_xlsx)
    default_dest_path = os.path.join(workspace_dir, "Reporte_de_Vendedores_2026.xlsx")

    print(f"--> Generando libro Excel: {output_xlsx}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Vendedores"

    # Estilos
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007CC3", end_color="007CC3", fill_type="solid")
    title_font = Font(name="Arial", size=14, bold=True, color="007CC3")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    current_row = 1
    header_map = {}

    for r in rows_data:
        if not r or not any(r):
            continue
        if len(r) == 1 and 'REPORTE' in r[0].upper():
            c = ws.cell(row=current_row, column=1, value=r[0])
            c.font = title_font
            current_row += 2
            continue
        
        is_header = ('Referencia' in r or 'Tipo Servicio' in r or 'Vendedor' in r)
        for col_idx, val in enumerate(r, 1):
            c = ws.cell(row=current_row, column=col_idx)
            val_str = val.strip()
            if is_header:
                c.value = val_str
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                header_map[col_idx] = val_str
            else:
                col_header = header_map.get(col_idx, '').lower()

                # Si es campo de Fecha (ej. Fecha de creacion, Fecha carga, Fecha descarga)
                is_date_col = 'fecha' in col_header
                date_parsed = False

                if is_date_col and val_str and not val_str.startswith('0000-00-00'):
                    try:
                        date_part = val_str.split(' ')[0] # Extraer YYYY-MM-DD sin hora
                        d_obj = datetime.strptime(date_part, '%Y-%m-%d').date()
                        c.value = d_obj
                        c.number_format = 'dd/mm/yy'
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        date_parsed = True
                    except ValueError:
                        date_parsed = False

                if not date_parsed:
                    clean_val = val_str.replace(',', '').replace('$', '')
                    try:
                        if '.' in clean_val:
                            c.value = float(clean_val)
                            if col_idx in [22, 24, 25, 26]:
                                c.number_format = '$#,##0.00'
                            elif col_idx == 27:
                                c.number_format = '0.00%'
                            else:
                                c.number_format = '#,##0.00'
                        elif clean_val.lstrip('-').isdigit():
                            c.value = int(clean_val)
                            c.number_format = '#,##0'
                        else:
                            c.value = val_str
                    except ValueError:
                        c.value = val_str

                    c.font = data_font
                    c.border = thin_border
                    if not is_date_col:
                        c.alignment = Alignment(horizontal="right" if isinstance(c.value, (int, float)) else "left", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="center", vertical="center")

                c.font = data_font
                c.border = thin_border

        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3:
                continue
            if cell.value:
                if isinstance(cell.value, datetime) or hasattr(cell.value, 'strftime'):
                    val_str_repr = cell.value.strftime('%d/%m/%y')
                else:
                    val_str_repr = str(cell.value)
                max_len = max(max_len, len(val_str_repr))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    ws.freeze_panes = 'A4'
    wb.save(dest_path)
    shutil.copyfile(dest_path, default_dest_path)
    print(f"✓ Excel guardado exitosamente en: {dest_path}")
    print(f"✓ Copia actualizada guardada en: {default_dest_path}")

if __name__ == '__main__':
    main()
