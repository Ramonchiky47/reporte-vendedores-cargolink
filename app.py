#!/usr/bin/env python3
"""App con login y un botón para generar el Reporte de Vendedores de CargoLink.

Backend de datos: Postgres (Supabase), vía DATABASE_URL. Pensada para correr
tanto local como en una plataforma serverless (Vercel): no depende de disco
persistente ni de subprocesos — la descarga de CargoLink corre inline y los
resultados se guardan directo en la base de datos.
"""

import base64
import calendar
import csv
import hashlib
import hmac
import io
import json
import os
import random
import re
import secrets
import time
from datetime import date, datetime, timedelta
from functools import wraps
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import psycopg
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from flask_wtf import CSRFProtect
from markupsafe import Markup, escape
from psycopg.rows import dict_row
from xhtml2pdf import pisa

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

DATABASE_URL = (os.environ.get("DATABASE_URL") or "").strip()
TZ_LOCAL = ZoneInfo(os.environ.get("TZ_LOCAL", "America/Mexico_City"))
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def fecha_valida_o_vacia(texto):
    """Regresa `texto` si es una fecha YYYY-MM-DD real (no solo con el formato
    correcto: rechaza cosas como "2026-99-99"), o "" si no. Para sanear
    parámetros de filtro que vienen de la URL antes de mandarlos a SQL."""
    texto = (texto or "").strip()
    if not DATE_RE.match(texto):
        return ""
    try:
        datetime.strptime(texto, "%Y-%m-%d")
    except ValueError:
        return ""
    return texto

CARGOLINK_LOGIN_URL = "https://fwd.cargolink.mx/seguridad/control.php?loginfrom=usuario"
CARGOLINK_REPORT_URL = "https://fwd.cargolink.mx/templates/pdfs/excel_vendedores.php"
CARGOLINK_REPORTE_CLIENTES_URL = "https://fwd.cargolink.mx/templates/pdfs/ReporteClientesExcel.php"
CARGOLINK_LIQ_VENDEDOR_URL = "https://fwd.cargolink.mx/templates/egresos_liq_vendedor/"


def get_secret_key():
    env_key = (os.environ.get("SECRET_KEY") or "").strip()
    if env_key:
        return env_key
    # Sin SECRET_KEY fija, las sesiones no sobreviven un reinicio del proceso
    # (normal en serverless). Sirve para correr rápido en local sin configurar nada.
    return secrets.token_hex(32)


def get_db():
    if not DATABASE_URL:
        raise RuntimeError("Falta la variable de entorno DATABASE_URL.")
    return psycopg.connect(DATABASE_URL, row_factory=dict_row)


def init_db():
    """Crea las tablas si no existen (primer arranque en una base vacía).

    El login usa el catálogo de accesos de Seguimiento de Importaciones
    (auth.users + public.app_user_permissions, mismo proyecto Supabase) en
    vez de una tabla de usuarios propia."""
    db = get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS catalogo_vendedores (
            id bigint generated always as identity primary key,
            vendedor text not null unique,
            plaza text not null,
            ocultar_detalle boolean not null default false
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS catalogo_desarrolladores (
            id bigint generated always as identity primary key,
            desarrollador text not null unique,
            plaza text not null
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS catalogo_presupuesto (
            id bigint generated always as identity primary key,
            mes text not null,
            vendedor text,
            desarrollador text,
            presupuesto numeric not null,
            CONSTRAINT chk_catalogo_presupuesto_vendedor_o_desarrollador
                CHECK (vendedor IS NOT NULL OR desarrollador IS NOT NULL)
        );
    """)
    db.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_catalogo_presupuesto_mes_vendedor_dev
            ON catalogo_presupuesto (mes, coalesce(vendedor, ''), coalesce(desarrollador, ''));
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS reporte_bookings (
            id bigint generated always as identity primary key,
            mes text not null,
            vendedor text not null,
            referencia text,
            fecha timestamptz,
            ejecutivo text,
            venta_por text,
            cliente_servicio text,
            venta numeric not null default 0,
            profit numeric not null default 0,
            margen numeric not null default 0
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_reporte_bookings_mes_vendedor ON reporte_bookings (mes, vendedor);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS reporte_generaciones (
            id bigint generated always as identity primary key,
            fecha_inicio date not null,
            fecha_fin date not null,
            generado_en timestamptz not null default now()
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS asignacion_de_clientes (
            id bigint generated always as identity primary key,
            folio integer,
            razon_social text not null,
            vendedor text,
            desarrollador text,
            tipo_cliente text,
            fecha_creacion timestamptz not null default now(),
            cant_booking integer not null default 0,
            fecha_ultimo_booking timestamptz
        );
    """)
    db.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_asignacion_de_clientes_folio
            ON asignacion_de_clientes (folio);
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS usuario_plazas (
            id bigint generated always as identity primary key,
            usuario_id bigint not null references usuarios(id) on delete cascade,
            plaza text not null,
            unique (usuario_id, plaza)
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS registro_ingresos (
            id bigint generated always as identity primary key,
            usuario_id bigint references usuarios(id) on delete set null,
            usuario text not null,
            fecha_hora timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_registro_ingresos_fecha ON registro_ingresos (fecha_hora desc);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS intentos_login (
            id bigint generated always as identity primary key,
            email text not null,
            exitoso boolean not null,
            creado_en timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_intentos_login_email_fecha ON intentos_login (lower(email), creado_en desc);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS comisiones_liquidacion_detalle (
            id bigint generated always as identity primary key,
            folio integer not null,
            descripcion text not null,
            booking text not null,
            folio_cobro text,
            profit numeric not null default 0,
            vendedor text,
            pct_vendedor numeric,
            total_vendedor numeric not null default 0,
            desarrollador text,
            pct_desarrollador numeric,
            total_desarrollador numeric not null default 0,
            cargado_en timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_comisiones_liq_folio ON comisiones_liquidacion_detalle (folio);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS comisiones_cobros_detalle (
            id bigint generated always as identity primary key,
            folio integer not null,
            etiqueta text not null,
            folio_cobro text,
            folio_factura text,
            referencia text not null,
            uuid text,
            tipo_docto text,
            tipo_referencia text,
            cliente text,
            fecha_factura date,
            fecha_cobro date,
            dias_diferencia integer,
            fecha_timbre text,
            banco text,
            moneda text,
            subtotal numeric not null default 0,
            iva numeric not null default 0,
            descuento numeric not null default 0,
            retencion numeric not null default 0,
            total numeric not null default 0,
            cargado_en timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_comisiones_cobros_folio ON comisiones_cobros_detalle (folio);")
    db.execute("CREATE INDEX IF NOT EXISTS idx_comisiones_cobros_referencia ON comisiones_cobros_detalle (referencia);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_grupos (
            id bigint generated always as identity primary key,
            nombre text not null unique
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_contactos (
            id bigint generated always as identity primary key,
            nombre text not null,
            apellido text,
            telefono text,
            correo text,
            observaciones text,
            creado_en timestamptz not null default now()
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_contacto_clientes (
            contacto_id bigint not null references crm_contactos(id) on delete cascade,
            cliente_folio integer not null references asignacion_de_clientes(folio) on delete cascade,
            primary key (contacto_id, cliente_folio)
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_contacto_grupos (
            contacto_id bigint not null references crm_contactos(id) on delete cascade,
            grupo_id bigint not null references crm_grupos(id) on delete cascade,
            primary key (contacto_id, grupo_id)
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_incoterms (
            id bigint generated always as identity primary key,
            nombre text not null unique
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_modalidades (
            id bigint generated always as identity primary key,
            nombre text not null unique
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_tipos_ingreso_egreso (
            id bigint generated always as identity primary key,
            nombre text not null,
            id_externo integer,
            concepto_ingles text,
            naturaleza text,
            producto text,
            porcentaje_iva numeric,
            porcentaje_ret numeric,
            bloqueado boolean not null default false
        );
    """)
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso DROP CONSTRAINT IF EXISTS crm_tipos_ingreso_egreso_nombre_key;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS id_externo integer;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS concepto_ingles text;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS naturaleza text;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS producto text;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS porcentaje_iva numeric;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS porcentaje_ret numeric;")
    db.execute("ALTER TABLE crm_tipos_ingreso_egreso ADD COLUMN IF NOT EXISTS bloqueado boolean not null default false;")
    db.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_crm_tipos_ingreso_egreso_id_externo "
        "ON crm_tipos_ingreso_egreso (id_externo) WHERE id_externo IS NOT NULL;"
    )
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_cotizaciones (
            id bigint generated always as identity primary key,
            id_cotizacion text not null unique,
            nombre_cotizacion text,
            fecha_creacion date not null default current_date,
            fecha_vencimiento date,
            vencimiento_modo text not null default 'libre',
            cliente_folio integer references asignacion_de_clientes(folio) on delete set null,
            cliente_prospecto text,
            contacto_id bigint references crm_contactos(id) on delete set null,
            origen text,
            destino text,
            hazmat boolean not null default false,
            hazmat_clase text,
            hazmat_un_imo text,
            incoterm_id bigint references crm_incoterms(id) on delete set null,
            modalidad_id bigint references crm_modalidades(id) on delete set null,
            tipo_ingreso_egreso_id bigint references crm_tipos_ingreso_egreso(id) on delete set null,
            tipo_ingreso_egreso_texto text,
            estibable boolean not null default false,
            tiempo_traslado text,
            via text,
            seguro_mercancia boolean not null default false,
            profit_estimado numeric,
            tipo_cambio numeric,
            descripcion text,
            creado_en timestamptz not null default now()
        );
    """)
    tipo_id_cotizacion = db.execute("""
        SELECT data_type FROM information_schema.columns
        WHERE table_name = 'crm_cotizaciones' AND column_name = 'id_cotizacion'
    """).fetchone()
    if tipo_id_cotizacion and tipo_id_cotizacion["data_type"] != "text":
        # Migración única: el ID de cotización pasó de numérico (6 dígitos) a
        # texto con prefijo "COT-" (12 caracteres). Los folios ya generados
        # conservan su número, solo se les da el nuevo formato.
        db.execute(
            "ALTER TABLE crm_cotizaciones ALTER COLUMN id_cotizacion TYPE text "
            "USING ('COT-' || lpad(id_cotizacion::text, 8, '0'));"
        )
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS cliente_folio integer references asignacion_de_clientes(folio) on delete set null;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS cliente_prospecto text;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS nombre_cotizacion text;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS contacto_id bigint references crm_contactos(id) on delete set null;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS tipo_ingreso_egreso_id bigint references crm_tipos_ingreso_egreso(id) on delete set null;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS tipo_ingreso_egreso_texto text;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS creado_por_user_id uuid;")
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_motivos_perdida (
            id bigint generated always as identity primary key,
            nombre text not null unique,
            creado_en timestamptz not null default now()
        );
    """)
    # Estatus de una cotización: 'vigente' y 'perdida' se guardan en esta
    # columna (perdida es la única acción manual real); "vencido" se
    # calcula al vuelo comparando fecha_vencimiento con hoy, y "ganada" se
    # calcula según si tiene renglones en crm_cotizacion_bookings — nunca
    # se guarda "ganada" ni "vencido" aquí.
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS estatus text not null default 'vigente';")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS motivo_perdida_id bigint references crm_motivos_perdida(id) on delete set null;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS perdida_en timestamptz;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS comentario_perdida text;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS mostrar_columna_impuesto boolean not null default true;")
    db.execute("ALTER TABLE crm_cotizaciones ADD COLUMN IF NOT EXISTS mostrar_totales boolean not null default true;")
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_cotizacion_bookings (
            id bigint generated always as identity primary key,
            cotizacion_id bigint not null references crm_cotizaciones(id) on delete cascade,
            booking_referencia text not null unique,
            aplicado_en timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_crm_cotizacion_bookings_cotizacion ON crm_cotizacion_bookings (cotizacion_id);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_solicitudes_maritimo_aereo (
            id bigint generated always as identity primary key,
            referencia text not null unique,
            cotizacion_id bigint references crm_cotizaciones(id) on delete set null,
            creado_por text,
            importacion_exportacion text,
            incoterm_id bigint references crm_incoterms(id) on delete set null,
            tipo_embarque text,
            pais_origen text,
            pais_destino text,
            lugar_recoleccion text,
            puerto_carga text,
            puerto_descarga text,
            lugar_entrega text,
            naviera_aerolinea text,
            fcl_numero_tipo_contenedores text,
            fcl_dias_libres_requeridos integer,
            producto text,
            lcl_air_dimensiones text,
            estibable boolean,
            requiere_inbond_usa boolean,
            hazmat boolean not null default false,
            carga_reefer boolean,
            temperatura text,
            requerimientos_especiales text,
            agente_a_cotizar text,
            descripcion_material text,
            anexos_notas text,
            fecha_creacion date not null default current_date,
            estado text not null default 'Solicitud',
            propiedad text,
            creado_en timestamptz not null default now()
        );
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_crm_solicitudes_maritimo_aereo_cotizacion ON crm_solicitudes_maritimo_aereo (cotizacion_id);")
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_firmas (
            user_id uuid primary key,
            nombre_firma text,
            puesto text,
            telefono text,
            correo text,
            updated_at timestamptz not null default now()
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_plazas_habilitadas (
            plaza text primary key,
            creado_en timestamptz not null default now()
        );
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_cotizacion_productos (
            id bigint generated always as identity primary key,
            cotizacion_id bigint not null references crm_cotizaciones(id) on delete cascade,
            producto_id bigint references crm_tipos_ingreso_egreso(id) on delete set null,
            producto_texto text,
            cantidad numeric not null default 1,
            precio_unitario numeric not null default 0,
            moneda text not null default 'MXN',
            causa_impuesto boolean not null default false,
            impuesto numeric not null default 0,
            orden integer not null default 0
        );
    """)
    db.execute("ALTER TABLE crm_cotizacion_productos ADD COLUMN IF NOT EXISTS moneda text not null default 'MXN';")
    db.execute("ALTER TABLE crm_cotizacion_productos ADD COLUMN IF NOT EXISTS causa_impuesto boolean not null default false;")
    db.execute("ALTER TABLE crm_cotizacion_productos ADD COLUMN IF NOT EXISTS impuesto numeric not null default 0;")
    db.execute("ALTER TABLE crm_cotizacion_productos ADD COLUMN IF NOT EXISTS producto_texto text;")
    db.execute("ALTER TABLE crm_cotizacion_productos ADD COLUMN IF NOT EXISTS observaciones text;")
    db.execute("CREATE INDEX IF NOT EXISTS idx_crm_cotizacion_productos_cotizacion ON crm_cotizacion_productos (cotizacion_id);")
    db.execute("""
        INSERT INTO crm_incoterms (nombre) VALUES
            ('EXW'), ('FCA'), ('FAS'), ('FOB'), ('CFR'), ('CIF'), ('CPT'), ('CIP'), ('DAP'), ('DPU'), ('DDP')
        ON CONFLICT (nombre) DO NOTHING;
    """)
    db.execute("""
        INSERT INTO crm_modalidades (nombre) VALUES
            ('Aéreo'), ('Marítimo'), ('Terrestre'), ('Multimodal'), ('Paquetería')
        ON CONFLICT (nombre) DO NOTHING;
    """)
    db.commit()
    db.close()


def parse_presupuesto(raw):
    cleaned = (raw or "").replace(",", "").replace("$", "").replace(" ", "").strip()
    return float(cleaned)


MESES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def get_vendedores(db):
    filas = db.execute("SELECT vendedor FROM catalogo_vendedores ORDER BY vendedor").fetchall()
    return [f["vendedor"] for f in filas]


def get_desarrolladores(db):
    filas = db.execute("SELECT desarrollador FROM catalogo_desarrolladores ORDER BY desarrollador").fetchall()
    return [f["desarrollador"] for f in filas]


def opciones_mes():
    anio_actual = datetime.now(TZ_LOCAL).year
    opciones = []
    for anio in range(anio_actual - 1, anio_actual + 2):
        for mes_num, nombre in enumerate(MESES_ES, start=1):
            opciones.append({"value": f"{anio}-{mes_num:02d}", "label": f"{nombre} - {anio}"})
    return opciones


MESES_VALIDOS = {opt["value"] for opt in opciones_mes()}
PRESUPUESTO_COLUMNAS = ["mes", "vendedor", "desarrollador", "presupuesto"]


def leer_filas_csv(file_storage):
    stream = io.StringIO(file_storage.stream.read().decode("utf-8-sig"))
    reader = csv.DictReader(stream)
    return [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]


def leer_filas_xlsx(file_storage):
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = wb.active
    filas_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h is not None else "" for h in next(filas_iter)]
    resultado = []
    for row in filas_iter:
        if not any(v not in (None, "") for v in row):
            continue
        resultado.append(dict(zip(header, row)))
    return resultado


MESES_LARGOS_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def fecha_larga_es(fecha):
    """'2026-08-14' (date) -> '14 de agosto de 2026', sin depender del
    locale del sistema (no confiable en serverless)."""
    if not fecha:
        return ""
    return f"{fecha.day} de {MESES_LARGOS_ES[fecha.month - 1]} de {fecha.year}"


def normalizar(texto):
    return (texto or "").strip().upper()


def json_para_js(datos):
    """json.dumps a salvo de </script> embebido dentro de un <script> inline."""
    return json.dumps(datos).replace("</", "<\\/")


def extraer_tipo_servicio(referencia):
    """El tipo de servicio (FCLI, LCLI, DA, AI, FTL, ...) es el tercer
    segmento de la referencia, ej. '2608-3798-FCLI' -> 'FCLI'."""
    partes = (referencia or "").split("-")
    if len(partes) >= 3:
        return "-".join(partes[2:]).strip().upper() or "Sin tipo"
    return "Sin tipo"


def descargar_bookings_cargolink(fecha_inicio, fecha_fin):
    """Se conecta a CargoLink, descarga el reporte de vendedores del rango
    dado y regresa una lista de dicts (uno por booking). No toca el disco."""
    usuario = (os.environ.get("CARGOLINK_USUARIO") or "").strip()
    password = (os.environ.get("CARGOLINK_PASSWORD") or "").strip()
    if not usuario or not password:
        raise RuntimeError("Faltan las variables de entorno CARGOLINK_USUARIO / CARGOLINK_PASSWORD.")

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    sesion = requests.Session()
    r_login = sesion.post(CARGOLINK_LOGIN_URL, data={"usuario": usuario, "password": password}, headers=headers, timeout=60)
    if r_login.status_code != 200 or '"activo"' not in r_login.text:
        raise RuntimeError("No se pudo iniciar sesión en CargoLink.")

    report_url = (
        f"{CARGOLINK_REPORT_URL}?fecha_inicio={fecha_inicio}&fecha_fin={fecha_fin}&"
        f"ejecutivo=undefined&vendedor=undefined&id_cliente=undefined&"
        f"sucursal=undefined&id_cliente_factura=undefined&status_booking=1,2"
    )
    res = sesion.get(report_url, headers=headers, timeout=120)
    if res.status_code != 200 or len(res.content) == 0:
        raise RuntimeError("Error al descargar el reporte desde CargoLink.")

    soup = BeautifulSoup(res.content, "html.parser")
    header_map = None
    bookings = []
    campos_necesarios = {
        "Referencia": "referencia", "Fecha de creacion": "fecha", "Vendedor": "vendedor",
        "Ejecutivo": "ejecutivo", "Venta por": "venta_por", "Cliente servicio": "cliente_servicio",
        "Venta": "venta", "Profit": "profit", "Margen": "margen",
    }

    for tr in soup.find_all("tr"):
        celdas = [td.get_text(strip=True) for td in tr.find_all(["th", "td"])]
        if not celdas:
            continue
        if len(celdas) == 1:
            continue  # fila de título ("REPORTE DE VENDEDORES")
        if "Referencia" in celdas or "Vendedor" in celdas:
            header_map = {i: nombre for i, nombre in enumerate(celdas)}
            continue
        if header_map is None:
            continue

        fila = {campos_necesarios[header_map[i]]: v for i, v in enumerate(celdas) if header_map.get(i) in campos_necesarios}
        if not fila.get("vendedor") or not fila.get("fecha"):
            continue

        fecha_texto = fila["fecha"].split(" ")[0]
        try:
            fecha_dt = datetime.strptime(fila["fecha"], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            try:
                fecha_dt = datetime.strptime(fecha_texto, "%Y-%m-%d")
            except ValueError:
                continue

        def num(v):
            v = (v or "0").replace(",", "").replace("$", "").replace("%", "").strip()
            try:
                return float(v)
            except ValueError:
                return 0.0

        bookings.append({
            "mes": fecha_dt.strftime("%Y-%m"),
            "vendedor": fila["vendedor"],
            "referencia": fila.get("referencia", ""),
            "fecha": fecha_dt,
            "ejecutivo": fila.get("ejecutivo", ""),
            "venta_por": fila.get("venta_por", ""),
            "cliente_servicio": fila.get("cliente_servicio", ""),
            "venta": num(fila.get("venta")),
            "profit": num(fila.get("profit")),
            "margen": num(fila.get("margen")),
        })

    return bookings


def descargar_reporte_clientes_cargolink():
    """Se conecta a CargoLink, descarga el Reporte de Clientes (excel real,
    no HTML) y regresa una lista de dicts (uno por cliente). No toca el
    disco. Columnas del excel: FOLIO, RAZÓN SOCIAL, VENDEDOR, DESARROLLADOR,
    TIPO DE CLIENTE, SUCURSAL, FECHA DE CREACIÓN, CANT BOOKING,
    F. ÚLT. BOOKING (SUCURSAL no se usa, no está en asignacion_de_clientes)."""
    usuario = (os.environ.get("CARGOLINK_USUARIO") or "").strip()
    password = (os.environ.get("CARGOLINK_PASSWORD") or "").strip()
    if not usuario or not password:
        raise RuntimeError("Faltan las variables de entorno CARGOLINK_USUARIO / CARGOLINK_PASSWORD.")

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    sesion = requests.Session()
    r_login = sesion.post(CARGOLINK_LOGIN_URL, data={"usuario": usuario, "password": password}, headers=headers, timeout=60)
    if r_login.status_code != 200 or '"activo"' not in r_login.text:
        raise RuntimeError("No se pudo iniciar sesión en CargoLink.")

    res = sesion.get(
        CARGOLINK_REPORTE_CLIENTES_URL,
        params={"cliente": "", "t_cliente": "", "vendedor": "", "sucursal": "", "desarrollador": "", "industria": ""},
        headers=headers,
        timeout=120,
    )
    if res.status_code != 200 or len(res.content) == 0:
        raise RuntimeError("Error al descargar el Reporte de Clientes desde CargoLink.")

    wb = openpyxl.load_workbook(io.BytesIO(res.content), data_only=True)
    ws = wb.active

    def parse_fecha(v):
        if not v:
            return None
        if isinstance(v, datetime):
            return v
        texto = str(v).strip()
        for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(texto, formato)
            except ValueError:
                continue
        return None

    header = None
    clientes = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "FOLIO":
                header = [str(h).strip() if h is not None else "" for h in row]
            continue
        if not row or row[0] is None:
            continue

        dato = dict(zip(header, row))
        try:
            folio = int(dato.get("FOLIO"))
        except (TypeError, ValueError):
            continue

        clientes.append({
            "folio": folio,
            "razon_social": (dato.get("RAZÓN SOCIAL") or "").strip() or "(sin nombre)",
            "vendedor": (dato.get("VENDEDOR") or "").strip() or None,
            "desarrollador": (dato.get("DESARROLLADOR") or "").strip() or None,
            "tipo_cliente": (dato.get("TIPO DE CLIENTE") or "").strip() or None,
            # fecha_creacion es NOT NULL en la tabla; si CargoLink no trae
            # una fecha parseable (no debería pasar, pero por seguridad),
            # se usa el momento de la descarga en vez de mandar NULL.
            "fecha_creacion": parse_fecha(dato.get("FECHA DE CREACIÓN")) or datetime.now(TZ_LOCAL),
            "cant_booking": int(dato.get("CANT BOOKING") or 0),
            "fecha_ultimo_booking": parse_fecha(dato.get("F. ÚLT. BOOKING")),
        })

    return clientes


def _conectar_liq_vendedor_cargolink():
    """Login a CargoLink + token de sesión del módulo Liquidación de
    vendedores (m=150). Compartido por listar_folios_liquidacion_cargolink
    y descargar_liquidacion_vendedor_cargolink."""
    usuario = (os.environ.get("CARGOLINK_USUARIO") or "").strip()
    password = (os.environ.get("CARGOLINK_PASSWORD") or "").strip()
    if not usuario or not password:
        raise RuntimeError("Faltan las variables de entorno CARGOLINK_USUARIO / CARGOLINK_PASSWORD.")

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    sesion = requests.Session()
    r_login = sesion.post(CARGOLINK_LOGIN_URL, data={"usuario": usuario, "password": password}, headers=headers, timeout=60)
    if r_login.status_code != 200 or '"activo"' not in r_login.text:
        raise RuntimeError("No se pudo iniciar sesión en CargoLink.")

    r_pagina = sesion.get(f"{CARGOLINK_LIQ_VENDEDOR_URL}?m=150", headers=headers, timeout=60)
    match_token = re.search(r"token=([a-f0-9]{32}\d*)", r_pagina.text)
    if not match_token:
        raise RuntimeError("No se pudo obtener el token de sesión de Liquidación de Vendedores.")
    return sesion, headers, match_token.group(1)


def listar_folios_liquidacion_cargolink():
    """Lista los folios (folio, descripción) disponibles en Egresos →
    Liquidación de vendedores, para el selector de carga."""
    sesion, headers, token = _conectar_liq_vendedor_cargolink()
    r_lista = sesion.post(
        f"https://fwd.cargolink.mx/ws/cliente_conexion.php?token={token}&cat=apiLiquidacion&fn=consultaLiqVendedor",
        json={},
        headers={"Content-Type": "application/json"},
        timeout=60,
    )
    if r_lista.status_code != 200:
        raise RuntimeError("Error al consultar la lista de liquidaciones en CargoLink.")
    folios = [
        {"folio": int(f["folio_int"]), "descripcion": f.get("nombre") or ""}
        for f in r_lista.json().get("values", [])
        if int(f["folio_int"]) >= 40
    ]
    folios.sort(key=lambda f: f["folio"], reverse=True)
    return folios


def descargar_liquidacion_vendedor_cargolink(folio):
    """Se conecta a CargoLink, ubica el folio dado en Egresos → Liquidación
    de vendedores (m=150) y descarga su detalle línea por línea (uno por
    booking). No toca el disco. Regresa {folio, descripcion, detalle}."""
    sesion, headers, token = _conectar_liq_vendedor_cargolink()

    r_lista = sesion.post(
        f"https://fwd.cargolink.mx/ws/cliente_conexion.php?token={token}&cat=apiLiquidacion&fn=consultaLiqVendedor",
        json={},
        headers={"Content-Type": "application/json"},
        timeout=60,
    )
    if r_lista.status_code != 200:
        raise RuntimeError("Error al consultar la lista de liquidaciones en CargoLink.")
    fila_folio = next(
        (f for f in r_lista.json().get("values", []) if str(f.get("folio_int")) == str(folio)), None
    )
    if fila_folio is None:
        raise RuntimeError(f"No se encontró el folio {folio} en Liquidación de vendedores.")

    id_liq = fila_folio["id_liq_vendedor"]
    descripcion = fila_folio.get("nombre") or ""

    r_excel = sesion.get(
        f"{CARGOLINK_LIQ_VENDEDOR_URL}excel_detalle.php?token={token}&id_liq={id_liq}",
        headers=headers,
        timeout=60,
    )
    if r_excel.status_code != 200 or len(r_excel.content) == 0:
        raise RuntimeError("Error al descargar el detalle de la liquidación desde CargoLink.")

    def num(v):
        v = (v or "0").replace(",", "").replace("$", "").strip()
        try:
            return float(v)
        except ValueError:
            return 0.0

    soup = BeautifulSoup(r_excel.content.decode("utf-8-sig", errors="replace"), "html.parser")
    tabla = soup.find("table")
    filas_html = tabla.find_all("tr") if tabla else []

    detalle = []
    header_visto = False
    for tr in filas_html:
        celdas = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
        if not celdas:
            continue
        if celdas[0] == "Booking":
            header_visto = True
            continue
        if not header_visto or celdas[0] == "" or len(celdas) < 9:
            continue  # fila de totales al final, o de encabezado/folio arriba de la tabla
        profit = num(celdas[2])
        detalle.append({
            "booking": celdas[0],
            "folio_cobro": celdas[1],
            "profit": profit,
            "vendedor": celdas[3],
            "pct_vendedor": num(celdas[4]),
            "total_vendedor": num(celdas[5]),
            "desarrollador": celdas[6],
            # La comisión del desarrollador ya no viene de CargoLink: se
            # calcula aquí como profit * 0.02.
            "pct_desarrollador": 2.0 if celdas[6] else None,
            "total_desarrollador": round(profit * 0.02, 2) if celdas[6] else 0.0,
        })

    return {"folio": int(folio), "descripcion": descripcion, "detalle": detalle}


def descargar_concentrado_cobros_cargolink(fecha_ini, fecha_fin):
    """Se conecta a CargoLink, consulta Ingresos → Reporte de Cobros (m=48)
    para el rango de fechas dado y descarga el Excel Concentrado. No toca
    el disco. Regresa una lista de dicts (uno por cobro)."""
    usuario = (os.environ.get("CARGOLINK_USUARIO") or "").strip()
    password = (os.environ.get("CARGOLINK_PASSWORD") or "").strip()
    if not usuario or not password:
        raise RuntimeError("Faltan las variables de entorno CARGOLINK_USUARIO / CARGOLINK_PASSWORD.")

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
    sesion = requests.Session()
    r_login = sesion.post(CARGOLINK_LOGIN_URL, data={"usuario": usuario, "password": password}, headers=headers, timeout=60)
    if r_login.status_code != 200 or '"activo"' not in r_login.text:
        raise RuntimeError("No se pudo iniciar sesión en CargoLink.")

    r_pagina = sesion.get(
        "https://fwd.cargolink.mx/templates/reporteCobros/index20.php?m=48", headers=headers, timeout=60
    )
    match_token = re.search(r"token=([a-f0-9]{32}\d*)", r_pagina.text)
    if not match_token:
        raise RuntimeError("No se pudo obtener el token de sesión de Reporte de Cobros.")
    token = match_token.group(1)

    r_consulta = sesion.post(
        f"https://fwd.cargolink.mx/ws/cliente_conexion.php?token={token}&cat=api&fn=consultaCobrosClientes&limit=0",
        json={"filtros": {}, "filtros2": {"fechaini": fecha_ini, "fechafin": fecha_fin}},
        headers={"Content-Type": "application/json"},
        timeout=60,
    )
    if r_consulta.status_code != 200:
        raise RuntimeError("Error al consultar Reporte de Cobros en CargoLink.")
    where = r_consulta.json().get("where")
    if where is None:
        raise RuntimeError("CargoLink no regresó resultados para ese rango de fechas.")

    r_excel = sesion.get(
        "https://fwd.cargolink.mx/templates/reporteCobros/excelConcentrado.php",
        params={"filtros": where, "token": token},
        headers=headers,
        timeout=120,
    )
    if r_excel.status_code != 200 or len(r_excel.content) == 0:
        raise RuntimeError("Error al descargar el concentrado de cobros desde CargoLink.")

    def num(v):
        v = (v or "0").replace(",", "").replace("$", "").strip()
        try:
            return float(v)
        except ValueError:
            return 0.0

    def fecha(v):
        v = (v or "").strip()
        try:
            return datetime.strptime(v, "%d-%m-%Y").date()
        except ValueError:
            return None

    soup = BeautifulSoup(r_excel.content.decode("utf-8-sig", errors="replace"), "html.parser")
    tabla = soup.find("table")
    filas_html = tabla.find_all("tr") if tabla else []

    cobros = []
    header_visto = False
    for tr in filas_html:
        celdas = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
        if not celdas:
            continue
        if celdas[0] == "Folio cobro":
            header_visto = True
            continue
        if not header_visto or len(celdas) < 18 or not celdas[2]:
            continue  # fila de título arriba de la tabla, o sin Referencia (nota/anotación, no un cobro real)
        cobros.append({
            "folio_cobro": celdas[0],
            "folio_factura": celdas[1],
            "referencia": celdas[2],
            "uuid": celdas[3],
            "tipo_docto": celdas[4],
            "tipo_referencia": celdas[5],
            "cliente": celdas[6],
            "fecha_factura": fecha(celdas[7]),
            "fecha_cobro": fecha(celdas[8]),
            "dias_diferencia": int(celdas[9]) if celdas[9].isdigit() else None,
            "fecha_timbre": celdas[10],
            "banco": celdas[11],
            "moneda": celdas[12],
            "subtotal": num(celdas[13]),
            "iva": num(celdas[14]),
            "descuento": num(celdas[15]),
            "retencion": num(celdas[16]),
            "total": num(celdas[17]),
        })

    return cobros


def construir_datos_dashboard(plazas_permitidas=None):
    db = get_db()
    meta = db.execute(
        "SELECT fecha_inicio, fecha_fin, generado_en FROM reporte_generaciones ORDER BY generado_en DESC LIMIT 1"
    ).fetchone()
    if meta is None:
        db.close()
        return None

    catalogo_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza, ocultar_detalle FROM catalogo_vendedores"):
        catalogo_vendedor[normalizar(r["vendedor"])] = {
            "plaza": r["plaza"],
            "nombre": r["vendedor"],
            "ocultar_detalle": r["ocultar_detalle"],
        }

    catalogo_desarrollador = {}
    for r in db.execute("SELECT desarrollador, plaza FROM catalogo_desarrolladores"):
        catalogo_desarrollador[normalizar(r["desarrollador"])] = {"plaza": r["plaza"], "nombre": r["desarrollador"]}

    presupuesto_por_mes_vendedor = {}
    for r in db.execute("SELECT mes, vendedor, presupuesto FROM catalogo_presupuesto WHERE vendedor IS NOT NULL"):
        clave = (r["mes"], normalizar(r["vendedor"]))
        presupuesto_por_mes_vendedor[clave] = presupuesto_por_mes_vendedor.get(clave, 0.0) + float(r["presupuesto"])

    presupuesto_por_mes_desarrollador = {}
    for r in db.execute("SELECT mes, desarrollador, presupuesto FROM catalogo_presupuesto WHERE desarrollador IS NOT NULL"):
        clave = (r["mes"], normalizar(r["desarrollador"]))
        presupuesto_por_mes_desarrollador[clave] = presupuesto_por_mes_desarrollador.get(clave, 0.0) + float(r["presupuesto"])

    bookings = db.execute(
        "SELECT mes, vendedor, referencia, fecha, ejecutivo, venta_por, cliente_servicio, venta, profit, margen "
        "FROM reporte_bookings"
    ).fetchall()
    db.close()

    agregados = {}
    agregados_desarrollador = {}
    detalle = []
    for r in bookings:
        vkey = normalizar(r["vendedor"])
        cat = catalogo_vendedor.get(vkey)
        plaza_vendedor = cat["plaza"] if cat else "#N/D"
        vendedor_permitido = plazas_permitidas is None or plaza_vendedor in plazas_permitidas

        dkey = normalizar(r["ejecutivo"]) if r["ejecutivo"] else None

        if vendedor_permitido:
            clave = (r["mes"], vkey)
            agg = agregados.setdefault(clave, {"cant_book": 0, "venta": 0.0, "profit": 0.0})
            agg["cant_book"] += 1
            agg["venta"] += float(r["venta"])
            agg["profit"] += float(r["profit"])

        if dkey and vendedor_permitido:
            agg_d = agregados_desarrollador.setdefault((r["mes"], dkey), {"cant_book": 0, "venta": 0.0, "profit": 0.0})
            agg_d["cant_book"] += 1
            agg_d["venta"] += float(r["venta"])
            agg_d["profit"] += float(r["profit"])

        nombre_canonico = cat["nombre"] if cat else r["vendedor"]
        # Los totales (agregados/filas de arriba) siempre incluyen a todos los
        # vendedores; solo el detalle a nivel booking se omite para quien
        # tenga marcado "ocultar_detalle" en su catálogo. El criterio
        # mandatorio de plaza es el vendedor: que el desarrollador/customer
        # esté catalogado en una plaza permitida NO basta para mostrar la
        # venta si el vendedor que la vendió es de otra plaza.
        if vendedor_permitido and not (cat and cat["ocultar_detalle"]):
            detalle.append({
                "mes": r["mes"],
                "vendedor": nombre_canonico,
                "referencia": r["referencia"] or "",
                "fecha": r["fecha"].strftime("%Y-%m-%d %H:%M") if r["fecha"] else "",
                "ejecutivo": r["ejecutivo"] or "",
                "venta_por": r["venta_por"] or "",
                "cliente_servicio": r["cliente_servicio"] or "",
                "venta": round(float(r["venta"]), 2),
                "profit": round(float(r["profit"]), 2),
                "margen": round(float(r["margen"]), 4),
            })

    for clave in presupuesto_por_mes_vendedor:
        _, vkey = clave
        cat = catalogo_vendedor.get(vkey)
        plaza = cat["plaza"] if cat else "#N/D"
        if plazas_permitidas is None or plaza in plazas_permitidas:
            agregados.setdefault(clave, {"cant_book": 0, "venta": 0.0, "profit": 0.0})
    for clave in presupuesto_por_mes_desarrollador:
        _, dkey = clave
        cat = catalogo_desarrollador.get(dkey)
        plaza = cat["plaza"] if cat else "#N/D"
        if plazas_permitidas is None or plaza in plazas_permitidas:
            agregados_desarrollador.setdefault(clave, {"cant_book": 0, "venta": 0.0, "profit": 0.0})

    filas = []
    for (mes, vkey), agg in agregados.items():
        cat = catalogo_vendedor.get(vkey)
        plaza = cat["plaza"] if cat else "#N/D"
        nombre = cat["nombre"] if cat else vkey
        ppto = presupuesto_por_mes_vendedor.get((mes, vkey), 0.0)
        filas.append({
            "mes": mes,
            "vendedor": nombre,
            "plaza": plaza,
            "cant_book": agg["cant_book"],
            "venta": round(agg["venta"], 2),
            "profit": round(agg["profit"], 2),
            "ppto": round(ppto, 2),
        })

    filas_desarrolladores = []
    for (mes, dkey), agg in agregados_desarrollador.items():
        cat = catalogo_desarrollador.get(dkey)
        plaza = cat["plaza"] if cat else "#N/D"
        nombre = cat["nombre"] if cat else dkey
        ppto = presupuesto_por_mes_desarrollador.get((mes, dkey), 0.0)
        filas_desarrolladores.append({
            "mes": mes,
            "desarrollador": nombre,
            "plaza": plaza,
            "cant_book": agg["cant_book"],
            "venta": round(agg["venta"], 2),
            "profit": round(agg["profit"], 2),
            "ppto": round(ppto, 2),
        })

    todos_los_meses = sorted(set(f["mes"] for f in filas) | set(f["mes"] for f in filas_desarrolladores))
    mes_actual = datetime.now(TZ_LOCAL).strftime("%Y-%m")
    if mes_actual not in todos_los_meses:
        todos_los_meses.append(mes_actual)
        todos_los_meses.sort()

    return {
        "archivo": f"Reporte {meta['fecha_inicio']} al {meta['fecha_fin']}",
        "generado_en": meta["generado_en"].astimezone(TZ_LOCAL).strftime("%d/%m/%Y %H:%M"),
        "filas": filas,
        "filas_desarrolladores": filas_desarrolladores,
        "detalle": detalle,
        "meses": todos_los_meses,
        "mes_actual": mes_actual,
        "plazas": sorted(set(f["plaza"] for f in filas)),
        "vendedores": sorted(set(f["vendedor"] for f in filas)),
        "desarrolladores": sorted(set(f["desarrollador"] for f in filas_desarrolladores)),
    }


try:
    init_db()
except Exception as e:
    print(f"Aviso: no se pudo inicializar la base de datos al arrancar ({e}).")

app = Flask(__name__)
app.secret_key = get_secret_key()

# EN_VERCEL: solo exigimos cookies "Secure" en producción (HTTPS). En local
# (HTTP en 127.0.0.1) exigirlo rompería el login, porque el navegador nunca
# manda una cookie Secure sobre una conexión sin TLS.
EN_VERCEL = bool(os.environ.get("VERCEL"))
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=EN_VERCEL,
    # Fuerza un re-login cada 12h: sin esto, una sesión podía quedar abierta
    # indefinidamente (nunca vencía sola) y los permisos que trae la sesión
    # (es_admin, puede_ver_crm, ...) nunca se refrescaban aunque un admin
    # se los quitara a alguien en Catálogos → Accesos.
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    # Sin esto, Flask-WTF vence el token CSRF a la hora (su default) aunque
    # la sesión siga viva 12h — rompería formularios largos (como Nueva
    # cotización) si alguien tarda más de una hora en llenarlos.
    WTF_CSRF_TIME_LIMIT=None,
)

csrf = CSRFProtect(app)


@app.after_request
def agregar_cabeceras_seguridad(resp):
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    resp.headers["Content-Security-Policy"] = "frame-ancestors 'none'"
    return resp


IMPORTACIONES_URL = (os.environ.get("IMPORTACIONES_URL") or "http://localhost:3001").strip()
SSO_SHARED_SECRET = (os.environ.get("SSO_SHARED_SECRET") or "").strip()
SSO_TOKEN_TTL_SEGUNDOS = 60


@app.context_processor
def inject_importaciones_url():
    return {"importaciones_url": IMPORTACIONES_URL}


def formatear_duracion(delta):
    """Formatea un timedelta como "2d 3h", "5h 20min" o "12min" — lo que
    tardó Pricing en contestar una solicitud, en la unidad más legible."""
    total_min = int(delta.total_seconds() // 60)
    if total_min < 0:
        total_min = 0
    dias, resto_min = divmod(total_min, 24 * 60)
    horas, minutos = divmod(resto_min, 60)
    if dias:
        return f"{dias}d {horas}h"
    if horas:
        return f"{horas}h {minutos}min"
    return f"{minutos}min"


@app.template_filter("hora_mx")
def hora_mx(valor, formato="%Y-%m-%d %H:%M"):
    """Formatea un datetime (con tz, típicamente UTC desde Postgres) en la
    hora de Ciudad de México en vez de la hora cruda de la base de datos."""
    if not valor:
        return ""
    return valor.astimezone(TZ_LOCAL).strftime(formato)


@app.template_filter("nl2br")
def nl2br(texto):
    """Convierte saltos de línea en <br>, escapando el resto del texto.
    xhtml2pdf respeta <br> de forma confiable; CSS white-space:pre-line no."""
    if not texto:
        return ""
    return Markup("<br>").join(escape(linea) for linea in str(texto).splitlines())


def registrar_ingreso():
    """Guarda un renglón en registro_ingresos una vez por día calendario
    por sesión (no una vez por cada página que visite, pero tampoco solo
    una vez por sesión): las sesiones no expiran solas, así que alguien
    que no cierra el navegador puede seguir usando la app días después
    sin volver a loguearse — si solo registráramos una vez por sesión,
    esa actividad real nunca aparecería en Actividad de Usuarios. Se
    engancha en los decoradores en vez de en /login para que funcione
    sin importar qué mecanismo de autenticación esté activo. Nunca debe
    tumbar la vista por un problema de logging."""
    hoy = datetime.now(TZ_LOCAL).date().isoformat()
    if session.get("ingreso_registrado") == hoy:
        return
    session["ingreso_registrado"] = hoy
    try:
        db = get_db()
        db.execute(
            "INSERT INTO registro_ingresos (usuario_id, usuario) VALUES (%s, %s)",
            (session.get("usuario_id"), session.get("usuario") or "?"),
        )
        db.commit()
        db.close()
    except Exception as e:
        print(f"Aviso: no se pudo registrar el ingreso ({e}).")


def sesion_activa():
    """True si hay una sesión logueada y completa (incluye usuario_id).
    Antes solo se checaba "logged_in": como las sesiones no expiran solas,
    una sesión vieja de antes de que existiera usuario_id podía seguir
    "logueada" indefinidamente pero sin usuario_id — eso hacía que, por
    ejemplo, las cotizaciones nuevas que creara esa persona se guardaran
    con creado_por_user_id en NULL (nadie identificable como su autor).
    Si falta, se limpia la sesión para forzar un re-login que la complete."""
    if session.get("logged_in") and session.get("usuario_id"):
        return True
    session.clear()
    return False


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        if not session.get("es_admin"):
            flash("Esa sección es solo para administradores.")
            return redirect(url_for("dashboard_plazas_vendedores"))
        return view(*args, **kwargs)

    return wrapped


def reportes_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        if not usuario_puede_ver_reportes():
            flash("No tienes permiso para ver Reportes.")
            return redirect(url_for(primera_pagina_permitida()))
        return view(*args, **kwargs)

    return wrapped


def catalogos_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        if not usuario_puede_ver_catalogos():
            flash("No tienes permiso para ver Catálogos.")
            return redirect(url_for(primera_pagina_permitida()))
        return view(*args, **kwargs)

    return wrapped


def crm_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        if not usuario_puede_ver_crm():
            flash("No tienes permiso para ver CRM.")
            return redirect(url_for(primera_pagina_permitida()))
        return view(*args, **kwargs)

    return wrapped


def pricing_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not sesion_activa():
            return redirect(url_for("login"))
        registrar_ingreso()
        if not usuario_puede_pricing():
            flash("No tienes permiso para ver Pricing.")
            return redirect(url_for(primera_pagina_permitida()))
        return view(*args, **kwargs)

    return wrapped


def plazas_permitidas_usuario():
    """None = el usuario en sesión ve todas las plazas (sin restricción).
    Si no es None, es el set de plazas que puede ver. Los administradores
    siempre ven todo, igual que quien tenga marcado "Todas las plazas" en
    Catálogos → Visibilidad de Plazas. El resto ve solo las plazas que se le
    hayan asignado ahí (ninguna hasta que un admin le asigne alguna)."""
    if session.get("es_admin") or session.get("todas_las_plazas"):
        return None
    usuario_id = session.get("usuario_id")
    if not usuario_id:
        return set()
    db = get_db()
    filas = db.execute("SELECT plaza FROM app_user_plazas WHERE user_id = %s", (usuario_id,)).fetchall()
    db.close()
    return {f["plaza"] for f in filas}


def usuario_puede_ver_cotizacion(db, cliente_folio):
    """True si el usuario en sesión puede ver/editar/borrar/clonar una
    cotización con este cliente_folio, según sus plazas permitidas —
    mismo criterio que el listado de Cotizaciones (construir_cotizaciones_crm):
    las de prospecto (cliente_folio None) son visibles para cualquiera con
    acceso al CRM; las de cliente real heredan la plaza del vendedor
    asignado a ese cliente. El listado ya aplicaba esta regla al armar la
    tabla; esta función la repite para las rutas que operan sobre UNA
    cotización puntual por id (antes solo validaban que existiera)."""
    plazas_permitidas = plazas_permitidas_usuario()
    if plazas_permitidas is None or cliente_folio is None:
        return True
    fila = db.execute("""
        SELECT cv.plaza
        FROM asignacion_de_clientes ac
        LEFT JOIN catalogo_vendedores cv ON upper(trim(cv.vendedor)) = upper(trim(ac.vendedor))
        WHERE ac.folio = %s
    """, (cliente_folio,)).fetchone()
    plaza = (fila["plaza"] if fila else None) or "#N/D"
    return plaza in plazas_permitidas


def cotizacion_visible_para_usuario(db, cotizacion_id):
    """True si la cotización existe y el usuario en sesión tiene permiso de
    plaza para operar sobre ella. Úsala al inicio de cualquier ruta que
    reciba un cotizacion_id, antes de mostrar/editar/borrar nada."""
    fila = db.execute("SELECT cliente_folio FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)).fetchone()
    if fila is None:
        return False
    return usuario_puede_ver_cotizacion(db, fila["cliente_folio"])


def usuario_puede_ver_contacto(db, contacto_id):
    """True si el usuario en sesión puede ver/editar/borrar este contacto,
    según sus plazas permitidas — mismo criterio que el listado de
    Contactos (construir_contactos_crm): uno sin ningún cliente asociado es
    visible para cualquiera con acceso al CRM; si tiene clientes, basta con
    que UNO caiga en una plaza permitida."""
    plazas_permitidas = plazas_permitidas_usuario()
    if plazas_permitidas is None:
        return True
    filas = db.execute("""
        SELECT DISTINCT cv.plaza
        FROM crm_contacto_clientes cc
        JOIN asignacion_de_clientes ac ON ac.folio = cc.cliente_folio
        LEFT JOIN catalogo_vendedores cv ON upper(trim(cv.vendedor)) = upper(trim(ac.vendedor))
        WHERE cc.contacto_id = %s
    """, (contacto_id,)).fetchall()
    plazas_contacto = {(f["plaza"] or "#N/D") for f in filas}
    if not plazas_contacto:
        return True
    return bool(plazas_contacto & plazas_permitidas)


def usuario_puede_exportar():
    """True = el usuario en sesión puede usar los botones "Exportar". Los
    administradores siempre pueden; para el resto se usa el permiso
    guardado en sesión al hacer login (app_user_permissions.puede_exportar)."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_exportar"))


def usuario_puede_actualizar():
    """True = el usuario en sesión puede usar el botón "Actualizar" para
    regenerar el reporte desde CargoLink (acción sensible: reemplaza todos
    los bookings). Los administradores siempre pueden; para el resto se usa
    el permiso guardado en sesión al hacer login
    (app_user_permissions.puede_actualizar), otorgado desde Catálogos →
    Permiso de Actualizar."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_actualizar"))


def usuario_puede_comisiones():
    """True = el usuario en sesión puede ver la pestaña y la pantalla de
    Comisiones. Los administradores siempre pueden; para el resto se usa
    el permiso guardado en sesión al hacer login
    (app_user_permissions.puede_comisiones), otorgado desde Catálogos →
    Permiso de Comisiones."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_comisiones"))


def usuario_puede_pricing():
    """True = el usuario en sesión puede ver la pestaña Pricing (bandeja de
    solicitudes Marítimo/Aéreo de todas las cotizaciones, para contestarlas).
    Los administradores siempre pueden; para el resto se usa el permiso
    guardado en sesión al hacer login (app_user_permissions.puede_pricing),
    otorgado desde Catálogos → Permisos de Usuario."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_pricing"))


def usuario_puede_ver_ventas():
    """True = el usuario en sesión puede ver la pestaña Información de
    Ventas. Los administradores siempre pueden; para el resto se usa el
    permiso guardado en sesión al hacer login
    (app_user_permissions.puede_ver_ventas), otorgado desde Catálogos →
    Permisos de Usuario."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_ver_ventas"))


def usuario_puede_ver_reportes():
    """True = el usuario en sesión puede ver la pestaña Reportes (y sus
    subpáginas: por vendedor, clientes asignados, clientes mensual). Los
    administradores siempre pueden; para el resto se usa el permiso
    guardado en sesión al hacer login (app_user_permissions.puede_ver_reportes),
    otorgado desde Catálogos → Permisos de Usuario."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_ver_reportes"))


def usuario_puede_ver_catalogos():
    """True = el usuario en sesión puede ver la pantalla principal de
    Catálogos. Los administradores siempre pueden; para el resto se usa el
    permiso guardado en sesión al hacer login
    (app_user_permissions.puede_ver_catalogos), otorgado desde Catálogos →
    Permisos de Usuario. Las herramientas individuales dentro de Catálogos
    (vendedores, presupuesto, permisos de usuario, etc.) siguen siendo
    exclusivas de administradores."""
    if session.get("es_admin"):
        return True
    return bool(session.get("puede_ver_catalogos"))


def plazas_con_crm_habilitado():
    """Plazas a las que Catálogos → CRM por Plaza les dio acceso a CRM
    completo (todo usuario asignado a esa plaza en Visibilidad de Plazas
    puede entrar, sin necesidad de un permiso individual)."""
    db = get_db()
    filas = db.execute("SELECT plaza FROM crm_plazas_habilitadas").fetchall()
    db.close()
    return {f["plaza"] for f in filas}


def plazas_asignadas_usuario_sesion():
    """Las plazas que Visibilidad de Plazas le asignó directamente al
    usuario en sesión (sin importar la bandera 'todas_las_plazas'), para
    decidir accesos que se otorgan por plaza, como CRM."""
    usuario_id = session.get("usuario_id")
    if not usuario_id:
        return set()
    db = get_db()
    filas = db.execute("SELECT plaza FROM app_user_plazas WHERE user_id = %s", (usuario_id,)).fetchall()
    db.close()
    return {f["plaza"] for f in filas}


def usuario_puede_ver_crm():
    """True = el usuario en sesión puede ver la pestaña CRM. Los
    administradores siempre pueden. Para el resto, hay dos caminos que se
    combinan (basta con uno): el permiso individual guardado en sesión al
    hacer login (app_user_permissions.puede_ver_crm, otorgado desde
    Catálogos → Permisos de Usuario), o que alguna de sus plazas asignadas
    (Catálogos → Visibilidad de Plazas) tenga CRM habilitado por plaza
    (Catálogos → CRM por Plaza)."""
    if session.get("es_admin"):
        return True
    if session.get("puede_ver_crm"):
        return True
    plazas_habilitadas = plazas_con_crm_habilitado()
    if not plazas_habilitadas:
        return False
    return bool(plazas_asignadas_usuario_sesion() & plazas_habilitadas)


def primera_pagina_permitida():
    """Nombre de la ruta a la que mandar al usuario en sesión: la primera
    sección para la que sí tiene permiso de ver, en el mismo orden del menú."""
    if session.get("es_admin"):
        return "dashboard"
    if usuario_puede_ver_ventas():
        return "dashboard_plazas_vendedores"
    if usuario_puede_ver_reportes():
        return "reportes_graficas"
    if usuario_puede_comisiones():
        return "comisiones"
    if usuario_puede_ver_crm():
        return "crm"
    return "login"


@app.context_processor
def inject_permisos():
    return {
        "puede_exportar": usuario_puede_exportar(),
        "puede_actualizar": usuario_puede_actualizar(),
        "puede_comisiones": usuario_puede_comisiones(),
        "puede_ver_ventas": usuario_puede_ver_ventas(),
        "puede_ver_reportes": usuario_puede_ver_reportes(),
        "puede_ver_catalogos": usuario_puede_ver_catalogos(),
        "puede_ver_crm": usuario_puede_ver_crm(),
        "puede_pricing": usuario_puede_pricing(),
        "reporte_ventas_url": url_for(primera_pagina_permitida()) if session.get("logged_in") else None,
    }


def autenticar_contra_catalogo_accesos(email, password):
    """Valida el correo/contraseña contra el catálogo de accesos de
    Seguimiento de Importaciones (auth.users + app_user_permissions, mismo
    proyecto Supabase). Regresa la fila del usuario si es válido y la cuenta
    no está deshabilitada, o None si no."""
    db = get_db()
    fila = db.execute(
        """
        SELECT
            u.id, u.email,
            (u.encrypted_password IS NOT NULL
                AND extensions.crypt(%(password)s, u.encrypted_password) = u.encrypted_password) AS password_ok,
            (u.banned_until IS NOT NULL AND u.banned_until > now()) AS baneado,
            coalesce(p.es_admin, false) AS es_admin,
            coalesce(p.puede_exportar, false) AS puede_exportar,
            coalesce(p.puede_actualizar, false) AS puede_actualizar,
            coalesce(p.puede_comisiones, false) AS puede_comisiones,
            coalesce(p.puede_ver_ventas, true) AS puede_ver_ventas,
            coalesce(p.puede_ver_reportes, true) AS puede_ver_reportes,
            coalesce(p.puede_ver_catalogos, false) AS puede_ver_catalogos,
            coalesce(p.puede_ver_crm, false) AS puede_ver_crm,
            coalesce(p.puede_pricing, false) AS puede_pricing,
            coalesce(p.todas_las_plazas, false) AS todas_las_plazas,
            coalesce(p.puede_borrar, false) AS puede_borrar,
            coalesce(p.puede_operativos, false) AS puede_operativos,
            coalesce(p.es_master, false) AS es_master
        FROM auth.users u
        LEFT JOIN public.app_user_permissions p ON p.user_id = u.id
        WHERE lower(u.email) = lower(%(email)s)
        """,
        {"email": email, "password": password},
    ).fetchone()
    db.close()
    if not fila or not fila["password_ok"] or fila["baneado"]:
        return None
    return fila


INTENTOS_LOGIN_LIMITE = 5
INTENTOS_LOGIN_VENTANA_MIN = 15


def intentos_login_bloqueado(db, email):
    """True si este correo ya acumuló demasiados intentos fallidos en la
    ventana reciente. Se guarda en Postgres (no en memoria del proceso)
    porque en Vercel cada invocación puede caer en una instancia distinta —
    un contador en memoria no serviría para frenar fuerza bruta real."""
    fila = db.execute(
        "SELECT count(*) AS n FROM intentos_login "
        "WHERE lower(email) = lower(%s) AND exitoso = false "
        "AND creado_en > now() - make_interval(mins => %s)",
        (email, INTENTOS_LOGIN_VENTANA_MIN),
    ).fetchone()
    return fila["n"] >= INTENTOS_LOGIN_LIMITE


def registrar_intento_login(db, email, exitoso):
    db.execute("INSERT INTO intentos_login (email, exitoso) VALUES (%s, %s)", (email, exitoso))
    db.commit()


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        clave = request.form.get("password", "")

        db = get_db()
        if email and intentos_login_bloqueado(db, email):
            db.close()
            flash(f"Demasiados intentos fallidos. Espera {INTENTOS_LOGIN_VENTANA_MIN} minutos e inténtalo de nuevo.")
            return render_template("login.html")

        fila = autenticar_contra_catalogo_accesos(email, clave)
        if email:
            registrar_intento_login(db, email, fila is not None)
        db.close()

        if fila:
            session.permanent = True
            session["logged_in"] = True
            session["usuario"] = fila["email"]
            session["usuario_id"] = str(fila["id"])
            session["es_admin"] = bool(fila["es_admin"])
            session["puede_exportar"] = bool(fila["puede_exportar"])
            session["puede_actualizar"] = bool(fila["puede_actualizar"])
            session["puede_comisiones"] = bool(fila["puede_comisiones"])
            session["puede_ver_ventas"] = bool(fila["puede_ver_ventas"])
            session["puede_ver_reportes"] = bool(fila["puede_ver_reportes"])
            session["puede_ver_catalogos"] = bool(fila["puede_ver_catalogos"])
            session["puede_ver_crm"] = bool(fila["puede_ver_crm"])
            session["puede_pricing"] = bool(fila["puede_pricing"])
            session["todas_las_plazas"] = bool(fila["todas_las_plazas"])
            destino = primera_pagina_permitida()
            return redirect(url_for(destino))
        flash("Correo o contraseña incorrectos.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def verificar_token_sso(token):
    """Valida un token de acceso único emitido por Seguimiento de
    Importaciones para no pedir login de nuevo (mismo catálogo de accesos):
    HMAC-SHA256 sobre {email, iat} con un secreto compartido
    (SSO_SHARED_SECRET) y vigencia corta. Regresa el email si es válido."""
    if not SSO_SHARED_SECRET or not token or "." not in token:
        return None
    payload_b64, firma = token.rsplit(".", 1)
    esperada = hmac.new(
        SSO_SHARED_SECRET.encode(), payload_b64.encode(), hashlib.sha256
    ).hexdigest()
    if not hmac.compare_digest(firma, esperada):
        return None
    try:
        relleno = "=" * (-len(payload_b64) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64 + relleno))
        email = payload["email"]
        iat = float(payload["iat"])
    except (ValueError, KeyError, TypeError):
        return None
    if not (-5 <= time.time() - iat <= SSO_TOKEN_TTL_SEGUNDOS):
        return None
    return email


def _siguiente_sso_valido(next_path):
    """Solo permite redirigir dentro de esta misma app tras el SSO — una
    ruta relativa, nunca una URL completa (evita usar /sso como open-redirect)."""
    if not next_path or not next_path.startswith("/") or next_path.startswith("//"):
        return None
    return next_path


@app.route("/sso")
def sso():
    email = verificar_token_sso(request.args.get("token", ""))
    if not email:
        flash("El enlace de acceso expiró o no es válido. Inicia sesión normalmente.")
        return redirect(url_for("login"))

    db = get_db()
    fila = db.execute(
        """
        SELECT
            u.id, u.email,
            (u.banned_until IS NOT NULL AND u.banned_until > now()) AS baneado,
            coalesce(p.es_admin, false) AS es_admin,
            coalesce(p.puede_exportar, false) AS puede_exportar,
            coalesce(p.puede_actualizar, false) AS puede_actualizar,
            coalesce(p.puede_comisiones, false) AS puede_comisiones,
            coalesce(p.puede_ver_ventas, true) AS puede_ver_ventas,
            coalesce(p.puede_ver_reportes, true) AS puede_ver_reportes,
            coalesce(p.puede_ver_catalogos, false) AS puede_ver_catalogos,
            coalesce(p.puede_ver_crm, false) AS puede_ver_crm,
            coalesce(p.puede_pricing, false) AS puede_pricing,
            coalesce(p.todas_las_plazas, false) AS todas_las_plazas
        FROM auth.users u
        LEFT JOIN public.app_user_permissions p ON p.user_id = u.id
        WHERE lower(u.email) = lower(%(email)s)
        """,
        {"email": email},
    ).fetchone()
    db.close()

    if not fila or fila["baneado"]:
        flash("Tu cuenta no tiene acceso a esta app.")
        return redirect(url_for("login"))

    session.permanent = True
    session["logged_in"] = True
    session["usuario"] = fila["email"]
    session["usuario_id"] = str(fila["id"])
    session["es_admin"] = bool(fila["es_admin"])
    session["puede_exportar"] = bool(fila["puede_exportar"])
    session["puede_actualizar"] = bool(fila["puede_actualizar"])
    session["puede_comisiones"] = bool(fila["puede_comisiones"])
    session["puede_ver_ventas"] = bool(fila["puede_ver_ventas"])
    session["puede_ver_reportes"] = bool(fila["puede_ver_reportes"])
    session["puede_ver_catalogos"] = bool(fila["puede_ver_catalogos"])
    session["puede_ver_crm"] = bool(fila["puede_ver_crm"])
    session["puede_pricing"] = bool(fila["puede_pricing"])
    session["todas_las_plazas"] = bool(fila["todas_las_plazas"])

    destino = _siguiente_sso_valido(request.args.get("next")) or url_for(primera_pagina_permitida())
    return redirect(destino)


@app.route("/", methods=["GET"])
@admin_required
def dashboard():
    return render_template("dashboard.html", resultado=session.pop("resultado", None))


def ejecutar_generacion_reporte(fecha_inicio, fecha_fin):
    """Descarga bookings de CargoLink y reemplaza reporte_bookings solo para
    el rango [fecha_inicio, fecha_fin] — lo que ya existe fuera de ese rango
    (ej. años anteriores cargados una sola vez) no se toca.
    Regresa (ok: bool, mensaje: str, detalle: str|None)."""
    try:
        bookings = descargar_bookings_cargolink(fecha_inicio, fecha_fin)
    except Exception as e:
        return False, "Error al generar el reporte.", str(e)[:1500]

    if not bookings:
        return False, "CargoLink no devolvió bookings para ese rango de fechas.", None

    columnas = ["mes", "vendedor", "referencia", "fecha", "ejecutivo", "venta_por", "cliente_servicio", "venta", "profit", "margen"]
    TAMANO_LOTE = 500

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "DELETE FROM reporte_bookings WHERE fecha >= %s AND fecha < (%s::date + interval '1 day')",
        (fecha_inicio, fecha_fin),
    )
    for inicio in range(0, len(bookings), TAMANO_LOTE):
        lote = bookings[inicio:inicio + TAMANO_LOTE]
        placeholders = ", ".join(["(" + ", ".join(["%s"] * len(columnas)) + ")"] * len(lote))
        valores = [b[c] for b in lote for c in columnas]
        cur.execute(
            f"INSERT INTO reporte_bookings ({', '.join(columnas)}) VALUES {placeholders}",
            valores,
        )
    cur.execute(
        "INSERT INTO reporte_generaciones (fecha_inicio, fecha_fin) VALUES (%s, %s)",
        (fecha_inicio, fecha_fin),
    )
    db.commit()

    # Reporte de Clientes: además de los bookings, se actualiza
    # asignacion_de_clientes usando folio como llave — los folios que ya
    # existen se reemplazan con los datos frescos de CargoLink, y los que
    # no existían se agregan. Es un paso adicional: si falla, no se revierte
    # el reporte de bookings que ya se guardó arriba, solo se avisa.
    mensaje_clientes = ""
    try:
        clientes = descargar_reporte_clientes_cargolink()
        columnas_clientes = [
            "folio", "razon_social", "vendedor", "desarrollador",
            "tipo_cliente", "fecha_creacion", "cant_booking", "fecha_ultimo_booking",
        ]
        for inicio in range(0, len(clientes), TAMANO_LOTE):
            lote = clientes[inicio:inicio + TAMANO_LOTE]
            placeholders = ", ".join(["(" + ", ".join(["%s"] * len(columnas_clientes)) + ")"] * len(lote))
            valores = [c[col] for c in lote for col in columnas_clientes]
            cur.execute(
                f"""
                INSERT INTO asignacion_de_clientes ({', '.join(columnas_clientes)})
                VALUES {placeholders}
                ON CONFLICT (folio) DO UPDATE SET
                    razon_social = EXCLUDED.razon_social,
                    vendedor = EXCLUDED.vendedor,
                    desarrollador = EXCLUDED.desarrollador,
                    tipo_cliente = EXCLUDED.tipo_cliente,
                    fecha_creacion = EXCLUDED.fecha_creacion,
                    cant_booking = EXCLUDED.cant_booking,
                    fecha_ultimo_booking = EXCLUDED.fecha_ultimo_booking
                """,
                valores,
            )
        db.commit()
        mensaje_clientes = f" Reporte de clientes actualizado ({len(clientes)} clientes)."
    except Exception as e:
        db.rollback()
        mensaje_clientes = f" Aviso: el reporte de bookings sí se guardó, pero el reporte de clientes falló ({str(e)[:300]})."

    db.close()

    return True, f"Reporte generado correctamente ({len(bookings)} bookings).{mensaje_clientes}", None


@app.route("/generar", methods=["POST"])
@login_required
def generar():
    es_admin = bool(session.get("es_admin"))
    if not es_admin and not usuario_puede_actualizar():
        flash("No tienes permiso para actualizar el reporte.")
        return redirect(url_for("dashboard_plazas_vendedores"))

    # Las fechas de inicio/fin solo las puede elegir un administrador; para
    # el resto (botón "Actualizar" en Información de Ventas) siempre se usa
    # el rango por default, sin importar lo que venga en el formulario.
    fecha_inicio = request.form.get("fecha_inicio", "").strip() if es_admin else ""
    fecha_fin = request.form.get("fecha_fin", "").strip() if es_admin else ""

    if fecha_inicio and not DATE_RE.match(fecha_inicio):
        session["resultado"] = {"ok": False, "mensaje": "Fecha de inicio inválida."}
        return redirect(url_for("dashboard"))
    if fecha_fin and not DATE_RE.match(fecha_fin):
        session["resultado"] = {"ok": False, "mensaje": "Fecha de fin inválida."}
        return redirect(url_for("dashboard"))

    now = datetime.now(TZ_LOCAL)
    fecha_inicio = fecha_inicio or f"{now.year}-01-01"
    fecha_fin = fecha_fin or now.strftime("%Y-%m-%d")

    ok, mensaje, detalle = ejecutar_generacion_reporte(fecha_inicio, fecha_fin)
    if es_admin:
        session["resultado"] = {"ok": ok, "mensaje": mensaje, "detalle": detalle}
        return redirect(url_for("dashboard"))
    flash(mensaje)
    return redirect(url_for("dashboard_plazas_vendedores"))


@app.route("/cron/generar-reporte", methods=["GET", "POST"])
@csrf.exempt
def cron_generar_reporte():
    cron_secret = os.environ.get("CRON_SECRET", "").strip()
    auth = request.headers.get("Authorization", "")
    if not cron_secret or not hmac.compare_digest(auth, f"Bearer {cron_secret}"):
        return {"ok": False, "error": "no autorizado"}, 401

    now = datetime.now(TZ_LOCAL)
    fecha_inicio = f"{now.year}-01-01"
    fecha_fin = now.strftime("%Y-%m-%d")

    ok, mensaje, detalle = ejecutar_generacion_reporte(fecha_inicio, fecha_fin)
    respuesta = {"ok": ok, "mensaje": mensaje, "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin}
    if detalle:
        respuesta["detalle"] = detalle
    return respuesta, (200 if ok else 500)


@app.route("/descargar")
@admin_required
def descargar():
    db = get_db()
    filas = db.execute(
        "SELECT mes, vendedor, referencia, fecha, ejecutivo, venta_por, cliente_servicio, venta, profit, margen "
        "FROM reporte_bookings ORDER BY fecha DESC"
    ).fetchall()
    db.close()
    if not filas:
        flash("Todavía no hay ningún reporte generado.")
        return redirect(url_for("dashboard"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Vendedores"
    encabezados = ["Referencia", "Fecha de creación", "Vendedor", "Ejecutivo", "Venta por", "Cliente servicio", "Venta", "Profit", "Margen"]
    ws.append(encabezados)
    for r in filas:
        ws.append([
            r["referencia"], r["fecha"].strftime("%Y-%m-%d %H:%M") if r["fecha"] else "", r["vendedor"],
            r["ejecutivo"], r["venta_por"], r["cliente_servicio"],
            float(r["venta"]), float(r["profit"]), float(r["margen"]),
        ])
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True, download_name="Reporte_Vendedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/dashboard")
@login_required
def dashboard_plazas_vendedores():
    if not usuario_puede_ver_ventas():
        flash("No tienes permiso para ver Información de Ventas.")
        return redirect(url_for(primera_pagina_permitida()))

    datos = construir_datos_dashboard(plazas_permitidas_usuario())
    if datos is None:
        if session.get("es_admin"):
            flash("Todavía no hay ningún reporte descargado. Genera uno primero en 'Reporte'.")
            return redirect(url_for("dashboard"))
        flash("Todavía no hay ningún reporte generado. Pídele a un administrador que genere uno.")
        return render_template("dashboard_plazas_vendedores.html", datos_json="null", datos=None)
    datos_json = json.dumps(datos).replace("</", "<\\/")
    return render_template("dashboard_plazas_vendedores.html", datos_json=datos_json, datos=datos)


def construir_filas_reportes(plazas_permitidas=None):
    """Un registro liviano por booking (usado por /reportes y
    /reportes/por-vendedor) para que todo el filtrado y las sumas se hagan
    en el navegador, igual que en /dashboard."""
    db = get_db()

    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    bookings = db.execute(
        "SELECT fecha, vendedor, ejecutivo, referencia, venta_por, cliente_servicio, venta, profit FROM reporte_bookings ORDER BY fecha"
    ).fetchall()
    db.close()

    filas = []
    for r in bookings:
        fecha = r["fecha"]
        if fecha is None:
            continue
        vkey = normalizar(r["vendedor"])
        plaza = plaza_por_vendedor.get(vkey, "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        filas.append({
            "fecha": fecha.astimezone(TZ_LOCAL).strftime("%Y-%m-%d"),
            "mes": fecha.astimezone(TZ_LOCAL).strftime("%Y-%m"),
            "plaza": plaza,
            "vendedor": r["vendedor"] or "#N/D",
            "cliente": r["cliente_servicio"] or "Sin cliente",
            "tipo": r["venta_por"] or "Sin tipo",
            "ejecutivo": normalizar(r["ejecutivo"]),
            "tipoServicio": extraer_tipo_servicio(r["referencia"]),
            "venta": round(float(r["venta"]), 2),
            "profit": round(float(r["profit"]), 2),
        })
    return filas


def construir_presupuesto_mensual(plazas_permitidas=None):
    """Suma de presupuesto por mes (mismo criterio que la tabla Venta por
    Plaza del Dashboard: presupuesto por vendedor, filtrado por plaza),
    para comparar Venta vs Presupuesto en /reportes."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    totales = {}
    for r in db.execute("SELECT mes, vendedor, presupuesto FROM catalogo_presupuesto WHERE vendedor IS NOT NULL"):
        plaza = plaza_por_vendedor.get(normalizar(r["vendedor"]), "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        totales[r["mes"]] = totales.get(r["mes"], 0.0) + float(r["presupuesto"])
    db.close()
    return [{"mes": mes, "ppto": round(v, 2)} for mes, v in sorted(totales.items())]


@app.route("/reportes")
@reportes_required
def reportes_graficas():
    plazas_permitidas = plazas_permitidas_usuario()
    filas = construir_filas_reportes(plazas_permitidas)
    presupuesto_mensual = construir_presupuesto_mensual(plazas_permitidas)
    datos_json = json.dumps(filas).replace("</", "<\\/")
    presupuesto_json = json.dumps(presupuesto_mensual).replace("</", "<\\/")
    return render_template(
        "reportes.html", datos_json=datos_json, presupuesto_json=presupuesto_json, hay_datos=len(filas) > 0
    )


@app.route("/comisiones")
@login_required
def comisiones():
    if not usuario_puede_comisiones():
        flash("No tienes permiso para ver Comisiones.")
        return redirect(url_for("dashboard_plazas_vendedores"))

    db = get_db()
    folios_disponibles = db.execute(
        "SELECT DISTINCT folio, descripcion FROM comisiones_liquidacion_detalle WHERE folio >= 40 ORDER BY folio DESC"
    ).fetchall()

    folio_solicitado = request.args.get("folio", type=int)
    folio_actual = folio_solicitado or (folios_disponibles[0]["folio"] if folios_disponibles else None)

    filas = []
    descripcion = None
    reporte_por_booking = {}
    cobros_por_booking = {}
    if folio_actual is not None:
        filas = db.execute(
            "SELECT * FROM comisiones_liquidacion_detalle WHERE folio = %s ORDER BY booking", (folio_actual,)
        ).fetchall()
        if filas:
            descripcion = filas[0]["descripcion"]
            bookings = [f["booking"] for f in filas]
            for r in db.execute(
                "SELECT referencia, venta, margen, cliente_servicio, fecha, ejecutivo, venta_por "
                "FROM reporte_bookings WHERE referencia = ANY(%s)",
                (bookings,),
            ):
                reporte_por_booking[r["referencia"]] = {
                    "venta": float(r["venta"]),
                    "margen": float(r["margen"]),
                    "cliente": r["cliente_servicio"],
                    "fecha": r["fecha"].strftime("%Y-%m-%d") if r["fecha"] else None,
                    "ejecutivo": r["ejecutivo"],
                    "venta_por": r["venta_por"],
                }
            for r in db.execute(
                "SELECT referencia, folio_cobro, folio_factura, uuid, tipo_docto, tipo_referencia, cliente, "
                "fecha_factura, fecha_cobro, dias_diferencia, fecha_timbre, banco, moneda, subtotal, iva, "
                "descuento, retencion, total "
                "FROM comisiones_cobros_detalle WHERE referencia = ANY(%s) ORDER BY fecha_cobro",
                (bookings,),
            ):
                cobros_por_booking.setdefault(r["referencia"], []).append({
                    "folio_cobro": r["folio_cobro"],
                    "folio_factura": r["folio_factura"],
                    "uuid": r["uuid"],
                    "tipo_docto": r["tipo_docto"],
                    "tipo_referencia": r["tipo_referencia"],
                    "cliente": r["cliente"],
                    "fecha_factura": r["fecha_factura"].strftime("%Y-%m-%d") if r["fecha_factura"] else None,
                    "fecha_cobro": r["fecha_cobro"].strftime("%Y-%m-%d") if r["fecha_cobro"] else None,
                    "dias_diferencia": r["dias_diferencia"],
                    "fecha_timbre": r["fecha_timbre"],
                    "banco": r["banco"],
                    "moneda": r["moneda"],
                    "subtotal": float(r["subtotal"]),
                    "iva": float(r["iva"]),
                    "descuento": float(r["descuento"]),
                    "retencion": float(r["retencion"]),
                    "total": float(r["total"]),
                })

    cobros_cargados = 0
    if folio_actual is not None:
        cobros_cargados = db.execute(
            "SELECT count(*) AS n FROM comisiones_cobros_detalle WHERE folio = %s", (folio_actual,)
        ).fetchone()["n"]
    db.close()

    total_profit = sum(float(f["profit"]) for f in filas)
    total_vendedor = sum(float(f["total_vendedor"]) for f in filas)
    total_desarrollador = sum(float(f["total_desarrollador"]) for f in filas)
    total_venta = sum(reporte_por_booking.get(f["booking"], {}).get("venta") or 0.0 for f in filas)
    margen_total = (total_profit / total_venta) if total_venta else 0.0

    def agrupar(filas, campo_nombre, campo_total):
        grupos = {}
        for f in filas:
            nombre = f[campo_nombre] or "(sin nombre)"
            g = grupos.setdefault(nombre, {"nombre": nombre, "total": 0.0, "cant_book": 0})
            g["total"] += float(f[campo_total])
            g["cant_book"] += 1
        return sorted(grupos.values(), key=lambda g: g["total"], reverse=True)

    por_vendedor = agrupar(filas, "vendedor", "total_vendedor")
    por_desarrollador = agrupar(filas, "desarrollador", "total_desarrollador")

    filas_json = json.dumps([
        {
            "booking": f["booking"],
            "folio_cobro": f["folio_cobro"],
            "profit": float(f["profit"]),
            "venta": reporte_por_booking.get(f["booking"], {}).get("venta"),
            "margen": reporte_por_booking.get(f["booking"], {}).get("margen"),
            "cliente": reporte_por_booking.get(f["booking"], {}).get("cliente"),
            "fecha": reporte_por_booking.get(f["booking"], {}).get("fecha"),
            "ejecutivo": reporte_por_booking.get(f["booking"], {}).get("ejecutivo"),
            "venta_por": reporte_por_booking.get(f["booking"], {}).get("venta_por"),
            "cobros": cobros_por_booking.get(f["booking"], []),
            "vendedor": f["vendedor"],
            "pct_vendedor": float(f["pct_vendedor"]) if f["pct_vendedor"] is not None else None,
            "total_vendedor": float(f["total_vendedor"]),
            "desarrollador": f["desarrollador"],
            "pct_desarrollador": float(f["pct_desarrollador"]) if f["pct_desarrollador"] is not None else None,
            "total_desarrollador": float(f["total_desarrollador"]),
        }
        for f in filas
    ]).replace("</", "<\\/")

    folios_cargolink = []
    if session.get("es_admin"):
        try:
            folios_cargolink = listar_folios_liquidacion_cargolink()
        except RuntimeError as e:
            flash(f"No se pudo consultar la lista de folios en CargoLink: {e}")

    return render_template(
        "comisiones.html",
        folios_disponibles=folios_disponibles,
        folios_cargolink=folios_cargolink,
        folio_actual=folio_actual,
        descripcion=descripcion,
        filas=filas,
        filas_json=filas_json,
        total_profit=total_profit,
        total_vendedor=total_vendedor,
        total_desarrollador=total_desarrollador,
        total_venta=total_venta,
        margen_total=margen_total,
        total_liquidar=total_vendedor + total_desarrollador,
        por_vendedor=por_vendedor,
        por_desarrollador=por_desarrollador,
        cobros_cargados=cobros_cargados,
    )


@app.route("/comisiones/exportar")
@login_required
def comisiones_exportar():
    if not usuario_puede_comisiones():
        flash("No tienes permiso para ver Comisiones.")
        return redirect(url_for("dashboard_plazas_vendedores"))

    folio = request.args.get("folio", type=int)
    campo = request.args.get("campo")
    nombre = request.args.get("nombre", "")
    if folio is None or campo not in ("vendedor", "desarrollador") or not nombre:
        flash("Parámetros inválidos para exportar.")
        return redirect(url_for("comisiones"))

    db = get_db()
    filas = db.execute(
        f"SELECT * FROM comisiones_liquidacion_detalle WHERE folio = %s AND {campo} = %s ORDER BY booking",
        (folio, nombre),
    ).fetchall()
    descripcion = filas[0]["descripcion"] if filas else ""

    venta_por_booking = {}
    if filas:
        bookings = [f["booking"] for f in filas]
        for r in db.execute("SELECT referencia, venta FROM reporte_bookings WHERE referencia = ANY(%s)", (bookings,)):
            venta_por_booking[r["referencia"]] = float(r["venta"])
    db.close()

    campo_total = "total_vendedor" if campo == "vendedor" else "total_desarrollador"
    etiqueta = "Vendedor" if campo == "vendedor" else "Desarrollador"

    profit_total = sum(float(f["profit"]) for f in filas)
    comision_total = sum(float(f[campo_total]) for f in filas)
    venta_total = sum(venta_por_booking.get(f["booking"], 0.0) for f in filas)
    margen = (profit_total / venta_total) if venta_total else 0.0

    amarillo = PatternFill(fill_type="solid", fgColor="FFF59D")
    naranja = PatternFill(fill_type="solid", fgColor="C1502E")
    lado = Side(style="thin", color="E6C200")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)
    centrado = Alignment(horizontal="center", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center")
    derecha = Alignment(horizontal="right", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comisiones"

    ws.merge_cells("A1:C1")
    titulo = ws["A1"]
    titulo.value = f"Comisiones {descripcion}"
    titulo.font = Font(bold=True, size=13)
    titulo.alignment = centrado
    ws.row_dimensions[1].height = 22

    ws["A3"] = etiqueta
    ws["A3"].font = Font(bold=True)
    ws.merge_cells("B3:C3")
    ws["B3"] = nombre
    ws["B3"].font = Font(bold=True)
    ws["B3"].alignment = centrado
    for coord in ("A3", "B3", "C3"):
        ws[coord].fill = amarillo
        ws[coord].border = borde

    stats = [
        ("Cantidad de Booking", len(filas), None),
        ("Venta Total USD", venta_total, "#,##0.00"),
        ("Profit USD", profit_total, "#,##0.00"),
        ("Margen", margen, "0.00%"),
        ("Comisión", comision_total, "#,##0.00"),
    ]
    fila = 5
    for etq, valor, formato in stats:
        celda_etq = ws.cell(row=fila, column=1, value=etq)
        celda_etq.font = Font(bold=True)
        celda_etq.fill = amarillo
        celda_etq.border = borde
        celda_valor = ws.cell(row=fila, column=2, value=valor)
        celda_valor.font = Font(bold=True)
        celda_valor.alignment = derecha
        celda_valor.fill = amarillo
        celda_valor.border = borde
        if formato:
            celda_valor.number_format = formato
        fila += 1

    fila_tabla = fila + 1
    encabezados = ["Booking", "Profit", f"Comisión {etiqueta}"]
    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_tabla, column=col, value=encabezado)
        celda.fill = naranja
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = centrado if col == 1 else derecha

    for i, f in enumerate(filas, start=fila_tabla + 1):
        ws.cell(row=i, column=1, value=f["booking"]).alignment = izquierda
        celda_profit = ws.cell(row=i, column=2, value=float(f["profit"]))
        celda_profit.number_format = "#,##0.00"
        celda_profit.alignment = derecha
        celda_com = ws.cell(row=i, column=3, value=float(f[campo_total]))
        celda_com.number_format = "#,##0.00"
        celda_com.alignment = derecha

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = ws.cell(row=fila_tabla + 1, column=1).coordinate

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = re.sub(r"[^a-zA-Z0-9]+", "_", f"Comisiones_{nombre}_{descripcion}").strip("_") + ".xlsx"
    return send_file(
        buffer, as_attachment=True, download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/comisiones/cargar", methods=["POST"])
@admin_required
def comisiones_cargar():
    folio_raw = (request.form.get("folio") or "").strip()
    try:
        folio = int(folio_raw)
    except ValueError:
        flash("Folio inválido.")
        return redirect(url_for("comisiones"))

    try:
        resultado = descargar_liquidacion_vendedor_cargolink(folio)
    except RuntimeError as e:
        flash(str(e))
        return redirect(url_for("comisiones"))

    db = get_db()
    db.execute("DELETE FROM comisiones_liquidacion_detalle WHERE folio = %s", (folio,))
    for d in resultado["detalle"]:
        db.execute(
            """
            INSERT INTO comisiones_liquidacion_detalle
                (folio, descripcion, booking, folio_cobro, profit, vendedor, pct_vendedor,
                 total_vendedor, desarrollador, pct_desarrollador, total_desarrollador)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                resultado["folio"], resultado["descripcion"], d["booking"], d["folio_cobro"], d["profit"],
                d["vendedor"], d["pct_vendedor"], d["total_vendedor"], d["desarrollador"],
                d["pct_desarrollador"], d["total_desarrollador"],
            ),
        )
    db.commit()
    db.close()

    flash(f"Folio {folio} cargado: {len(resultado['detalle'])} booking(s).")
    return redirect(url_for("comisiones", folio=folio))


@app.route("/comisiones/cobros/cargar", methods=["POST"])
@admin_required
def comisiones_cobros_cargar():
    folio_raw = (request.form.get("folio") or "").strip()
    descripcion = (request.form.get("descripcion") or "").strip()
    fecha_ini = (request.form.get("fecha_ini") or "").strip()
    fecha_fin = (request.form.get("fecha_fin") or "").strip()
    try:
        folio = int(folio_raw)
    except ValueError:
        flash("Folio inválido.")
        return redirect(url_for("comisiones"))
    if not DATE_RE.match(fecha_ini) or not DATE_RE.match(fecha_fin):
        flash("Selecciona una fecha inicial y una fecha final válidas.")
        return redirect(url_for("comisiones", folio=folio))

    etiqueta = f"{folio} {descripcion}".strip()

    try:
        cobros = descargar_concentrado_cobros_cargolink(fecha_ini, fecha_fin)
    except RuntimeError as e:
        flash(str(e))
        return redirect(url_for("comisiones", folio=folio))

    db = get_db()
    db.execute("DELETE FROM comisiones_cobros_detalle WHERE folio = %s", (folio,))
    for c in cobros:
        db.execute(
            """
            INSERT INTO comisiones_cobros_detalle
                (folio, etiqueta, folio_cobro, folio_factura, referencia, uuid, tipo_docto, tipo_referencia,
                 cliente, fecha_factura, fecha_cobro, dias_diferencia, fecha_timbre, banco, moneda,
                 subtotal, iva, descuento, retencion, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                folio, etiqueta, c["folio_cobro"], c["folio_factura"], c["referencia"], c["uuid"],
                c["tipo_docto"], c["tipo_referencia"], c["cliente"], c["fecha_factura"], c["fecha_cobro"],
                c["dias_diferencia"], c["fecha_timbre"], c["banco"], c["moneda"], c["subtotal"], c["iva"],
                c["descuento"], c["retencion"], c["total"],
            ),
        )
    db.commit()
    db.close()

    flash(f'Cobros cargados para "{etiqueta}": {len(cobros)} registro(s).')
    return redirect(url_for("comisiones", folio=folio))


def calcular_comisiones_acotadas(
    filas, cobros_por_booking, tasa_usd, tasa_eur, dias_umbral,
    pct_vendedor_arriba, pct_vendedor_debajo, pct_desarrollador_arriba, pct_desarrollador_debajo,
):
    """Comisión 'acotada': por cada booking se reparte el profit entre sus
    cobros según su % de participación en MXN (convirtiendo USD/EUR con la
    tasa dada), agrupando ese % en dos buckets según dias_diferencia
    (> dias_umbral o <= dias_umbral). A cada bucket se le aplica una tasa
    distinta para vendedor y desarrollador, y la comisión es la suma de
    ambos. Si el booking no tiene cobros utilizables (sin registros, sin
    dias_diferencia, o en una moneda sin tasa) se mantiene su comisión
    actual sin cambios."""
    tasas_moneda = {"MXN": 1.0, "USD": tasa_usd, "EUR": tasa_eur}
    resultado = []
    for f in filas:
        booking = f["booking"]
        profit = float(f["profit"])
        total_vendedor_actual = float(f["total_vendedor"])
        total_desarrollador_actual = float(f["total_desarrollador"])

        cobros_raw = cobros_por_booking.get(booking) or []
        filas_calc = []
        for c in cobros_raw:
            tasa = tasas_moneda.get(c["moneda"])
            usable = tasa is not None and c["dias_diferencia"] is not None
            monto_mxn = float(c["total"]) * tasa if tasa is not None else None
            filas_calc.append({"c": c, "monto_mxn": monto_mxn, "dias": c["dias_diferencia"], "usable": usable})

        usables = [v for v in filas_calc if v["usable"]]
        suma_mxn = sum(v["monto_mxn"] for v in usables)
        con_regla = bool(usables) and suma_mxn != 0

        if con_regla:
            for v in usables:
                v["pct"] = v["monto_mxn"] / suma_mxn * 100
            pct_arriba = sum(v["pct"] for v in usables if v["dias"] > dias_umbral)
            pct_debajo = sum(v["pct"] for v in usables if v["dias"] <= dias_umbral)
            profit_arriba = profit * pct_arriba / 100
            profit_debajo = profit * pct_debajo / 100
            total_vendedor_acotada = round(
                profit_arriba * pct_vendedor_arriba / 100 + profit_debajo * pct_vendedor_debajo / 100, 2
            )
            total_desarrollador_acotada = round(
                profit_arriba * pct_desarrollador_arriba / 100 + profit_debajo * pct_desarrollador_debajo / 100, 2
            )
        else:
            pct_arriba = pct_debajo = None
            total_vendedor_acotada = total_vendedor_actual
            total_desarrollador_acotada = total_desarrollador_actual

        cobros_out = []
        for v in filas_calc:
            c = v["c"]
            bucket = None
            if v["usable"]:
                bucket = "arriba" if v["dias"] > dias_umbral else "debajo"
            cobros_out.append({
                "folio_cobro": c["folio_cobro"],
                "folio_factura": c["folio_factura"],
                "uuid": c["uuid"],
                "tipo_docto": c["tipo_docto"],
                "tipo_referencia": c["tipo_referencia"],
                "cliente": c["cliente"],
                "fecha_factura": c["fecha_factura"].strftime("%Y-%m-%d") if c["fecha_factura"] else None,
                "fecha_cobro": c["fecha_cobro"].strftime("%Y-%m-%d") if c["fecha_cobro"] else None,
                "dias_diferencia": c["dias_diferencia"],
                "fecha_timbre": c["fecha_timbre"],
                "banco": c["banco"],
                "moneda": c["moneda"],
                "subtotal": float(c["subtotal"]),
                "iva": float(c["iva"]),
                "descuento": float(c["descuento"]),
                "retencion": float(c["retencion"]),
                "total": float(c["total"]),
                "monto_mxn": round(v["monto_mxn"], 2) if v["monto_mxn"] is not None else None,
                "pct_participacion": round(v["pct"], 2) if v.get("pct") is not None else None,
                "bucket": bucket,
            })

        resultado.append({
            "booking": booking, "vendedor": f["vendedor"], "desarrollador": f["desarrollador"],
            "profit": profit,
            "pct_arriba": round(pct_arriba, 2) if pct_arriba is not None else None,
            "pct_debajo": round(pct_debajo, 2) if pct_debajo is not None else None,
            "total_vendedor_actual": total_vendedor_actual, "total_vendedor_acotada": total_vendedor_acotada,
            "total_desarrollador_actual": total_desarrollador_actual,
            "total_desarrollador_acotada": total_desarrollador_acotada,
            "con_regla": con_regla,
            "cobros": cobros_out,
        })
    return resultado


@app.route("/comisiones-acotadas")
@login_required
def comisiones_acotadas():
    if not session.get("es_admin"):
        flash("No tienes permiso para ver Comisiones Acotadas.")
        return redirect(url_for("dashboard_plazas_vendedores"))

    db = get_db()
    folios_disponibles = db.execute(
        "SELECT DISTINCT folio, descripcion FROM comisiones_liquidacion_detalle WHERE folio >= 40 ORDER BY folio DESC"
    ).fetchall()

    folio_solicitado = request.args.get("folio", type=int)
    folio_actual = folio_solicitado or (folios_disponibles[0]["folio"] if folios_disponibles else None)

    tasa_usd = request.args.get("usd", type=float)
    tasa_eur = request.args.get("eur", type=float)
    dias_umbral = request.args.get("dias", type=int) or 35
    pct_vendedor_arriba = request.args.get("pct_v_arriba", type=float)
    pct_vendedor_arriba = 5.0 if pct_vendedor_arriba is None else pct_vendedor_arriba
    pct_desarrollador_arriba = request.args.get("pct_d_arriba", type=float)
    pct_desarrollador_arriba = 1.0 if pct_desarrollador_arriba is None else pct_desarrollador_arriba
    pct_vendedor_debajo = 7.0
    pct_desarrollador_debajo = 2.0

    filas = []
    descripcion = None
    resultados = []
    if folio_actual is not None:
        filas = db.execute(
            "SELECT * FROM comisiones_liquidacion_detalle WHERE folio = %s ORDER BY booking", (folio_actual,)
        ).fetchall()
        if filas:
            descripcion = filas[0]["descripcion"]
            bookings = [f["booking"] for f in filas]
            cobros_por_booking = {}
            for r in db.execute(
                "SELECT referencia, folio_cobro, folio_factura, uuid, tipo_docto, tipo_referencia, cliente, "
                "fecha_factura, fecha_cobro, dias_diferencia, fecha_timbre, banco, moneda, subtotal, iva, "
                "descuento, retencion, total "
                "FROM comisiones_cobros_detalle WHERE referencia = ANY(%s) ORDER BY fecha_cobro",
                (bookings,),
            ):
                cobros_por_booking.setdefault(r["referencia"], []).append(r)

            if tasa_usd and tasa_eur:
                resultados = calcular_comisiones_acotadas(
                    filas, cobros_por_booking, tasa_usd, tasa_eur, dias_umbral,
                    pct_vendedor_arriba, pct_vendedor_debajo, pct_desarrollador_arriba, pct_desarrollador_debajo,
                )
    db.close()

    con_regla = sum(1 for r in resultados if r["con_regla"])
    total_vendedor_actual = sum(r["total_vendedor_actual"] for r in resultados)
    total_vendedor_acotada = sum(r["total_vendedor_acotada"] for r in resultados)
    total_desarrollador_actual = sum(r["total_desarrollador_actual"] for r in resultados)
    total_desarrollador_acotada = sum(r["total_desarrollador_acotada"] for r in resultados)

    def agrupar(campo_nombre, campo_actual, campo_acotada):
        grupos = {}
        for r in resultados:
            nombre = r[campo_nombre] or "(sin nombre)"
            g = grupos.setdefault(nombre, {"nombre": nombre, "cant_book": 0, "actual": 0.0, "acotada": 0.0})
            g["cant_book"] += 1
            g["actual"] += r[campo_actual]
            g["acotada"] += r[campo_acotada]
        return sorted(grupos.values(), key=lambda g: g["acotada"], reverse=True)

    por_vendedor = agrupar("vendedor", "total_vendedor_actual", "total_vendedor_acotada")
    por_desarrollador = agrupar("desarrollador", "total_desarrollador_actual", "total_desarrollador_acotada")

    resultados_json = json.dumps(resultados).replace("</", "<\\/")

    return render_template(
        "comisiones_acotadas.html",
        folios_disponibles=folios_disponibles,
        folio_actual=folio_actual,
        descripcion=descripcion,
        tasa_usd=tasa_usd,
        tasa_eur=tasa_eur,
        dias_umbral=dias_umbral,
        pct_vendedor_arriba=pct_vendedor_arriba,
        pct_vendedor_debajo=pct_vendedor_debajo,
        pct_desarrollador_arriba=pct_desarrollador_arriba,
        pct_desarrollador_debajo=pct_desarrollador_debajo,
        resultados=resultados,
        resultados_json=resultados_json,
        total_bookings=len(filas),
        con_regla=con_regla,
        sin_regla=len(resultados) - con_regla,
        total_vendedor_actual=total_vendedor_actual,
        total_vendedor_acotada=total_vendedor_acotada,
        total_desarrollador_actual=total_desarrollador_actual,
        total_desarrollador_acotada=total_desarrollador_acotada,
        por_vendedor=por_vendedor,
        por_desarrollador=por_desarrollador,
    )


@app.route("/reportes/por-vendedor")
@reportes_required
def reportes_por_vendedor():
    filas = construir_filas_reportes(plazas_permitidas_usuario())
    datos_json = json.dumps(filas).replace("</", "<\\/")
    return render_template("reportes_por_vendedor.html", datos_json=datos_json, hay_datos=len(filas) > 0)


@app.route("/reportes/por-cliente")
@reportes_required
def reportes_por_cliente():
    filas = construir_filas_reportes(plazas_permitidas_usuario())
    datos_json = json.dumps(filas).replace("</", "<\\/")
    return render_template("reportes_por_cliente.html", datos_json=datos_json, hay_datos=len(filas) > 0)


@app.route("/reportes/clientes-mensual")
@reportes_required
def reportes_clientes_mensual():
    db = get_db()

    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    # El booking pertenece al cliente (cliente_servicio) sin importar quién
    # de reporte_bookings.vendedor lo vendió — aquí se agrupa por el
    # vendedor ASIGNADO al cliente en asignacion_de_clientes, que es lo que
    # pidió el usuario ("clientes asignados por vendedor"). Se guarda
    # también el set de vendedores reales detrás de cada mes para poder
    # marcar en rojo cuando alguna venta la generó alguien distinto al
    # vendedor asignado actualmente.
    mensual_por_cliente = {}
    for r in db.execute("SELECT mes, vendedor, cliente_servicio, profit FROM reporte_bookings"):
        ckey = normalizar(r["cliente_servicio"])
        if not ckey:
            continue
        por_mes = mensual_por_cliente.setdefault(ckey, {})
        acc = por_mes.setdefault(r["mes"], {"cant": 0, "profit": 0.0, "vendedores": set()})
        acc["cant"] += 1
        acc["profit"] += float(r["profit"])
        acc["vendedores"].add(normalizar(r["vendedor"]))

    filas_clientes = db.execute(
        "SELECT folio, razon_social, vendedor FROM asignacion_de_clientes WHERE vendedor IS NOT NULL ORDER BY vendedor, razon_social"
    ).fetchall()
    db.close()

    plazas_permitidas = plazas_permitidas_usuario()
    filas = []
    for r in filas_clientes:
        vkey = normalizar(r["vendedor"])
        ckey = normalizar(r["razon_social"])
        plaza = plaza_por_vendedor.get(vkey, "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        mensual = {}
        for mes, acc in mensual_por_cliente.get(ckey, {}).items():
            mensual[mes] = {
                "cant": acc["cant"],
                "profit": acc["profit"],
                "otroVendedor": bool(acc["vendedores"] - {vkey}),
            }
        filas.append({
            "folio": r["folio"],
            "razonSocial": r["razon_social"],
            "vendedor": r["vendedor"],
            "plaza": plaza,
            "mensual": mensual,
        })

    datos_json = json.dumps(filas).replace("</", "<\\/")
    return render_template("reportes_clientes_mensual.html", datos_json=datos_json, hay_datos=len(filas) > 0)


CRM_NAV = [
    {"grupo": None, "texto": "Inicio", "slug": "inicio"},
    {"grupo": None, "texto": "Booking", "slug": "booking"},
    {"grupo": None, "texto": "Negocios", "slug": "negocios"},
    {"grupo": None, "texto": "Cotizaciones", "slug": "cotizaciones"},
    {"grupo": None, "texto": "Tareas", "slug": "tareas"},
    {"grupo": "Catálogos", "texto": "Clientes", "slug": "clientes"},
    {"grupo": "Catálogos", "texto": "Grupo", "slug": "grupo"},
    {"grupo": "Catálogos", "texto": "Contactos", "slug": "contactos"},
    {"grupo": "Catálogos", "texto": "Productos", "slug": "productos"},
    {"grupo": "Catálogos", "texto": "Motivos de Pérdida", "slug": "motivos-perdida"},
    {"grupo": "Catálogos", "texto": "Estatus", "slug": "estatus"},
    {"grupo": "Catálogos", "texto": "Estado de la República", "slug": "estado-republica"},
    {"grupo": "Catálogos", "texto": "Etapa del Negocio", "slug": "etapa-negocio"},
    {"grupo": "Catálogos", "texto": "Actividades", "slug": "actividades"},
    {"grupo": "Administración", "texto": "Usuarios", "slug": "usuarios"},
    {"grupo": "Administración", "texto": "Representantes", "slug": "representantes"},
    {"grupo": "Administración", "texto": "Almacenamiento", "slug": "almacenamiento"},
    {"grupo": "Administración", "texto": "Configuración", "slug": "configuracion"},
]

# Datos de ejemplo para la vista de Tareas mientras se define la fuente de datos real.
TAREAS_MOCK = []


def construir_booking_crm(plazas_permitidas=None, fecha_inicio=None, fecha_fin=None, limite=300):
    """Detalle a nivel booking para la pantalla CRM → Booking: mismos datos
    (tabla reporte_bookings) y mismo filtro de plazas que Reporte de Ventas,
    pero mostrados fila por fila en vez de agregados. `limite` acota la
    respuesta a los bookings más recientes (o más recientes dentro del rango
    de fechas dado): la tabla puede tener miles de filas y no vale la pena
    paginar/virtualizar solo para esta vista de detalle mientras se define
    si de verdad se necesita ver el histórico completo aquí."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    condiciones = []
    parametros = []
    if fecha_inicio:
        condiciones.append("fecha >= %s")
        parametros.append(fecha_inicio)
    if fecha_fin:
        condiciones.append("fecha < (%s::date + interval '1 day')")
        parametros.append(fecha_fin)
    where = f"WHERE {' AND '.join(condiciones)}" if condiciones else ""

    bookings = db.execute(
        "SELECT referencia, fecha, vendedor, ejecutivo, venta_por, cliente_servicio, venta, profit, margen "
        f"FROM reporte_bookings {where} ORDER BY fecha DESC",
        parametros,
    ).fetchall()
    db.close()

    filas = []
    for r in bookings:
        fecha = r["fecha"]
        if fecha is None:
            continue
        plaza = plaza_por_vendedor.get(normalizar(r["vendedor"]), "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        filas.append({
            "referencia": r["referencia"] or "#N/D",
            "fecha": fecha.astimezone(TZ_LOCAL).strftime("%Y-%m-%d"),
            "plaza": plaza,
            "vendedor": r["vendedor"] or "#N/D",
            "ejecutivo": normalizar(r["ejecutivo"]) or "#N/D",
            "venta_por": r["venta_por"] or "Sin tipo",
            "cliente_servicio": r["cliente_servicio"] or "Sin cliente",
            "venta": round(float(r["venta"]), 2),
            "profit": round(float(r["profit"]), 2),
            "margen": round(float(r["margen"]), 2),
        })
    total = len(filas)
    return filas[:limite], total


def construir_clientes_crm(plazas_permitidas=None):
    """Cliente + vendedor + desarrollador asignado para CRM → Clientes:
    misma fuente (asignacion_de_clientes) y filtro de plazas que Reportes →
    Clientes Asignados, pero solo con las columnas que pidió mostrar aquí.
    También trae si el cliente tiene bookings reales (reporte_bookings,
    casados por nombre) y si tiene cotizaciones reales (crm_cotizaciones,
    por cliente_folio) — Tareas todavía no tiene una fuente de datos
    conectada, así que no se puede calcular aquí."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    bookings_por_cliente = set()
    for r in db.execute(
        "SELECT DISTINCT cliente_servicio FROM reporte_bookings WHERE cliente_servicio IS NOT NULL"
    ):
        bookings_por_cliente.add(normalizar(r["cliente_servicio"]))

    clientes_con_cotizacion = set()
    for r in db.execute(
        "SELECT DISTINCT cliente_folio FROM crm_cotizaciones WHERE cliente_folio IS NOT NULL"
    ):
        clientes_con_cotizacion.add(r["cliente_folio"])

    filas_clientes = db.execute(
        "SELECT folio, razon_social, vendedor, desarrollador "
        "FROM asignacion_de_clientes WHERE vendedor IS NOT NULL ORDER BY razon_social"
    ).fetchall()
    db.close()

    filas = []
    for r in filas_clientes:
        plaza = plaza_por_vendedor.get(normalizar(r["vendedor"]), "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        filas.append({
            "folio": r["folio"],
            "cliente": r["razon_social"],
            "vendedor": r["vendedor"] or "#N/D",
            "desarrollador": r["desarrollador"] or "Sin asignar",
            "tiene_booking": normalizar(r["razon_social"]) in bookings_por_cliente,
            "tiene_cotizacion": r["folio"] in clientes_con_cotizacion,
        })
    return filas


def construir_cotizaciones_resumen_crm(cliente_folio=None, contacto_id=None):
    """Cotizaciones ligadas a un cliente (por cliente_folio) o a un contacto
    (por contacto_id), para los paneles 'Cotizaciones' de sus páginas de
    detalle. Trae el gran total (suma de líneas + su impuesto, por moneda)
    de cada una."""
    if cliente_folio is None and contacto_id is None:
        return []
    db = get_db()
    if cliente_folio is not None:
        filas = db.execute("""
            SELECT id, id_cotizacion, nombre_cotizacion, fecha_creacion
            FROM crm_cotizaciones WHERE cliente_folio = %s
            ORDER BY fecha_creacion DESC
        """, (cliente_folio,)).fetchall()
    else:
        filas = db.execute("""
            SELECT id, id_cotizacion, nombre_cotizacion, fecha_creacion
            FROM crm_cotizaciones WHERE contacto_id = %s
            ORDER BY fecha_creacion DESC
        """, (contacto_id,)).fetchall()

    resultado = []
    for f in filas:
        totales = db.execute("""
            SELECT moneda, SUM(cantidad * precio_unitario) AS subtotal, SUM(cantidad * precio_unitario * impuesto / 100.0) AS impuesto
            FROM crm_cotizacion_productos WHERE cotizacion_id = %s GROUP BY moneda
        """, (f["id"],)).fetchall()
        gran_total_texto = " / ".join(
            f"{t['moneda']} ${float(t['subtotal']) + float(t['impuesto'] or 0):,.2f}" for t in totales
        )
        resultado.append({
            "id": f["id"],
            "id_cotizacion": f["id_cotizacion"],
            "nombre_cotizacion": f["nombre_cotizacion"] or "",
            "fecha_creacion": f["fecha_creacion"].strftime("%Y-%m-%d") if f["fecha_creacion"] else "",
            "gran_total_texto": gran_total_texto,
        })
    db.close()
    return resultado


def construir_cliente_detalle_crm(folio, plazas_permitidas=None):
    """Info de un cliente (asignacion_de_clientes) + sus bookings reales
    (reporte_bookings, casados por nombre de cliente) para la página de
    detalle de CRM → Clientes. Regresa None si el folio no existe o si el
    usuario no tiene permiso de ver la plaza de ese cliente."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    cliente = db.execute(
        "SELECT folio, razon_social, vendedor, desarrollador, tipo_cliente, cant_booking, fecha_ultimo_booking "
        "FROM asignacion_de_clientes WHERE folio = %s",
        (folio,),
    ).fetchone()
    if cliente is None:
        db.close()
        return None

    plaza = plaza_por_vendedor.get(normalizar(cliente["vendedor"]), "#N/D")
    if plazas_permitidas is not None and plaza not in plazas_permitidas:
        db.close()
        return None

    bookings_cliente = db.execute(
        "SELECT referencia, fecha, venta, profit "
        "FROM reporte_bookings WHERE cliente_servicio ILIKE %s ORDER BY fecha DESC",
        (cliente["razon_social"],),
    ).fetchall()
    db.close()

    bookings = [{
        "referencia": b["referencia"] or "#N/D",
        "fecha": b["fecha"].astimezone(TZ_LOCAL).strftime("%Y-%m-%d") if b["fecha"] else "",
        "venta": round(float(b["venta"]), 2),
        "profit": round(float(b["profit"]), 2),
    } for b in bookings_cliente]

    return {
        "folio": cliente["folio"],
        "razon_social": cliente["razon_social"],
        "vendedor": cliente["vendedor"] or "#N/D",
        "desarrollador": cliente["desarrollador"] or "Sin asignar",
        "plaza": plaza,
        "tipo_cliente": cliente["tipo_cliente"] or "Sin clasificar",
        "cant_booking": cliente["cant_booking"] or 0,
        "fecha_ultimo_booking": cliente["fecha_ultimo_booking"].strftime("%Y-%m-%d") if cliente["fecha_ultimo_booking"] else "",
        "bookings": bookings,
    }


def construir_contactos_crm(plazas_permitidas=None):
    """Contactos del CRM con sus clientes y grupos asociados (muchos-a-muchos).
    Un contacto sin ningún cliente asociado todavía es visible para todos
    (no hay plaza que restringir); si tiene clientes, solo es visible si
    alguno de esos clientes cae en una plaza permitida para el usuario.
    También trae si el contacto tiene cotizaciones reales (crm_cotizaciones,
    por contacto_id)."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    clientes_con_booking = set()
    for r in db.execute(
        "SELECT DISTINCT cliente_servicio FROM reporte_bookings WHERE cliente_servicio IS NOT NULL"
    ):
        clientes_con_booking.add(normalizar(r["cliente_servicio"]))

    contactos_con_cotizacion = set()
    for r in db.execute(
        "SELECT DISTINCT contacto_id FROM crm_cotizaciones WHERE contacto_id IS NOT NULL"
    ):
        contactos_con_cotizacion.add(r["contacto_id"])

    filas = db.execute("""
        SELECT c.id, c.nombre, c.apellido, c.telefono, c.correo, c.observaciones,
               COALESCE(array_agg(DISTINCT ac.razon_social) FILTER (WHERE ac.razon_social IS NOT NULL), '{}') AS clientes,
               COALESCE(array_agg(DISTINCT ac.vendedor) FILTER (WHERE ac.vendedor IS NOT NULL), '{}') AS vendedores,
               COALESCE(array_agg(DISTINCT g.nombre) FILTER (WHERE g.nombre IS NOT NULL), '{}') AS grupos
        FROM crm_contactos c
        LEFT JOIN crm_contacto_clientes cc ON cc.contacto_id = c.id
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = cc.cliente_folio
        LEFT JOIN crm_contacto_grupos cg ON cg.contacto_id = c.id
        LEFT JOIN crm_grupos g ON g.id = cg.grupo_id
        GROUP BY c.id
        ORDER BY c.nombre, c.apellido
    """).fetchall()
    db.close()

    resultado = []
    for r in filas:
        plazas_contacto = {plaza_por_vendedor.get(normalizar(v), "#N/D") for v in r["vendedores"]}
        if plazas_permitidas is not None and plazas_contacto and not (plazas_contacto & plazas_permitidas):
            continue
        tiene_booking = any(normalizar(cl) in clientes_con_booking for cl in r["clientes"])
        resultado.append({
            "id": r["id"],
            "nombre": r["nombre"],
            "apellido": r["apellido"] or "",
            "telefono": r["telefono"] or "",
            "correo": r["correo"] or "",
            "observaciones": r["observaciones"] or "",
            "clientes": sorted(r["clientes"]),
            "grupos": sorted(r["grupos"]),
            "tiene_booking": tiene_booking,
            "tiene_cotizacion": r["id"] in contactos_con_cotizacion,
        })
    return resultado


def construir_contacto_detalle_crm(contacto_id, plazas_permitidas=None):
    """Info de un contacto + sus clientes/grupos asociados + los bookings
    reales de esos clientes (reporte_bookings, casados por nombre), para la
    página de detalle de CRM → Contactos."""
    db = get_db()
    contacto = db.execute("SELECT * FROM crm_contactos WHERE id = %s", (contacto_id,)).fetchone()
    if contacto is None:
        db.close()
        return None

    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    clientes = db.execute("""
        SELECT ac.folio, ac.razon_social, ac.vendedor
        FROM crm_contacto_clientes cc
        JOIN asignacion_de_clientes ac ON ac.folio = cc.cliente_folio
        WHERE cc.contacto_id = %s
        ORDER BY ac.razon_social
    """, (contacto_id,)).fetchall()

    plazas_contacto = {plaza_por_vendedor.get(normalizar(c["vendedor"]), "#N/D") for c in clientes}
    if plazas_permitidas is not None and plazas_contacto and not (plazas_contacto & plazas_permitidas):
        db.close()
        return None

    grupos = db.execute("""
        SELECT g.nombre FROM crm_contacto_grupos cg
        JOIN crm_grupos g ON g.id = cg.grupo_id
        WHERE cg.contacto_id = %s ORDER BY g.nombre
    """, (contacto_id,)).fetchall()

    bookings = []
    if clientes:
        nombres_lower = [c["razon_social"].lower() for c in clientes]
        filas_bookings = db.execute(
            "SELECT referencia, fecha, cliente_servicio, venta "
            "FROM reporte_bookings WHERE lower(cliente_servicio) = ANY(%s) ORDER BY fecha DESC",
            (nombres_lower,),
        ).fetchall()
        bookings = [{
            "referencia": b["referencia"] or "#N/D",
            "fecha": b["fecha"].astimezone(TZ_LOCAL).strftime("%Y-%m-%d") if b["fecha"] else "",
            "cliente": b["cliente_servicio"],
            "venta": round(float(b["venta"]), 2),
        } for b in filas_bookings]
    db.close()

    return {
        "id": contacto["id"],
        "nombre": contacto["nombre"],
        "apellido": contacto["apellido"] or "",
        "telefono": contacto["telefono"] or "",
        "correo": contacto["correo"] or "",
        "observaciones": contacto["observaciones"] or "",
        "clientes": [c["razon_social"] for c in clientes],
        "grupos": [g["nombre"] for g in grupos],
        "bookings": bookings,
    }


def opciones_clientes_grupos_crm():
    db = get_db()
    clientes = db.execute(
        "SELECT folio, razon_social FROM asignacion_de_clientes WHERE folio IS NOT NULL ORDER BY razon_social"
    ).fetchall()
    grupos = db.execute("SELECT id, nombre FROM crm_grupos ORDER BY nombre").fetchall()
    db.close()
    return clientes, grupos


def guardar_contacto_crm(contacto_id):
    """Crea o actualiza un contacto y reemplaza sus asociaciones de clientes
    y grupos por las que vengan en el formulario. Regresa None si guardó
    bien, o un mensaje de error si faltó el nombre."""
    # Los contactos se capturan siempre en mayúsculas (salvo teléfono y
    # correo, que no siguen esa convención).
    nombre = request.form.get("nombre", "").strip().upper()
    apellido = request.form.get("apellido", "").strip().upper()
    telefono = request.form.get("telefono", "").strip()
    correo = request.form.get("correo", "").strip()
    observaciones = request.form.get("observaciones", "").strip().upper()
    grupo_nuevo = request.form.get("grupo_nuevo", "").strip().upper()
    clientes_folios = {int(v) for v in request.form.getlist("clientes") if v.strip().lstrip("-").isdigit()}
    grupos_ids = {int(v) for v in request.form.getlist("grupos") if v.strip().isdigit()}

    if not nombre:
        return "El nombre es obligatorio."

    db = get_db()
    if grupo_nuevo:
        fila_grupo = db.execute(
            "INSERT INTO crm_grupos (nombre) VALUES (%s) "
            "ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre RETURNING id",
            (grupo_nuevo,),
        ).fetchone()
        grupos_ids.add(fila_grupo["id"])

    if contacto_id is None:
        fila = db.execute(
            "INSERT INTO crm_contactos (nombre, apellido, telefono, correo, observaciones) "
            "VALUES (%s, %s, %s, %s, %s) RETURNING id",
            (nombre, apellido, telefono, correo, observaciones),
        ).fetchone()
        contacto_id = fila["id"]
    else:
        db.execute(
            "UPDATE crm_contactos SET nombre = %s, apellido = %s, telefono = %s, correo = %s, observaciones = %s WHERE id = %s",
            (nombre, apellido, telefono, correo, observaciones, contacto_id),
        )
        db.execute("DELETE FROM crm_contacto_clientes WHERE contacto_id = %s", (contacto_id,))
        db.execute("DELETE FROM crm_contacto_grupos WHERE contacto_id = %s", (contacto_id,))

    for folio in clientes_folios:
        db.execute(
            "INSERT INTO crm_contacto_clientes (contacto_id, cliente_folio) VALUES (%s, %s) ON CONFLICT DO NOTHING",
            (contacto_id, folio),
        )
    for grupo_id in grupos_ids:
        db.execute(
            "INSERT INTO crm_contacto_grupos (contacto_id, grupo_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
            (contacto_id, grupo_id),
        )
    db.commit()
    db.close()
    return None


VENCIMIENTO_DIAS = ("15", "30", "45", "60")


def generar_id_cotizacion(db):
    """ID de 12 caracteres con prefijo 'COT-' (COT- + 8 dígitos), único en
    crm_cotizaciones (reintenta si choca)."""
    for _ in range(20):
        candidato = f"COT-{random.randint(0, 99999999):08d}"
        if not db.execute("SELECT 1 FROM crm_cotizaciones WHERE id_cotizacion = %s", (candidato,)).fetchone():
            return candidato
    raise RuntimeError("No se pudo generar un ID de cotización único.")


def generar_referencia_solicitud_maritimo(db):
    """Referencia secuencial 'COT-5000', 'COT-5001', ... para
    crm_solicitudes_maritimo_aereo (arranca en 5000)."""
    fila = db.execute(
        "SELECT referencia FROM crm_solicitudes_maritimo_aereo ORDER BY id DESC LIMIT 1"
    ).fetchone()
    siguiente = int(fila["referencia"].split("-")[-1]) + 1 if fila and fila["referencia"] else 5000
    return f"COT-{siguiente}"


def construir_cotizaciones_crm(plazas_permitidas=None):
    """Cotizaciones con su cliente (o prospecto) resuelto. Una cotización con
    cliente real hereda su restricción de plaza (igual que Clientes/Contactos);
    una de prospecto no tiene plaza que restringir, así que es visible para
    todos los que puedan ver el CRM."""
    db = get_db()
    plaza_por_vendedor = {}
    for r in db.execute("SELECT vendedor, plaza FROM catalogo_vendedores"):
        plaza_por_vendedor[normalizar(r["vendedor"])] = r["plaza"]

    filas = db.execute("""
        SELECT co.id, co.id_cotizacion, co.nombre_cotizacion, co.fecha_creacion, co.fecha_vencimiento, co.vencimiento_modo,
               co.cliente_folio, co.cliente_prospecto, ac.razon_social AS cliente_nombre, ac.vendedor AS cliente_vendedor,
               co.contacto_id, ct.nombre AS contacto_nombre, ct.apellido AS contacto_apellido,
               co.origen, co.destino, co.hazmat, co.hazmat_clase, co.hazmat_un_imo,
               i.nombre AS incoterm, m.nombre AS modalidad,
               co.estibable, co.tiempo_traslado, co.via, co.seguro_mercancia,
               co.profit_estimado, co.tipo_cambio, co.descripcion, co.estatus,
               sp.estado_mas_reciente AS pricing_estado_mas_reciente
        FROM crm_cotizaciones co
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN crm_contactos ct ON ct.id = co.contacto_id
        LEFT JOIN crm_incoterms i ON i.id = co.incoterm_id
        LEFT JOIN crm_modalidades m ON m.id = co.modalidad_id
        LEFT JOIN LATERAL (
            SELECT estado AS estado_mas_reciente
            FROM crm_solicitudes_maritimo_aereo
            WHERE cotizacion_id = co.id
            ORDER BY creado_en DESC
            LIMIT 1
        ) sp ON true
        ORDER BY co.fecha_creacion DESC, co.id DESC
    """).fetchall()

    totales_por_moneda = {}
    for r in db.execute("""
        SELECT cotizacion_id, moneda,
               SUM(cantidad * precio_unitario) AS subtotal,
               SUM(cantidad * precio_unitario * impuesto / 100.0) AS impuesto
        FROM crm_cotizacion_productos GROUP BY cotizacion_id, moneda
    """):
        totales_por_moneda.setdefault(r["cotizacion_id"], {})[r["moneda"]] = {
            "subtotal": float(r["subtotal"]),
            "impuesto": float(r["impuesto"] or 0),
            "total": float(r["subtotal"]) + float(r["impuesto"] or 0),
        }

    cotizaciones_con_booking = {
        r["cotizacion_id"] for r in db.execute("SELECT DISTINCT cotizacion_id FROM crm_cotizacion_bookings")
    }
    db.close()
    hoy = datetime.now(TZ_LOCAL).date()

    resultado = []
    for r in filas:
        es_prospecto = r["cliente_folio"] is None
        if not es_prospecto:
            plaza = plaza_por_vendedor.get(normalizar(r["cliente_vendedor"]), "#N/D")
            if plazas_permitidas is not None and plaza not in plazas_permitidas:
                continue
            cliente_texto = r["cliente_nombre"] or "#N/D"
        else:
            cliente_texto = f"Prospecto: {r['cliente_prospecto']}" if r["cliente_prospecto"] else "Prospecto"

        contacto_texto = f"{r['contacto_nombre']} {r['contacto_apellido'] or ''}".strip() if r["contacto_id"] else ""

        pricing_estado = (
            "respondida" if r["pricing_estado_mas_reciente"] in ("Cotizado", "Rechazada")
            else "pendiente" if r["pricing_estado_mas_reciente"]
            else None
        )
        resultado.append({
            "id": r["id"],
            "pricing_estado": pricing_estado,
            "id_cotizacion": r["id_cotizacion"],
            "nombre_cotizacion": r["nombre_cotizacion"] or "",
            "cliente": cliente_texto,
            "es_prospecto": es_prospecto,
            "contacto": contacto_texto,
            "fecha_creacion": r["fecha_creacion"].strftime("%Y-%m-%d") if r["fecha_creacion"] else "",
            "fecha_vencimiento": r["fecha_vencimiento"].strftime("%Y-%m-%d") if r["fecha_vencimiento"] else "",
            "origen": r["origen"] or "",
            "destino": r["destino"] or "",
            "hazmat": r["hazmat"],
            "hazmat_clase": r["hazmat_clase"] or "",
            "hazmat_un_imo": r["hazmat_un_imo"] or "",
            "incoterm": r["incoterm"] or "",
            "modalidad": r["modalidad"] or "",
            "estibable": r["estibable"],
            "tiempo_traslado": r["tiempo_traslado"] or "",
            "via": r["via"] or "",
            "seguro_mercancia": r["seguro_mercancia"],
            "profit_estimado": float(r["profit_estimado"]) if r["profit_estimado"] is not None else None,
            "tipo_cambio": float(r["tipo_cambio"]) if r["tipo_cambio"] is not None else None,
            "descripcion": r["descripcion"] or "",
            "estatus": calcular_estatus_cotizacion(r["estatus"], r["fecha_vencimiento"], r["id"] in cotizaciones_con_booking, hoy),
            "totales_moneda": totales_por_moneda.get(r["id"], {}),
            "gran_total_texto": " / ".join(
                f"{moneda} ${datos['total']:,.2f}"
                for moneda, datos in totales_por_moneda.get(r["id"], {}).items()
            ),
        })
    return resultado


TITULOS_RE = re.compile(r"^(LIC|ING|MTRO|MTRA|DR|DRA|C\.P|CP)\.?\s+", re.IGNORECASE)


def quitar_titulo(nombre):
    """'Lic. Marielbis Camacaro' -> 'Marielbis Camacaro', para poder cruzar
    el nombre de una firma con el nombre de vendedor del catálogo."""
    return TITULOS_RE.sub("", (nombre or "").strip())


ETIQUETA_PERIODO_ANTERIOR = {
    "hoy": "hace una semana", "semana": "la semana pasada", "mes": "el mes pasado", "personalizado": "el periodo anterior",
}


def rango_periodo_crm(periodo, hoy, fecha_inicio_custom=None, fecha_fin_custom=None):
    """Regresa (fecha_inicio, fecha_fin) como date para el periodo elegido
    en CRM → Inicio. 'personalizado' usa las fechas que venga del filtro,
    saneadas a un rango válido (inicio <= fin)."""
    if periodo == "hoy":
        return hoy, hoy
    if periodo == "semana":
        return hoy - timedelta(days=hoy.weekday()), hoy
    if periodo == "personalizado" and fecha_inicio_custom and fecha_fin_custom:
        ini, fin = fecha_inicio_custom, fecha_fin_custom
        return (ini, fin) if ini <= fin else (fin, ini)
    return hoy.replace(day=1), hoy


def construir_inicio_crm(periodo, fecha_inicio, fecha_fin, plaza_filtro, vendedor_filtro, plazas_permitidas=None):
    """Dashboard de CRM → Inicio: actividad de cotización (crm_cotizaciones)
    y de cierre (reporte_bookings) del periodo elegido, comparada contra el
    mismo tramo del periodo anterior; más una tendencia diaria fija de los
    últimos 30 días y alertas de vencimiento. Respeta las mismas
    plazas_permitidas que el resto del CRM, y además Plaza/Vendedor del
    filtro. El cruce cotización↔vendedor es por nombre (firma capturada, o
    si no hay, un nombre derivado del correo de login) — es un cruce por
    mejor esfuerzo, no una relación garantizada en la base."""
    hoy = datetime.now(TZ_LOCAL).date()
    dias_periodo = (fecha_fin - fecha_inicio).days + 1
    if periodo == "mes":
        # "Mes a la fecha" se compara contra el mismo tramo de días del mes
        # calendario anterior (1-18 ago vs 1-18 jul), no contra una ventana
        # genérica de N días inmediatamente antes: esa ventana cruzaría a
        # mitad de julio y no correspondería a ningún periodo real.
        anio_ant = fecha_inicio.year if fecha_inicio.month > 1 else fecha_inicio.year - 1
        mes_ant = fecha_inicio.month - 1 if fecha_inicio.month > 1 else 12
        ultimo_dia_mes_ant = calendar.monthrange(anio_ant, mes_ant)[1]
        fecha_inicio_anterior = date(anio_ant, mes_ant, 1)
        fecha_fin_anterior = date(anio_ant, mes_ant, min(fecha_fin.day, ultimo_dia_mes_ant))
    elif periodo in ("hoy", "semana"):
        # Se compara contra el mismo tramo de la semana pasada (exactamente
        # 7 días antes), no contra los días inmediatamente anteriores: así
        # "Hoy" se compara contra el mismo día de la semana anterior, y la
        # semana laboral (lunes-viernes) contra la semana laboral previa,
        # sin cruzar al fin de semana.
        fecha_inicio_anterior = fecha_inicio - timedelta(days=7)
        fecha_fin_anterior = fecha_fin - timedelta(days=7)
    else:
        fecha_fin_anterior = fecha_inicio - timedelta(days=1)
        fecha_inicio_anterior = fecha_fin_anterior - timedelta(days=dias_periodo - 1)
    ventana_inicio = fecha_inicio_anterior
    ventana_fin = fecha_fin

    db = get_db()
    vendedores_catalogo = db.execute("SELECT vendedor, plaza FROM catalogo_vendedores ORDER BY vendedor").fetchall()
    plaza_por_vendedor = {normalizar(r["vendedor"]): r["plaza"] for r in vendedores_catalogo}
    desarrolladores_catalogo = db.execute("SELECT desarrollador, plaza FROM catalogo_desarrolladores ORDER BY desarrollador").fetchall()
    plaza_por_desarrollador = {normalizar(r["desarrollador"]): r["plaza"] for r in desarrolladores_catalogo}

    bookings = db.execute(
        "SELECT vendedor, ejecutivo, fecha, venta, profit FROM reporte_bookings "
        "WHERE fecha >= %s AND fecha < (%s::date + interval '1 day')",
        (ventana_inicio.isoformat(), ventana_fin.isoformat()),
    ).fetchall()

    cotizaciones = db.execute("""
        SELECT co.id, co.id_cotizacion, co.fecha_creacion, co.fecha_vencimiento,
               co.cliente_folio, co.cliente_prospecto, co.nombre_cotizacion,
               co.estatus, co.perdida_en, cb.ganada_desde,
               ac.vendedor AS cliente_vendedor, ac.razon_social,
               f.nombre_firma, cu.email AS creador_correo
        FROM crm_cotizaciones co
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN crm_firmas f ON f.user_id = co.creado_por_user_id
        LEFT JOIN auth.users cu ON cu.id = co.creado_por_user_id
        LEFT JOIN (
            SELECT cotizacion_id, MIN(aplicado_en) AS ganada_desde
            FROM crm_cotizacion_bookings GROUP BY cotizacion_id
        ) cb ON cb.cotizacion_id = co.id
    """).fetchall()
    db.close()

    plaza_filtro = plaza_filtro or ""
    vendedor_filtro_norm = normalizar(vendedor_filtro) if vendedor_filtro else ""

    filas_booking = []
    for r in bookings:
        fecha = r["fecha"]
        if fecha is None:
            continue
        d = fecha.astimezone(TZ_LOCAL).date()
        vendedor = r["vendedor"] or "#N/D"
        plaza = plaza_por_vendedor.get(normalizar(vendedor), "#N/D")
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        if plaza_filtro and plaza != plaza_filtro:
            continue
        if vendedor_filtro_norm and normalizar(vendedor) != vendedor_filtro_norm:
            continue
        ejecutivo = normalizar(r["ejecutivo"]) or None
        filas_booking.append({
            "d": d, "vendedor": vendedor, "plaza": plaza,
            "ejecutivo": r["ejecutivo"] if ejecutivo else None,
            "plaza_desarrollador": plaza_por_desarrollador.get(ejecutivo, "#N/D") if ejecutivo else None,
            "venta": float(r["venta"] or 0), "profit": float(r["profit"] or 0),
        })

    filas_cot = []
    for r in cotizaciones:
        d = r["fecha_creacion"]
        if d is None:
            continue
        if r["cliente_folio"] is not None:
            plaza = plaza_por_vendedor.get(normalizar(r["cliente_vendedor"]), "#N/D")
        else:
            plaza = None
        if plazas_permitidas is not None and plaza not in plazas_permitidas:
            continue
        if plaza_filtro and plaza != plaza_filtro:
            continue
        identidad_mostrar = quitar_titulo(r["nombre_firma"]) or nombre_desde_correo(r["creador_correo"])
        identidad = normalizar(identidad_mostrar)
        if vendedor_filtro_norm and identidad != vendedor_filtro_norm:
            continue
        filas_cot.append({
            "d": d, "id": r["id"], "id_cotizacion": r["id_cotizacion"],
            "cliente": r["razon_social"] or r["cliente_prospecto"] or "Prospecto",
            "nombre_cotizacion": r["nombre_cotizacion"] or "",
            "fecha_vencimiento": r["fecha_vencimiento"],
            "identidad": identidad, "identidad_mostrar": identidad_mostrar or "#N/D",
            "ganada_desde": r["ganada_desde"].astimezone(TZ_LOCAL).date() if r["ganada_desde"] else None,
            "perdida_desde": r["perdida_en"].astimezone(TZ_LOCAL).date() if r["estatus"] == "perdida" and r["perdida_en"] else None,
        })

    def en_rango(filas, ini, fin):
        return [f for f in filas if ini <= f["d"] <= fin]

    def en_rango_campo(filas, campo, ini, fin):
        return [f for f in filas if f[campo] is not None and ini <= f[campo] <= fin]

    booking_periodo = en_rango(filas_booking, fecha_inicio, fecha_fin)
    booking_anterior = en_rango(filas_booking, fecha_inicio_anterior, fecha_fin_anterior)
    cot_periodo = en_rango(filas_cot, fecha_inicio, fecha_fin)
    cot_anterior = en_rango(filas_cot, fecha_inicio_anterior, fecha_fin_anterior)

    # Ganadas/Perdidas se cuentan por cuándo pasó eso (primer booking
    # aplicado / cuándo se marcó perdida), no por cuándo se creó la
    # cotización — es una métrica de "actividad de cierre" del periodo,
    # igual que Venta/Profit ya lo son para bookings.
    ganadas_periodo = en_rango_campo(filas_cot, "ganada_desde", fecha_inicio, fecha_fin)
    ganadas_anterior = en_rango_campo(filas_cot, "ganada_desde", fecha_inicio_anterior, fecha_fin_anterior)
    perdidas_periodo = en_rango_campo(filas_cot, "perdida_desde", fecha_inicio, fecha_fin)
    perdidas_anterior = en_rango_campo(filas_cot, "perdida_desde", fecha_inicio_anterior, fecha_fin_anterior)

    def delta_pct(actual, anterior):
        if not anterior:
            return None
        return (actual - anterior) / anterior * 100

    venta_periodo = sum(f["venta"] for f in booking_periodo)
    venta_anterior = sum(f["venta"] for f in booking_anterior)
    profit_periodo = sum(f["profit"] for f in booking_periodo)
    profit_anterior = sum(f["profit"] for f in booking_anterior)

    kpis = {
        "cotizaciones": len(cot_periodo), "cotizaciones_delta": delta_pct(len(cot_periodo), len(cot_anterior)),
        "cotizaciones_anterior": len(cot_anterior),
        "ganadas": len(ganadas_periodo), "ganadas_delta": delta_pct(len(ganadas_periodo), len(ganadas_anterior)),
        "ganadas_anterior": len(ganadas_anterior),
        "perdidas": len(perdidas_periodo), "perdidas_delta": delta_pct(len(perdidas_periodo), len(perdidas_anterior)),
        "perdidas_anterior": len(perdidas_anterior),
        "bookings": len(booking_periodo), "bookings_delta": delta_pct(len(booking_periodo), len(booking_anterior)),
        "bookings_anterior": len(booking_anterior),
        "venta": venta_periodo, "venta_delta": delta_pct(venta_periodo, venta_anterior), "venta_anterior": venta_anterior,
        "profit": profit_periodo, "profit_delta": delta_pct(profit_periodo, profit_anterior), "profit_anterior": profit_anterior,
    }

    dias = [fecha_inicio + timedelta(days=i) for i in range((fecha_fin - fecha_inicio).days + 1)]
    profit_por_dia, cot_por_dia = {}, {}
    for f in booking_periodo:
        profit_por_dia[f["d"]] = profit_por_dia.get(f["d"], 0) + f["profit"]
    for f in cot_periodo:
        cot_por_dia[f["d"]] = cot_por_dia.get(f["d"], 0) + 1
    serie = [{"fecha": d.strftime("%d/%m"), "profit": round(profit_por_dia.get(d, 0)), "cotizaciones": cot_por_dia.get(d, 0)} for d in dias]

    resumen_vendedor = {}
    for f in booking_periodo:
        key = normalizar(f["vendedor"])
        fila = resumen_vendedor.setdefault(key, {"nombre": f["vendedor"], "plaza": f["plaza"], "bookings": 0, "venta": 0.0, "profit": 0.0, "cotizaciones": 0})
        fila["bookings"] += 1
        fila["venta"] += f["venta"]
        fila["profit"] += f["profit"]

    # Actividad por desarrollador: mismo patrón que Comisiones — se agrupa
    # por reporte_bookings.ejecutivo (no por vendedor). El permiso de plaza
    # ya se aplicó arriba usando la plaza del VENDEDOR (regla existente en
    # Comisiones: el desarrollador no basta para decidir visibilidad), así
    # que aquí solo agrupamos lo que ya pasó ese filtro.
    resumen_desarrollador = {}
    for f in booking_periodo:
        if not f["ejecutivo"]:
            continue
        key = normalizar(f["ejecutivo"])
        fila = resumen_desarrollador.setdefault(key, {"nombre": f["ejecutivo"], "plaza": f["plaza_desarrollador"], "bookings": 0, "venta": 0.0, "profit": 0.0, "cotizaciones": 0})
        fila["bookings"] += 1
        fila["venta"] += f["venta"]
        fila["profit"] += f["profit"]

    # Las cotizaciones se cruzan por nombre contra quien ya aparece en Vendedor
    # o en Desarrollador (para no duplicar a la misma persona en las dos
    # tablas); solo si no coincide con ninguna de las dos cae como fila
    # aparte en Vendedor, para que su actividad no quede invisible.
    cot_por_identidad = {}
    for f in cot_periodo:
        # Antes, una cotización sin firma capturada NI creador identificado
        # (sesión vieja sin usuario_id — ver sesion_activa) se saltaba por
        # completo aquí: contaba en "Cotizaciones creadas" pero desaparecía
        # de este desglose, así que la suma nunca cuadraba con el total.
        # Ahora esas caen en un bucket "#N/D" visible en vez de perderse.
        clave = f["identidad"] or "#N/D"
        nombre = f["identidad_mostrar"] or "Sin creador registrado"
        cot_por_identidad.setdefault(clave, {"nombre": nombre, "n": 0})
        cot_por_identidad[clave]["n"] += 1
    for key, datos in cot_por_identidad.items():
        if key in resumen_vendedor:
            resumen_vendedor[key]["cotizaciones"] = datos["n"]
        elif key in resumen_desarrollador:
            resumen_desarrollador[key]["cotizaciones"] = datos["n"]
        else:
            resumen_vendedor[key] = {"nombre": datos["nombre"], "plaza": "—", "bookings": 0, "venta": 0.0, "profit": 0.0, "cotizaciones": datos["n"]}

    ranking = sorted(resumen_vendedor.values(), key=lambda r: (-r["profit"], -r["cotizaciones"]))
    ranking_desarrolladores = sorted(resumen_desarrollador.values(), key=lambda r: (-r["profit"], -r["cotizaciones"]))

    por_vencer, vencidas = [], []
    for f in filas_cot:
        fv = f["fecha_vencimiento"]
        if fv is None or f["ganada_desde"] is not None or f["perdida_desde"] is not None:
            continue
        if fv < hoy:
            vencidas.append(f)
        elif fv <= hoy + timedelta(days=14):
            por_vencer.append({**f, "dias": (fv - hoy).days})
    por_vencer.sort(key=lambda f: f["fecha_vencimiento"])
    vencidas.sort(key=lambda f: f["fecha_vencimiento"])

    todas_las_plazas = {p for p in plaza_por_vendedor.values() if p}
    plazas_opciones = sorted(todas_las_plazas if plazas_permitidas is None else todas_las_plazas & plazas_permitidas)
    vendedores_opciones = sorted({
        r["vendedor"] for r in vendedores_catalogo
        if r["vendedor"] and (plazas_permitidas is None or r["plaza"] in plazas_permitidas)
        and (not plaza_filtro or r["plaza"] == plaza_filtro)
    })

    return {
        "kpis": kpis,
        "serie": serie,
        # Sin límite: antes se cortaba a los primeros 12 por profit, lo que
        # escondía a quien solo tiene cotizaciones sin bookings todavía
        # (profit $0) — la suma de la columna Cotizaciones nunca cuadraba
        # con el KPI "Cotizaciones creadas" de arriba.
        "ranking": ranking,
        "ranking_desarrolladores": ranking_desarrolladores,
        "por_vencer": por_vencer[:8],
        "vencidas": vencidas[:8],
        "plazas_opciones": plazas_opciones,
        "vendedores_opciones": vendedores_opciones,
        "fecha_inicio_larga": fecha_larga_es(fecha_inicio),
        "fecha_fin_larga": fecha_larga_es(fecha_fin),
        "fecha_inicio_anterior_larga": fecha_larga_es(fecha_inicio_anterior),
        "fecha_fin_anterior_larga": fecha_larga_es(fecha_fin_anterior),
        "etiqueta_anterior": ETIQUETA_PERIODO_ANTERIOR.get(periodo, "el periodo anterior"),
    }


def obtener_lineas_cotizacion_crm(cotizacion_id):
    db = get_db()
    filas = db.execute("""
        SELECT lp.producto_id, lp.producto_texto, p.nombre AS producto_nombre, lp.cantidad, lp.precio_unitario,
               lp.moneda, lp.causa_impuesto, lp.impuesto, lp.observaciones
        FROM crm_cotizacion_productos lp
        LEFT JOIN crm_tipos_ingreso_egreso p ON p.id = lp.producto_id
        WHERE lp.cotizacion_id = %s
        ORDER BY lp.orden
    """, (cotizacion_id,)).fetchall()
    db.close()
    return [{
        "producto_id": r["producto_id"],
        "producto_texto": r["producto_texto"],
        "producto_nombre": r["producto_texto"] or r["producto_nombre"] or "",
        "cantidad": float(r["cantidad"]),
        "precio_unitario": float(r["precio_unitario"]),
        "moneda": r["moneda"] or "MXN",
        "causa_impuesto": r["causa_impuesto"],
        "impuesto": float(r["impuesto"] or 0),
        "observaciones": r["observaciones"] or "",
    } for r in filas]


def nombre_desde_correo(correo):
    """Deriva un nombre legible de un correo tipo nombre.apellido@dominio
    ("marielbis.camacaro@av2logistics.com" -> "Marielbis Camacaro"), para
    cuando el creador de una cotización no tiene firma capturada todavía."""
    local = (correo or "").split("@")[0]
    partes = [p for p in re.split(r"[._-]+", local) if p]
    return " ".join(p.capitalize() for p in partes)


FIRMA_DEFAULT = {
    "nombre_firma": "Lic Armando Villanueva Silva",
    "puesto": "Ejecutivo de Cuenta Monterrey",
    "telefono": "+52 81 1778 7250",
    "correo": "avillanueva@av2logistics.com",
}


def calcular_estatus_cotizacion(estatus_guardado, fecha_vencimiento, tiene_bookings, hoy=None):
    """El estatus visible de una cotización combina lo guardado con lo que
    se puede derivar: Ganada gana sobre todo (tiene ≥1 booking aplicado,
    sin importar si ya venció o se había marcado perdida por error);
    Perdida es la única acción manual real (crm_cotizaciones.estatus);
    Vencido/Vigente se calculan comparando fecha_vencimiento con hoy —
    nunca se guardan en la base."""
    if tiene_bookings:
        return "ganada"
    if estatus_guardado == "perdida":
        return "perdida"
    hoy = hoy or datetime.now(TZ_LOCAL).date()
    if fecha_vencimiento and fecha_vencimiento < hoy:
        return "vencido"
    return "vigente"


def obtener_bookings_disponibles_cliente(cliente_nombre):
    """Bookings reales (reporte_bookings) del cliente de esta cotización,
    casados por nombre (mismo criterio que 'tiene_booking' en Clientes),
    para el selector "Aplicar a booking" en el detalle. Excluye bookings
    que ya están aplicados a CUALQUIER cotización (la restricción unique
    de crm_cotizacion_bookings.booking_referencia ya lo impediría, esto
    solo evita ofrecerlo de entrada en el selector)."""
    if not cliente_nombre:
        return []
    db = get_db()
    filas = db.execute("""
        SELECT rb.referencia, rb.fecha, rb.venta
        FROM reporte_bookings rb
        WHERE upper(trim(rb.cliente_servicio)) = upper(trim(%s))
          AND NOT EXISTS (SELECT 1 FROM crm_cotizacion_bookings cb WHERE cb.booking_referencia = rb.referencia)
        ORDER BY rb.fecha DESC LIMIT 200
    """, (cliente_nombre,)).fetchall()
    db.close()
    return [
        {"referencia": r["referencia"], "fecha": r["fecha"].strftime("%Y-%m-%d") if r["fecha"] else "",
         "venta": float(r["venta"] or 0)}
        for r in filas if r["referencia"]
    ]


def construir_documento_cotizacion_crm(cotizacion_id):
    """Junta toda la información de una cotización (cliente/prospecto,
    contacto, catálogos resueltos, líneas de producto con su total) para
    la vista imprimible."""
    db = get_db()
    fila = db.execute("""
        SELECT co.*, ac.razon_social AS cliente_nombre, ac.vendedor AS cliente_vendedor,
               ct.nombre AS contacto_nombre, ct.apellido AS contacto_apellido,
               ct.telefono AS contacto_telefono, ct.correo AS contacto_correo,
               i.nombre AS incoterm, m.nombre AS modalidad,
               f.nombre_firma, f.puesto AS firma_puesto, f.telefono AS firma_telefono, f.correo AS firma_correo,
               cu.email AS creador_correo, mp.nombre AS motivo_perdida_nombre
        FROM crm_cotizaciones co
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN crm_contactos ct ON ct.id = co.contacto_id
        LEFT JOIN crm_incoterms i ON i.id = co.incoterm_id
        LEFT JOIN crm_modalidades m ON m.id = co.modalidad_id
        LEFT JOIN crm_firmas f ON f.user_id = co.creado_por_user_id
        LEFT JOIN auth.users cu ON cu.id = co.creado_por_user_id
        LEFT JOIN crm_motivos_perdida mp ON mp.id = co.motivo_perdida_id
        WHERE co.id = %s
    """, (cotizacion_id,)).fetchone()
    if fila is None:
        db.close()
        return None
    if not usuario_puede_ver_cotizacion(db, fila["cliente_folio"]):
        db.close()
        return None

    bookings_aplicados = db.execute("""
        SELECT cb.booking_referencia AS referencia, cb.aplicado_en, rb.fecha, rb.venta
        FROM crm_cotizacion_bookings cb
        LEFT JOIN reporte_bookings rb ON rb.referencia = cb.booking_referencia
        WHERE cb.cotizacion_id = %s ORDER BY cb.aplicado_en
    """, (cotizacion_id,)).fetchall()
    motivos_perdida = db.execute("SELECT id, nombre FROM crm_motivos_perdida ORDER BY nombre").fetchall()
    solicitudes_maritimo_raw = db.execute("""
        SELECT
            s.id, s.referencia, s.tipo_embarque, s.estado, s.fecha_creacion, s.creado_por,
            s.creado_en AS solicitud_en, s.visto_por_vendedor_en,
            r.ultima_respuesta_en
        FROM crm_solicitudes_maritimo_aereo s
        LEFT JOIN LATERAL (
            SELECT max(creado_en) AS ultima_respuesta_en
            FROM crm_solicitudes_pricing_respuestas
            WHERE solicitud_id = s.id
        ) r ON true
        WHERE s.cotizacion_id = %s
        ORDER BY s.creado_en DESC
    """, (cotizacion_id,)).fetchall()
    db.close()

    solicitudes_maritimo = []
    for s in solicitudes_maritimo_raw:
        s = dict(s)
        s["fecha_entrega"] = s["ultima_respuesta_en"]
        s["diferencia"] = (
            formatear_duracion(s["ultima_respuesta_en"] - s["solicitud_en"])
            if s["ultima_respuesta_en"] else None
        )
        s["es_nuevo"] = bool(
            s["ultima_respuesta_en"]
            and (not s["visto_por_vendedor_en"] or s["visto_por_vendedor_en"] < s["ultima_respuesta_en"])
        )
        solicitudes_maritimo.append(s)

    # El nombre de quien creó la cotización manda sobre el vendedor asignado
    # al cliente: usa la firma capturada si existe, si no deriva un nombre
    # legible del correo de login (p.ej. "marielbis.camacaro@..." →
    # "Marielbis Camacaro"). Solo cae al vendedor del cliente en cotizaciones
    # viejas, de antes de que existiera creado_por_user_id.
    nombre_creador = fila["nombre_firma"] or nombre_desde_correo(fila["creador_correo"])

    if fila["cliente_folio"] is not None:
        cliente_nombre = fila["cliente_nombre"] or "#N/D"
    else:
        cliente_nombre = fila["cliente_prospecto"] or "Prospecto"
    vendedor = nombre_creador or (fila["cliente_vendedor"] or "") if fila["cliente_folio"] is not None else nombre_creador

    estatus = calcular_estatus_cotizacion(fila["estatus"], fila["fecha_vencimiento"], bool(bookings_aplicados))
    # Se puede seguir aplicando más bookings a una cotización ya Ganada
    # (varios embarques de un mismo cliente cuentan para la misma
    # cotización); solo se bloquea si está Perdida.
    bookings_disponibles = [] if estatus == "perdida" else obtener_bookings_disponibles_cliente(cliente_nombre)

    lineas_crudas = obtener_lineas_cotizacion_crm(cotizacion_id)
    lineas = []
    totales_por_moneda = {}
    for l in lineas_crudas:
        subtotal = l["cantidad"] * l["precio_unitario"]
        impuesto_monto = subtotal * (l["impuesto"] / 100) if l["causa_impuesto"] else 0
        lineas.append({
            "descripcion": l["producto_nombre"],
            "cantidad": l["cantidad"],
            "precio_unitario": l["precio_unitario"],
            "impuesto_pct": l["impuesto"] if l["causa_impuesto"] else 0,
            "impuesto": impuesto_monto,
            "total": subtotal + impuesto_monto,
            "moneda": l["moneda"],
            "observaciones": l["observaciones"],
        })
        t = totales_por_moneda.setdefault(l["moneda"], {"subtotal": 0.0, "impuesto": 0.0})
        t["subtotal"] += subtotal
        t["impuesto"] += impuesto_monto
    totales_moneda = [
        {"moneda": moneda, "subtotal": datos["subtotal"], "impuesto": datos["impuesto"],
         "total": datos["subtotal"] + datos["impuesto"]}
        for moneda, datos in sorted(totales_por_moneda.items())
    ]

    return {
        "id_cotizacion": fila["id_cotizacion"],
        "nombre_cotizacion": fila["nombre_cotizacion"] or "",
        "mostrar_columna_impuesto": fila["mostrar_columna_impuesto"],
        "mostrar_totales": fila["mostrar_totales"],
        "fecha_creacion": fila["fecha_creacion"],
        "fecha_vencimiento": fila["fecha_vencimiento"],
        "fecha_creacion_larga": fecha_larga_es(fila["fecha_creacion"]),
        "fecha_vencimiento_larga": fecha_larga_es(fila["fecha_vencimiento"]),
        "creado_en": fila["creado_en"],
        "cliente_nombre": cliente_nombre,
        "vendedor": vendedor,
        "contacto_nombre": f"{fila['contacto_nombre']} {fila['contacto_apellido'] or ''}".strip() if fila["contacto_id"] else "",
        "contacto_telefono": fila["contacto_telefono"] or "",
        "contacto_correo": fila["contacto_correo"] or "",
        "origen": fila["origen"] or "",
        "destino": fila["destino"] or "",
        "hazmat": fila["hazmat"],
        "incoterm": fila["incoterm"] or "",
        "modalidad": fila["modalidad"] or "",
        "estibable": fila["estibable"],
        "tiempo_traslado": fila["tiempo_traslado"] or "",
        "via": fila["via"] or "",
        "seguro_mercancia": fila["seguro_mercancia"],
        "descripcion": fila["descripcion"] or "",
        "lineas": lineas,
        "totales_moneda": totales_moneda,
        # Puesto y teléfono solo se muestran si la persona los capturó en su
        # firma (Catálogos → Firmas de Cotización): no se inventan. El correo
        # sí cae al correo real de login del creador si no capturó uno propio.
        # Si la cotización no tiene creador conocido (de antes de este
        # catálogo), se usa la firma genérica anterior como respaldo.
        "firma_nombre": nombre_creador or FIRMA_DEFAULT["nombre_firma"],
        "firma_puesto": fila["firma_puesto"] or (FIRMA_DEFAULT["puesto"] if not fila["creador_correo"] else ""),
        "firma_telefono": fila["firma_telefono"] or (FIRMA_DEFAULT["telefono"] if not fila["creador_correo"] else ""),
        "firma_correo": fila["firma_correo"] or fila["creador_correo"] or FIRMA_DEFAULT["correo"],
        "estatus": estatus,
        "motivo_perdida_nombre": fila["motivo_perdida_nombre"] or "",
        "comentario_perdida": fila["comentario_perdida"] or "",
        "perdida_en": fila["perdida_en"],
        "bookings_aplicados": [
            {"referencia": b["referencia"],
             "fecha": b["fecha"].strftime("%Y-%m-%d") if b["fecha"] else "",
             "venta": float(b["venta"]) if b["venta"] is not None else None}
            for b in bookings_aplicados
        ],
        "bookings_disponibles": bookings_disponibles,
        "motivos_perdida": [{"id": m["id"], "nombre": m["nombre"]} for m in motivos_perdida],
        "solicitudes_maritimo": [
            {"id": s["id"], "referencia": s["referencia"], "tipo_embarque": s["tipo_embarque"] or "",
             "estado": s["estado"], "fecha_creacion": s["fecha_creacion"], "creado_por": s["creado_por"] or "",
             "solicitud_en": s["solicitud_en"], "fecha_entrega": s["fecha_entrega"],
             "diferencia": s["diferencia"], "es_nuevo": s["es_nuevo"]}
            for s in solicitudes_maritimo
        ],
    }


def clonar_cotizacion_crm(cotizacion_id):
    """Duplica una cotización (encabezado + líneas de producto) con un ID y
    fecha de creación nuevos; si el vencimiento era a N días, se recalcula
    desde hoy. Regresa el `id` interno de la copia, o None si el original
    no existe."""
    db = get_db()
    original = db.execute("SELECT vencimiento_modo, fecha_vencimiento, cliente_folio FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)).fetchone()
    if original is None:
        db.close()
        return None
    if not usuario_puede_ver_cotizacion(db, original["cliente_folio"]):
        db.close()
        return None

    nuevo_id_cotizacion = generar_id_cotizacion(db)
    fecha_creacion = datetime.now(TZ_LOCAL).date()
    if original["vencimiento_modo"] in VENCIMIENTO_DIAS:
        fecha_vencimiento = (fecha_creacion + timedelta(days=int(original["vencimiento_modo"]))).isoformat()
    else:
        fecha_vencimiento = original["fecha_vencimiento"]

    fila_nueva = db.execute("""
        INSERT INTO crm_cotizaciones (
            id_cotizacion, nombre_cotizacion, fecha_creacion, fecha_vencimiento, vencimiento_modo,
            cliente_folio, cliente_prospecto, contacto_id,
            origen, destino, hazmat, hazmat_clase, hazmat_un_imo,
            incoterm_id, modalidad_id, tipo_ingreso_egreso_id,
            estibable, tiempo_traslado, via, seguro_mercancia,
            profit_estimado, tipo_cambio, descripcion
        )
        SELECT %s, nombre_cotizacion, %s, %s, vencimiento_modo,
               cliente_folio, cliente_prospecto, contacto_id,
               origen, destino, hazmat, hazmat_clase, hazmat_un_imo,
               incoterm_id, modalidad_id, tipo_ingreso_egreso_id,
               estibable, tiempo_traslado, via, seguro_mercancia,
               profit_estimado, tipo_cambio, descripcion
        FROM crm_cotizaciones WHERE id = %s
        RETURNING id
    """, (nuevo_id_cotizacion, fecha_creacion, fecha_vencimiento, cotizacion_id)).fetchone()
    nuevo_id = fila_nueva["id"]

    db.execute("""
        INSERT INTO crm_cotizacion_productos
            (cotizacion_id, producto_id, producto_texto, cantidad, precio_unitario, moneda, causa_impuesto, impuesto, orden)
        SELECT %s, producto_id, producto_texto, cantidad, precio_unitario, moneda, causa_impuesto, impuesto, orden
        FROM crm_cotizacion_productos WHERE cotizacion_id = %s
    """, (nuevo_id, cotizacion_id))

    db.commit()
    db.close()
    return nuevo_id


def opciones_incoterm_modalidad_crm():
    db = get_db()
    incoterms = db.execute("SELECT id, nombre FROM crm_incoterms ORDER BY nombre").fetchall()
    modalidades = db.execute("SELECT id, nombre FROM crm_modalidades ORDER BY nombre").fetchall()
    db.close()
    return incoterms, modalidades


def opciones_tipo_producto_crm():
    """Catálogo único Tipo Ingreso/Egreso: su nombre es lo que se usa como
    'Producto' tanto en el encabezado de la cotización como en cada línea.
    Solo se ofrecen los que sí generan ingreso (naturaleza INGRESO o AMBOS
    — AMBOS aplica tanto a ingreso como egreso) y que ya traen un Producto
    asignado (AEREO, ADUANA, MARITIMO, TERRESTRE, SEGURO) — es lo que
    alimenta el primer desplegable (Producto); el segundo (Nombre) se
    filtra en el navegador contra estos mismos datos."""
    db = get_db()
    tipos = db.execute("""
        SELECT id, nombre, producto FROM crm_tipos_ingreso_egreso
        WHERE naturaleza IN ('INGRESO', 'AMBOS') AND producto IS NOT NULL AND producto <> ''
        ORDER BY producto, nombre
    """).fetchall()
    db.close()
    return tipos


def opciones_clientes_cotizacion_crm():
    db = get_db()
    clientes = db.execute(
        "SELECT folio, razon_social FROM asignacion_de_clientes WHERE folio IS NOT NULL ORDER BY razon_social"
    ).fetchall()
    db.close()
    return clientes


def opciones_contactos_cotizacion_crm():
    db = get_db()
    contactos = db.execute(
        "SELECT id, nombre, apellido FROM crm_contactos ORDER BY nombre, apellido"
    ).fetchall()
    db.close()
    return contactos


def parse_numero_formato_miles(texto):
    texto = (texto or "").strip().replace(",", "")
    if not texto:
        return None, None
    try:
        return float(texto), None
    except ValueError:
        return None, "no es un número válido"


def guardar_cotizacion_crm(cotizacion_id):
    """Crea o actualiza una cotización. Regresa (None, cotizacion_id) si
    guardó bien (cotizacion_id es el nuevo id si se estaba creando), o
    (mensaje_de_error, None) si algo obligatorio faltó o un número vino mal."""
    nombre_cotizacion = request.form.get("nombre_cotizacion", "").strip()[:60]
    contacto_id_raw = request.form.get("contacto_id", "").strip()
    contacto_id = int(contacto_id_raw) if contacto_id_raw.isdigit() else None

    es_prospecto = request.form.get("es_prospecto") == "si"
    cliente_folio_raw = request.form.get("cliente_folio", "").strip()
    cliente_prospecto = request.form.get("cliente_prospecto", "").strip()[:120]
    if es_prospecto:
        if not cliente_prospecto:
            return "Captura el nombre del prospecto.", None
        cliente_folio = None
    else:
        if not cliente_folio_raw.isdigit():
            return "Selecciona un cliente (o marca Prospecto).", None
        cliente_folio = int(cliente_folio_raw)
        cliente_prospecto = None

    origen = request.form.get("origen", "").strip()[:50]
    destino = request.form.get("destino", "").strip()[:50]
    if not origen or not destino:
        return "Origen y Destino son obligatorios.", None

    hazmat = request.form.get("hazmat") == "si"
    hazmat_clase = request.form.get("hazmat_clase", "").strip()[:50] if hazmat else None
    hazmat_un_imo = request.form.get("hazmat_un_imo", "").strip()[:50] if hazmat else None

    incoterm_id = request.form.get("incoterm_id") or None
    incoterm_id = int(incoterm_id) if incoterm_id else None
    incoterm_nuevo = request.form.get("incoterm_nuevo", "").strip()
    modalidad_id = request.form.get("modalidad_id") or None
    modalidad_id = int(modalidad_id) if modalidad_id else None
    modalidad_nuevo = request.form.get("modalidad_nuevo", "").strip()

    estibable = request.form.get("estibable") == "si"
    seguro_mercancia = request.form.get("seguro_mercancia") == "si"
    tiempo_traslado = request.form.get("tiempo_traslado", "").strip()[:30]
    via = request.form.get("via", "").strip()[:60]
    descripcion = request.form.get("descripcion", "").strip()
    # Checkboxes reales (no <select>): si no vienen marcados, el campo ni
    # siquiera se manda con el formulario, así que ausente = False.
    mostrar_columna_impuesto = request.form.get("mostrar_columna_impuesto") == "si"
    mostrar_totales = request.form.get("mostrar_totales") == "si"

    vencimiento_modo = request.form.get("vencimiento_modo", "libre").strip()
    if vencimiento_modo not in VENCIMIENTO_DIAS + ("libre",):
        return "Fecha de vencimiento inválida.", None
    fecha_vencimiento_libre = fecha_valida_o_vacia(request.form.get("fecha_vencimiento_libre", ""))
    if vencimiento_modo == "libre" and not fecha_vencimiento_libre:
        return "Captura la fecha de vencimiento.", None

    profit_estimado, error_profit = parse_numero_formato_miles(request.form.get("profit_estimado", ""))
    if error_profit:
        return f"Profit estimado {error_profit}.", None
    tipo_cambio, error_tc = parse_numero_formato_miles(request.form.get("tipo_cambio", ""))
    if error_tc:
        return f"Tipo de cambio {error_tc}.", None

    db = get_db()
    if incoterm_nuevo:
        fila = db.execute(
            "INSERT INTO crm_incoterms (nombre) VALUES (%s) "
            "ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre RETURNING id",
            (incoterm_nuevo,),
        ).fetchone()
        incoterm_id = fila["id"]
    if modalidad_nuevo:
        fila = db.execute(
            "INSERT INTO crm_modalidades (nombre) VALUES (%s) "
            "ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre RETURNING id",
            (modalidad_nuevo,),
        ).fetchone()
        modalidad_id = fila["id"]
    if cotizacion_id is None:
        fecha_creacion = datetime.now(TZ_LOCAL).date()
    else:
        fila_actual = db.execute(
            "SELECT fecha_creacion FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)
        ).fetchone()
        fecha_creacion = fila_actual["fecha_creacion"]

    if vencimiento_modo == "libre":
        fecha_vencimiento = fecha_vencimiento_libre
    else:
        fecha_vencimiento = (fecha_creacion + timedelta(days=int(vencimiento_modo))).isoformat()

    if cotizacion_id is None:
        id_cotizacion = generar_id_cotizacion(db)
        creado_por_user_id = session.get("usuario_id") or None
        fila_nueva = db.execute("""
            INSERT INTO crm_cotizaciones (
                id_cotizacion, nombre_cotizacion, fecha_creacion, fecha_vencimiento, vencimiento_modo,
                cliente_folio, cliente_prospecto, contacto_id,
                origen, destino, hazmat, hazmat_clase, hazmat_un_imo,
                incoterm_id, modalidad_id,
                estibable, tiempo_traslado, via, seguro_mercancia,
                profit_estimado, tipo_cambio, descripcion, creado_por_user_id,
                mostrar_columna_impuesto, mostrar_totales
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            id_cotizacion, nombre_cotizacion, fecha_creacion, fecha_vencimiento, vencimiento_modo,
            cliente_folio, cliente_prospecto, contacto_id,
            origen, destino, hazmat, hazmat_clase, hazmat_un_imo,
            incoterm_id, modalidad_id,
            estibable, tiempo_traslado, via, seguro_mercancia,
            profit_estimado, tipo_cambio, descripcion, creado_por_user_id,
            mostrar_columna_impuesto, mostrar_totales,
        )).fetchone()
        cotizacion_id = fila_nueva["id"]
    else:
        db.execute("""
            UPDATE crm_cotizaciones SET
                nombre_cotizacion = %s, fecha_vencimiento = %s, vencimiento_modo = %s,
                cliente_folio = %s, cliente_prospecto = %s, contacto_id = %s,
                origen = %s, destino = %s, hazmat = %s, hazmat_clase = %s, hazmat_un_imo = %s,
                incoterm_id = %s, modalidad_id = %s,
                estibable = %s, tiempo_traslado = %s, via = %s,
                seguro_mercancia = %s, profit_estimado = %s, tipo_cambio = %s, descripcion = %s,
                mostrar_columna_impuesto = %s, mostrar_totales = %s
            WHERE id = %s
        """, (
            nombre_cotizacion, fecha_vencimiento, vencimiento_modo,
            cliente_folio, cliente_prospecto, contacto_id,
            origen, destino, hazmat, hazmat_clase, hazmat_un_imo,
            incoterm_id, modalidad_id,
            estibable, tiempo_traslado, via,
            seguro_mercancia, profit_estimado, tipo_cambio, descripcion,
            mostrar_columna_impuesto, mostrar_totales,
            cotizacion_id,
        ))

    guardar_lineas_cotizacion_crm(db, cotizacion_id)

    db.commit()
    db.close()
    return None, cotizacion_id


MONEDAS_VALIDAS = ("USD", "MXN", "EUR")


def guardar_lineas_cotizacion_crm(db, cotizacion_id):
    """Reemplaza las líneas de producto de una cotización con las que vengan
    en el formulario (arrays paralelos linea_producto_id / linea_producto_texto /
    linea_cantidad / linea_precio_unitario / linea_moneda, uno por fila
    agregada con '+ Agregar producto'). Una línea sin producto de catálogo
    usa linea_producto_texto ("Sin producto", texto libre en mayúsculas).
    Filas sin producto (ni catálogo ni texto) o sin cantidad/precio
    numéricos se ignoran."""
    productos = request.form.getlist("linea_producto_id")
    productos_texto = request.form.getlist("linea_producto_texto")
    cantidades = request.form.getlist("linea_cantidad")
    precios = request.form.getlist("linea_precio_unitario")
    monedas = request.form.getlist("linea_moneda")
    causa_impuestos = request.form.getlist("linea_causa_impuesto")
    impuestos = request.form.getlist("linea_impuesto")
    observaciones_lineas = request.form.getlist("linea_observaciones")

    db.execute("DELETE FROM crm_cotizacion_productos WHERE cotizacion_id = %s", (cotizacion_id,))
    orden = 0
    for i, (cantidad_raw, precio_raw) in enumerate(zip(cantidades, precios)):
        producto_id_raw = productos[i] if i < len(productos) else ""
        producto_texto = (productos_texto[i].strip().upper()[:100] if i < len(productos_texto) else "") or None
        producto_id = int(producto_id_raw) if producto_id_raw.isdigit() else None
        if producto_texto:
            producto_id = None
        elif producto_id is None:
            continue
        cantidad, _ = parse_numero_formato_miles(cantidad_raw)
        precio, _ = parse_numero_formato_miles(precio_raw)
        if cantidad is None or precio is None:
            continue
        moneda = monedas[i] if i < len(monedas) and monedas[i] in MONEDAS_VALIDAS else "MXN"
        causa_impuesto = i < len(causa_impuestos) and causa_impuestos[i] == "si"
        impuesto = 0.0
        if causa_impuesto:
            impuesto, _ = parse_numero_formato_miles(impuestos[i] if i < len(impuestos) else "")
            impuesto = impuesto or 0.0
        observaciones = (observaciones_lineas[i].strip()[:70] if i < len(observaciones_lineas) else "") or None
        db.execute(
            "INSERT INTO crm_cotizacion_productos "
            "(cotizacion_id, producto_id, producto_texto, cantidad, precio_unitario, moneda, causa_impuesto, impuesto, orden, observaciones) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (cotizacion_id, producto_id, producto_texto, cantidad, precio, moneda, causa_impuesto, impuesto, orden, observaciones),
        )
        orden += 1


def agrupar_nav_crm(slug_activo):
    """Arma la barra lateral del CRM agrupada (sin grupo, Catálogos, Administración),
    marcando como activo el item que corresponde a la sección actual."""
    grupos_orden = [None, "Catálogos", "Administración"]
    grupos = []
    for nombre_grupo in grupos_orden:
        items = [
            {**item, "activo": item["slug"] == slug_activo}
            for item in CRM_NAV if item["grupo"] == nombre_grupo
        ]
        if items:
            grupos.append({"nombre": nombre_grupo, "elementos": items})
    return grupos


@app.route("/crm")
@crm_required
def crm():
    return redirect(url_for("crm_seccion", slug="inicio"))


@app.route("/crm/<slug>")
@crm_required
def crm_seccion(slug):
    item = next((i for i in CRM_NAV if i["slug"] == slug), None)
    if item is None:
        return redirect(url_for("crm_seccion", slug="tareas"))

    nav_groups = agrupar_nav_crm(slug)
    if slug == "inicio":
        hoy = datetime.now(TZ_LOCAL).date()
        periodo = request.args.get("periodo", "semana")
        if periodo not in ("hoy", "semana", "mes", "personalizado"):
            periodo = "semana"
        fecha_inicio_custom = fecha_valida_o_vacia(request.args.get("fecha_inicio", ""))
        fecha_fin_custom = fecha_valida_o_vacia(request.args.get("fecha_fin", ""))
        fecha_inicio, fecha_fin = rango_periodo_crm(
            periodo, hoy,
            date.fromisoformat(fecha_inicio_custom) if fecha_inicio_custom else None,
            date.fromisoformat(fecha_fin_custom) if fecha_fin_custom else None,
        )
        plaza_filtro = request.args.get("plaza", "").strip()
        vendedor_filtro = request.args.get("vendedor", "").strip()
        datos = construir_inicio_crm(periodo, fecha_inicio, fecha_fin, plaza_filtro, vendedor_filtro, plazas_permitidas_usuario())
        return render_template(
            "crm_inicio.html", nav_groups=nav_groups, titulo_pagina=item["texto"],
            periodo=periodo, fecha_inicio=fecha_inicio.isoformat(), fecha_fin=fecha_fin.isoformat(),
            plaza_filtro=plaza_filtro, vendedor_filtro=vendedor_filtro,
            serie_json=json_para_js(datos["serie"]), **datos,
        )

    if slug == "tareas":
        hoy = datetime.now(TZ_LOCAL).date().isoformat()
        tareas = [{**t, "vencida": t["fecha_compromiso"] < hoy} for t in TAREAS_MOCK]
        return render_template("crm_tareas.html", nav_groups=nav_groups, titulo_pagina=item["texto"], tareas=tareas)

    if slug == "booking":
        hoy = datetime.now(TZ_LOCAL).date()
        fecha_inicio_default = hoy.replace(day=1).isoformat()
        fecha_fin_default = hoy.isoformat()
        fecha_inicio = fecha_valida_o_vacia(request.args.get("fecha_inicio", "")) or fecha_inicio_default
        fecha_fin = fecha_valida_o_vacia(request.args.get("fecha_fin", "")) or fecha_fin_default
        bookings, total_bookings = construir_booking_crm(plazas_permitidas_usuario(), fecha_inicio, fecha_fin)
        return render_template(
            "crm_booking.html", nav_groups=nav_groups, titulo_pagina=item["texto"],
            bookings=bookings, total_bookings=total_bookings,
            fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
            filtro_personalizado=(fecha_inicio != fecha_inicio_default or fecha_fin != fecha_fin_default),
        )

    if slug == "clientes":
        clientes = construir_clientes_crm(plazas_permitidas_usuario())
        return render_template("crm_clientes.html", nav_groups=nav_groups, titulo_pagina=item["texto"], clientes=clientes)

    if slug == "contactos":
        contactos = construir_contactos_crm(plazas_permitidas_usuario())
        return render_template("crm_contactos.html", nav_groups=nav_groups, titulo_pagina=item["texto"], contactos=contactos)

    if slug == "cotizaciones":
        cotizaciones = construir_cotizaciones_crm(plazas_permitidas_usuario())
        return render_template("crm_cotizaciones.html", nav_groups=nav_groups, titulo_pagina=item["texto"], cotizaciones=cotizaciones)

    return render_template("crm_placeholder.html", nav_groups=nav_groups, titulo_pagina=item["texto"])


@app.route("/crm/clientes/<int:folio>")
@crm_required
def crm_cliente_detalle(folio):
    cliente = construir_cliente_detalle_crm(folio, plazas_permitidas_usuario())
    if cliente is None:
        flash("Cliente no encontrado o sin permiso para verlo.")
        return redirect(url_for("crm_seccion", slug="clientes"))

    cotizaciones = construir_cotizaciones_resumen_crm(cliente_folio=folio)
    nav_groups = agrupar_nav_crm("clientes")
    return render_template(
        "crm_cliente_detalle.html", nav_groups=nav_groups,
        titulo_pagina=cliente["razon_social"], cliente=cliente, cotizaciones=cotizaciones,
    )


@app.route("/crm/contactos/nuevo", methods=["GET", "POST"])
@crm_required
def crm_contacto_nuevo():
    if request.method == "POST":
        error = guardar_contacto_crm(None)
        if error:
            flash(error)
        else:
            return redirect(url_for("crm_seccion", slug="contactos"))

    clientes, grupos = opciones_clientes_grupos_crm()
    nav_groups = agrupar_nav_crm("contactos")
    return render_template(
        "crm_contacto_form.html", nav_groups=nav_groups, titulo_pagina="Nuevo contacto",
        contacto=None, clientes=clientes, grupos=grupos, clientes_sel=set(), grupos_sel=set(),
    )


@app.route("/crm/contactos/<int:contacto_id>")
@crm_required
def crm_contacto_detalle(contacto_id):
    contacto = construir_contacto_detalle_crm(contacto_id, plazas_permitidas_usuario())
    if contacto is None:
        flash("Contacto no encontrado o sin permiso para verlo.")
        return redirect(url_for("crm_seccion", slug="contactos"))

    cotizaciones = construir_cotizaciones_resumen_crm(contacto_id=contacto_id)
    nav_groups = agrupar_nav_crm("contactos")
    nombre_completo = f"{contacto['nombre']} {contacto['apellido']}".strip()
    return render_template(
        "crm_contacto_detalle.html", nav_groups=nav_groups,
        titulo_pagina=nombre_completo, contacto=contacto, cotizaciones=cotizaciones,
    )


@app.route("/crm/contactos/<int:contacto_id>/editar", methods=["GET", "POST"])
@crm_required
def crm_contacto_editar(contacto_id):
    db = get_db()
    if not usuario_puede_ver_contacto(db, contacto_id):
        db.close()
        return "No encontrado", 404
    db.close()

    if request.method == "POST":
        error = guardar_contacto_crm(contacto_id)
        if error:
            flash(error)
        else:
            return redirect(url_for("crm_seccion", slug="contactos"))

    db = get_db()
    contacto = db.execute("SELECT * FROM crm_contactos WHERE id = %s", (contacto_id,)).fetchone()
    if contacto is None:
        db.close()
        return "No encontrado", 404
    clientes_sel = {
        r["cliente_folio"] for r in db.execute(
            "SELECT cliente_folio FROM crm_contacto_clientes WHERE contacto_id = %s", (contacto_id,)
        )
    }
    grupos_sel = {
        r["grupo_id"] for r in db.execute(
            "SELECT grupo_id FROM crm_contacto_grupos WHERE contacto_id = %s", (contacto_id,)
        )
    }
    db.close()

    clientes, grupos = opciones_clientes_grupos_crm()
    nav_groups = agrupar_nav_crm("contactos")
    return render_template(
        "crm_contacto_form.html", nav_groups=nav_groups, titulo_pagina="Editar contacto",
        contacto=contacto, clientes=clientes, grupos=grupos, clientes_sel=clientes_sel, grupos_sel=grupos_sel,
    )


@app.route("/crm/contactos/<int:contacto_id>/eliminar", methods=["POST"])
@crm_required
def crm_contacto_eliminar(contacto_id):
    db = get_db()
    if not usuario_puede_ver_contacto(db, contacto_id):
        db.close()
        return "No encontrado", 404
    db.execute("DELETE FROM crm_contactos WHERE id = %s", (contacto_id,))
    db.commit()
    db.close()
    return redirect(url_for("crm_seccion", slug="contactos"))


@app.route("/crm/cotizaciones/nueva", methods=["GET", "POST"])
@crm_required
def crm_cotizacion_nueva():
    if request.method == "POST":
        error, nuevo_id = guardar_cotizacion_crm(None)
        if error:
            flash(error)
        else:
            flash("Cotización guardada.")
            return redirect(url_for("crm_cotizacion_editar", cotizacion_id=nuevo_id))

    incoterms, modalidades = opciones_incoterm_modalidad_crm()
    tipos_ingreso_egreso = opciones_tipo_producto_crm()
    clientes = opciones_clientes_cotizacion_crm()
    contactos = opciones_contactos_cotizacion_crm()
    nav_groups = agrupar_nav_crm("cotizaciones")
    return render_template(
        "crm_cotizacion_form.html", nav_groups=nav_groups, titulo_pagina="Nueva cotización",
        cotizacion=None, incoterms=incoterms, modalidades=modalidades,
        clientes=clientes, contactos=contactos, lineas=[],
        productos_json=json_para_js([{"id": t["id"], "nombre": t["nombre"], "producto": t["producto"]} for t in tipos_ingreso_egreso]),
        lineas_json=json_para_js([]),
    )


@app.route("/crm/cotizaciones/<int:cotizacion_id>/editar", methods=["GET", "POST"])
@crm_required
def crm_cotizacion_editar(cotizacion_id):
    db = get_db()
    cotizacion = db.execute("SELECT * FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)).fetchone()
    if cotizacion is None:
        db.close()
        return "No encontrado", 404
    if not usuario_puede_ver_cotizacion(db, cotizacion["cliente_folio"]):
        db.close()
        return "No encontrado", 404
    tiene_bookings = db.execute(
        "SELECT 1 FROM crm_cotizacion_bookings WHERE cotizacion_id = %s LIMIT 1", (cotizacion_id,)
    ).fetchone() is not None
    db.close()
    estatus_actual = calcular_estatus_cotizacion(cotizacion["estatus"], cotizacion["fecha_vencimiento"], tiene_bookings)
    if estatus_actual in ("ganada", "perdida"):
        flash(f"Esta cotización ya está {estatus_actual} y no se puede editar. Puedes clonarla para crear una nueva.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    if request.method == "POST":
        error, _ = guardar_cotizacion_crm(cotizacion_id)
        if error:
            flash(error)
        else:
            flash("Cotización guardada.")
            return redirect(url_for("crm_cotizacion_editar", cotizacion_id=cotizacion_id))

    incoterms, modalidades = opciones_incoterm_modalidad_crm()
    tipos_ingreso_egreso = opciones_tipo_producto_crm()
    clientes = opciones_clientes_cotizacion_crm()
    contactos = opciones_contactos_cotizacion_crm()
    lineas = obtener_lineas_cotizacion_crm(cotizacion_id)
    nav_groups = agrupar_nav_crm("cotizaciones")
    return render_template(
        "crm_cotizacion_form.html", nav_groups=nav_groups,
        titulo_pagina=f"Cotización {cotizacion['id_cotizacion']}",
        cotizacion=cotizacion, incoterms=incoterms, modalidades=modalidades,
        clientes=clientes, contactos=contactos, lineas=lineas,
        productos_json=json_para_js([{"id": t["id"], "nombre": t["nombre"], "producto": t["producto"]} for t in tipos_ingreso_egreso]),
        lineas_json=json_para_js(lineas),
    )


@app.route("/crm/cotizaciones/<int:cotizacion_id>/eliminar", methods=["POST"])
@crm_required
def crm_cotizacion_eliminar(cotizacion_id):
    db = get_db()
    if not cotizacion_visible_para_usuario(db, cotizacion_id):
        db.close()
        return "No encontrado", 404
    db.execute("DELETE FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,))
    db.commit()
    db.close()
    return redirect(url_for("crm_seccion", slug="cotizaciones"))


@app.route("/crm/cotizaciones/<int:cotizacion_id>")
@crm_required
def crm_cotizacion_detalle(cotizacion_id):
    documento = construir_documento_cotizacion_crm(cotizacion_id)
    if documento is None:
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    nav_groups = agrupar_nav_crm("cotizaciones")
    return render_template(
        "crm_cotizacion_detalle.html", nav_groups=nav_groups,
        titulo_pagina=documento["nombre_cotizacion"] or f"Cotización {documento['id_cotizacion']}",
        doc=documento, cotizacion_id=cotizacion_id,
    )


@app.route("/crm/cotizaciones/<int:cotizacion_id>/solicitud-maritimo/nueva", methods=["GET", "POST"])
@crm_required
def crm_solicitud_maritimo_nueva(cotizacion_id):
    db = get_db()
    cotizacion = db.execute(
        "SELECT id, id_cotizacion, nombre_cotizacion, cliente_folio FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)
    ).fetchone()
    if cotizacion is None:
        db.close()
        return "No encontrado", 404
    if not usuario_puede_ver_cotizacion(db, cotizacion["cliente_folio"]):
        db.close()
        return "No encontrado", 404

    if request.method == "POST":
        def campo(nombre, limite=50):
            return (request.form.get(nombre, "") or "").strip()[:limite] or None

        def si_no(nombre):
            v = request.form.get(nombre)
            return True if v == "si" else (False if v == "no" else None)

        dias_libres_raw = request.form.get("fcl_dias_libres_requeridos", "").strip()
        dias_libres = int(dias_libres_raw) if dias_libres_raw.isdigit() else None
        incoterm_id_raw = request.form.get("incoterm_id", "").strip()
        incoterm_id = int(incoterm_id_raw) if incoterm_id_raw else None

        referencia = generar_referencia_solicitud_maritimo(db)
        db.execute("""
            INSERT INTO crm_solicitudes_maritimo_aereo (
                referencia, cotizacion_id, creado_por, importacion_exportacion, incoterm_id,
                tipo_embarque, pais_origen, pais_destino, lugar_recoleccion, puerto_carga,
                puerto_descarga, lugar_entrega, naviera_aerolinea, fcl_numero_tipo_contenedores,
                fcl_dias_libres_requeridos, producto, lcl_air_dimensiones, estibable,
                requiere_inbond_usa, hazmat, carga_reefer, temperatura, requerimientos_especiales,
                agente_a_cotizar, descripcion_material, anexos_notas, propiedad
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            referencia, cotizacion_id, session.get("usuario", ""),
            campo("importacion_exportacion"), incoterm_id,
            campo("tipo_embarque"), campo("pais_origen"), campo("pais_destino"),
            campo("lugar_recoleccion"), campo("puerto_carga"), campo("puerto_descarga"),
            campo("lugar_entrega"), campo("naviera_aerolinea"), campo("fcl_numero_tipo_contenedores"),
            dias_libres, campo("producto"), campo("lcl_air_dimensiones"),
            si_no("estibable"), si_no("requiere_inbond_usa"), si_no("hazmat") or False, si_no("carga_reefer"),
            campo("temperatura"), campo("requerimientos_especiales"), campo("agente_a_cotizar"),
            campo("descripcion_material"), (request.form.get("anexos_notas", "") or "").strip() or None,
            campo("propiedad"),
        ))
        db.commit()
        db.close()
        flash(f"Solicitud {referencia} enviada a Pricing.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    incoterms = db.execute("SELECT id, nombre FROM crm_incoterms ORDER BY nombre").fetchall()
    db.close()
    nav_groups = agrupar_nav_crm("cotizaciones")
    return render_template(
        "crm_solicitud_maritimo_form.html", nav_groups=nav_groups,
        titulo_pagina="Solicitud a Pricing · Marítimo / Aéreo",
        cotizacion=cotizacion, incoterms=incoterms,
    )


ESTADOS_SOLICITUD_PRICING = ["Solicitud", "En proceso", "Cotizado", "Rechazada"]
# Al guardar una respuesta hay que llegar a una decisión final: no se puede
# dejar la solicitud en "Solicitud" ni "En proceso" desde este formulario.
ESTADOS_FINALES_PRICING = ["Cotizado", "Rechazada"]


@app.route("/pricing")
@pricing_required
def pricing():
    db = get_db()
    filas = db.execute("""
        SELECT
            s.id, s.referencia, s.tipo_embarque, s.fecha_creacion, s.estado,
            co.id AS cotizacion_id, co.id_cotizacion,
            ac.razon_social AS cliente_nombre,
            op.nombre_operativo AS operativo_asignado
        FROM crm_solicitudes_maritimo_aereo s
        LEFT JOIN crm_cotizaciones co ON co.id = s.cotizacion_id
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN catalogo_operativos op ON op.id = s.operativo_asignado_id
        ORDER BY (s.estado = 'Solicitud') DESC, (s.estado = 'En proceso') DESC, s.creado_en DESC
    """).fetchall()
    db.close()
    return render_template("pricing.html", filas=filas)


@app.route("/pricing/<int:solicitud_id>")
@pricing_required
def pricing_detalle(solicitud_id):
    db = get_db()
    fila = db.execute("""
        SELECT
            s.*,
            co.id AS cotizacion_id, co.id_cotizacion,
            ac.razon_social AS cliente_nombre,
            i.nombre AS incoterm_nombre
        FROM crm_solicitudes_maritimo_aereo s
        LEFT JOIN crm_cotizaciones co ON co.id = s.cotizacion_id
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN crm_incoterms i ON i.id = s.incoterm_id
        WHERE s.id = %s
    """, (solicitud_id,)).fetchone()
    if fila is None:
        db.close()
        flash("Solicitud no encontrada.")
        return redirect(url_for("pricing"))
    operativos = db.execute("""
        SELECT co.id, co.nombre_operativo
        FROM catalogo_operativos co
        LEFT JOIN app_user_permissions p ON p.user_id = co.user_id
        WHERE co.activo = true
          AND (coalesce(p.es_admin, false) OR coalesce(p.puede_pricing, false) OR co.id = %s)
        ORDER BY co.nombre_operativo
    """, (fila["operativo_asignado_id"],)).fetchall()
    respuestas = db.execute("""
        SELECT id, respuesta, respondido_por, creado_en
        FROM crm_solicitudes_pricing_respuestas
        WHERE solicitud_id = %s
        ORDER BY creado_en DESC
    """, (solicitud_id,)).fetchall()
    db.close()
    return render_template(
        "pricing_detalle.html", fila=fila, estados=ESTADOS_FINALES_PRICING, operativos=operativos,
        respuestas=respuestas,
    )


@app.route("/pricing/<int:solicitud_id>/responder", methods=["POST"])
@pricing_required
def pricing_responder(solicitud_id):
    estado = request.form.get("estado", "").strip()
    if estado not in ESTADOS_FINALES_PRICING:
        flash("Elige Cotizado o Rechazada para guardar la respuesta.")
        return redirect(url_for("pricing_detalle", solicitud_id=solicitud_id))
    respuesta = (request.form.get("respuesta_pricing", "") or "").strip()[:4000] or None
    operativo_raw = (request.form.get("operativo_asignado_id", "") or "").strip()
    operativo_id = int(operativo_raw) if operativo_raw.isdigit() else None
    usuario = session.get("usuario", "")

    db = get_db()
    db.execute("""
        UPDATE crm_solicitudes_maritimo_aereo
        SET estado = %s, respuesta_pricing = %s, respondido_por = %s, respondido_en = now(),
            operativo_asignado_id = %s
        WHERE id = %s
    """, (estado, respuesta, usuario, operativo_id, solicitud_id))
    if respuesta:
        db.execute("""
            INSERT INTO crm_solicitudes_pricing_respuestas (solicitud_id, respuesta, respondido_por)
            VALUES (%s, %s, %s)
        """, (solicitud_id, respuesta, usuario))
    db.commit()
    db.close()
    flash("Solicitud actualizada.")
    return redirect(url_for("pricing_detalle", solicitud_id=solicitud_id))


def puede_ver_solicitud_pricing(db, fila):
    """True si el usuario en sesión puede ver esta solicitud: es de Pricing,
    o tiene acceso al CRM y puede ver la cotización de la que salió (mismo
    criterio de plazas que el resto del CRM)."""
    if usuario_puede_pricing():
        return True
    if not usuario_puede_ver_crm():
        return False
    if not fila["cotizacion_id"]:
        return False
    return cotizacion_visible_para_usuario(db, fila["cotizacion_id"])


@app.route("/pricing/<int:solicitud_id>/pdf")
@login_required
def pricing_pdf(solicitud_id):
    db = get_db()
    fila = db.execute("""
        SELECT
            s.*,
            co.id AS cotizacion_id, co.id_cotizacion,
            ac.razon_social AS cliente_nombre,
            i.nombre AS incoterm_nombre
        FROM crm_solicitudes_maritimo_aereo s
        LEFT JOIN crm_cotizaciones co ON co.id = s.cotizacion_id
        LEFT JOIN asignacion_de_clientes ac ON ac.folio = co.cliente_folio
        LEFT JOIN crm_incoterms i ON i.id = s.incoterm_id
        WHERE s.id = %s
    """, (solicitud_id,)).fetchone()
    if fila is None or not puede_ver_solicitud_pricing(db, fila):
        db.close()
        flash("Solicitud no encontrada.")
        return redirect(url_for(primera_pagina_permitida()))
    operativo = None
    if fila["operativo_asignado_id"]:
        operativo = db.execute(
            "SELECT nombre_operativo FROM catalogo_operativos WHERE id = %s", (fila["operativo_asignado_id"],)
        ).fetchone()
    respuestas = db.execute("""
        SELECT respuesta, respondido_por, creado_en
        FROM crm_solicitudes_pricing_respuestas
        WHERE solicitud_id = %s
        ORDER BY creado_en ASC
    """, (solicitud_id,)).fetchall()
    db.close()

    html = render_template(
        "pricing_pdf.html", fila=fila, operativo=operativo["nombre_operativo"] if operativo else None,
        respuestas=respuestas, generado_en=datetime.now(TZ_LOCAL),
    )
    buffer = io.BytesIO()
    resultado = pisa.CreatePDF(src=html, dest=buffer, encoding="utf-8")
    if resultado.err:
        flash("No se pudo generar el PDF de la solicitud.")
        return redirect(url_for("pricing_detalle", solicitud_id=solicitud_id))
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=request.args.get("descargar") == "1",
        download_name=f"{fila['referencia']}.pdf", mimetype="application/pdf",
    )


@app.route("/pricing/<int:solicitud_id>/ver")
@login_required
def pricing_ver(solicitud_id):
    """Pantalla intermedia para revisar el PDF de una solicitud (visor +
    botón de descarga explícito) antes de que el vendedor decida guardarlo.
    Marca la solicitud como vista, lo que le quita el aviso de "Nuevo" que
    ve el vendedor en la cotización cuando Pricing responde."""
    db = get_db()
    fila = db.execute(
        "SELECT id, referencia, cotizacion_id FROM crm_solicitudes_maritimo_aereo WHERE id = %s",
        (solicitud_id,),
    ).fetchone()
    if fila is None or not puede_ver_solicitud_pricing(db, fila):
        db.close()
        flash("Solicitud no encontrada.")
        return redirect(url_for(primera_pagina_permitida()))
    db.execute(
        "UPDATE crm_solicitudes_maritimo_aereo SET visto_por_vendedor_en = now() WHERE id = %s",
        (solicitud_id,),
    )
    db.commit()
    db.close()
    return render_template("pricing_pdf_ver.html", fila=fila)


@app.route("/crm/cotizaciones/<int:cotizacion_id>/aplicar-booking", methods=["POST"])
@crm_required
def crm_cotizacion_aplicar_booking(cotizacion_id):
    """Liga uno o varios bookings reales a la cotización: la marca como
    Ganada (basta con tener ≥1 booking aplicado). Se puede seguir
    aplicando más bookings después, pero no si ya está Perdida."""
    referencias = [r.strip() for r in request.form.getlist("booking_referencia") if r.strip()]
    if not referencias:
        flash("Selecciona al menos un booking antes de aplicar.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    db = get_db()
    cotizacion = db.execute("SELECT estatus, cliente_folio FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)).fetchone()
    if cotizacion is None:
        db.close()
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    if not usuario_puede_ver_cotizacion(db, cotizacion["cliente_folio"]):
        db.close()
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    if cotizacion["estatus"] == "perdida":
        db.close()
        flash("Esta cotización está marcada como Perdida; no se le pueden aplicar bookings.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    aplicados, ya_tomados, no_existen = [], [], []
    for referencia in referencias:
        booking = db.execute("SELECT referencia FROM reporte_bookings WHERE referencia = %s", (referencia,)).fetchone()
        if booking is None:
            no_existen.append(referencia)
            continue
        try:
            db.execute(
                "INSERT INTO crm_cotizacion_bookings (cotizacion_id, booking_referencia) VALUES (%s, %s)",
                (cotizacion_id, referencia),
            )
            db.commit()
            aplicados.append(referencia)
        except psycopg.errors.UniqueViolation:
            db.rollback()
            ya_tomados.append(referencia)
    db.close()

    mensajes = []
    if aplicados:
        mensajes.append(f"Aplicados {len(aplicados)} booking(s): {', '.join(aplicados)}.")
    if ya_tomados:
        mensajes.append(f"Ya estaban aplicados a otra cotización: {', '.join(ya_tomados)}.")
    if no_existen:
        mensajes.append(f"No existen: {', '.join(no_existen)}.")
    flash(" ".join(mensajes) if mensajes else "No se aplicó ningún booking.")
    return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))


@app.route("/crm/cotizaciones/<int:cotizacion_id>/marcar-perdida", methods=["POST"])
@crm_required
def crm_cotizacion_marcar_perdida(cotizacion_id):
    """Marca la cotización como Perdida; exige un motivo del catálogo
    (Catálogos → Motivos de Pérdida). No aplica si ya tiene bookings
    ganados — una cotización ganada no se puede volver a perder."""
    motivo_id_raw = request.form.get("motivo_perdida_id", "").strip()
    motivo_id = int(motivo_id_raw) if motivo_id_raw.isdigit() else None
    comentario = request.form.get("comentario_perdida", "").strip()[:2000]
    if motivo_id is None:
        flash("Selecciona un motivo antes de marcar la cotización como Perdida.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    db = get_db()
    cotizacion = db.execute("SELECT id, cliente_folio FROM crm_cotizaciones WHERE id = %s", (cotizacion_id,)).fetchone()
    if cotizacion is None:
        db.close()
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    if not usuario_puede_ver_cotizacion(db, cotizacion["cliente_folio"]):
        db.close()
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    tiene_bookings = db.execute(
        "SELECT 1 FROM crm_cotizacion_bookings WHERE cotizacion_id = %s LIMIT 1", (cotizacion_id,)
    ).fetchone() is not None
    if tiene_bookings:
        db.close()
        flash("Esta cotización ya está Ganada; no se puede marcar como Perdida.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))
    motivo = db.execute("SELECT id FROM crm_motivos_perdida WHERE id = %s", (motivo_id,)).fetchone()
    if motivo is None:
        db.close()
        flash("Ese motivo no existe.")
        return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))

    db.execute(
        "UPDATE crm_cotizaciones SET estatus = 'perdida', motivo_perdida_id = %s, comentario_perdida = %s, perdida_en = now() WHERE id = %s",
        (motivo_id, comentario or None, cotizacion_id),
    )
    db.commit()
    db.close()
    flash("Cotización marcada como Perdida.")
    return redirect(url_for("crm_cotizacion_detalle", cotizacion_id=cotizacion_id))


@app.route("/crm/cotizaciones/<int:cotizacion_id>/clonar", methods=["POST"])
@crm_required
def crm_cotizacion_clonar(cotizacion_id):
    nuevo_id = clonar_cotizacion_crm(cotizacion_id)
    if nuevo_id is None:
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    flash("Cotización clonada correctamente.")
    return redirect(url_for("crm_cotizacion_editar", cotizacion_id=nuevo_id))


@app.route("/crm/cotizaciones/<int:cotizacion_id>/vista")
@crm_required
def crm_cotizacion_vista(cotizacion_id):
    documento = construir_documento_cotizacion_crm(cotizacion_id)
    if documento is None:
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))
    return render_template("crm_cotizacion_vista.html", doc=documento, cotizacion_id=cotizacion_id)


@app.route("/crm/cotizaciones/<int:cotizacion_id>/pdf")
@crm_required
def crm_cotizacion_pdf(cotizacion_id):
    documento = construir_documento_cotizacion_crm(cotizacion_id)
    if documento is None:
        flash("Cotización no encontrada.")
        return redirect(url_for("crm_seccion", slug="cotizaciones"))

    html = render_template("crm_cotizacion_pdf.html", doc=documento)
    buffer = io.BytesIO()
    resultado = pisa.CreatePDF(src=html, dest=buffer, encoding="utf-8")
    if resultado.err:
        flash("No se pudo generar el PDF de la cotización.")
        return redirect(url_for("crm_cotizacion_vista", cotizacion_id=cotizacion_id))
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True, download_name=f"Cotizacion_{documento['id_cotizacion']}.pdf",
        mimetype="application/pdf",
    )


@app.route("/crm/productos", methods=["GET", "POST"])
@crm_required
def crm_productos():
    """Catálogo Productos = Tipo Ingreso/Egreso (crm_tipos_ingreso_egreso):
    un solo catálogo, visible aquí y usado como Producto en Cotizaciones."""
    db = get_db()
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        if not nombre:
            flash("El nombre es obligatorio.")
        else:
            db.execute("INSERT INTO crm_tipos_ingreso_egreso (nombre) VALUES (%s)", (nombre,))
            db.commit()
        db.close()
        return redirect(url_for("crm_productos"))

    filas = db.execute("""
        SELECT id, nombre, concepto_ingles, naturaleza, producto, porcentaje_iva, porcentaje_ret, bloqueado
        FROM crm_tipos_ingreso_egreso WHERE producto IS NOT NULL AND producto <> '' ORDER BY nombre
    """).fetchall()
    db.close()
    productos = [{
        "id": f["id"],
        "nombre": f["nombre"],
        "concepto_ingles": f["concepto_ingles"] or "",
        "naturaleza": f["naturaleza"] or "",
        "producto": f["producto"] or "",
        "porcentaje_iva": float(f["porcentaje_iva"]) if f["porcentaje_iva"] is not None else 0,
        "porcentaje_ret": float(f["porcentaje_ret"]) if f["porcentaje_ret"] is not None else 0,
        "bloqueado": f["bloqueado"],
    } for f in filas]
    nav_groups = agrupar_nav_crm("productos")
    return render_template("crm_productos.html", nav_groups=nav_groups, titulo_pagina="Productos", productos=productos)


@app.route("/crm/productos/<int:producto_id>/eliminar", methods=["POST"])
@crm_required
def crm_producto_eliminar(producto_id):
    db = get_db()
    db.execute("DELETE FROM crm_tipos_ingreso_egreso WHERE id = %s", (producto_id,))
    db.commit()
    db.close()
    return redirect(url_for("crm_productos"))


@app.route("/crm/motivos-perdida", methods=["GET", "POST"])
@crm_required
def crm_motivos_perdida():
    """Catálogo de motivos por los que se puede perder una cotización;
    obligatorio para poder marcarla como Perdida (CRM → Cotizaciones)."""
    db = get_db()
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip().upper()
        if not nombre:
            flash("El nombre es obligatorio.")
        else:
            db.execute(
                "INSERT INTO crm_motivos_perdida (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (nombre,)
            )
            db.commit()
        db.close()
        return redirect(url_for("crm_motivos_perdida"))

    filas = db.execute("SELECT id, nombre FROM crm_motivos_perdida ORDER BY nombre").fetchall()
    db.close()
    nav_groups = agrupar_nav_crm("motivos-perdida")
    return render_template("crm_motivos_perdida.html", nav_groups=nav_groups, titulo_pagina="Motivos de Pérdida", motivos=filas)


@app.route("/crm/motivos-perdida/<int:motivo_id>/eliminar", methods=["POST"])
@crm_required
def crm_motivo_perdida_eliminar(motivo_id):
    db = get_db()
    db.execute("DELETE FROM crm_motivos_perdida WHERE id = %s", (motivo_id,))
    db.commit()
    db.close()
    return redirect(url_for("crm_motivos_perdida"))


@app.route("/catalogos")
@catalogos_required
def catalogos():
    return render_template("catalogos.html")


PERMISOS_LISTA = [
    ("puede_ver_ventas", "Información de Ventas"),
    ("puede_ver_reportes", "Reportes"),
    ("puede_comisiones", "Comisiones"),
    ("puede_ver_crm", "CRM"),
    ("puede_pricing", "Pricing"),
    ("puede_ver_catalogos", "Catálogos"),
    ("puede_actualizar", "Actualizar"),
]
PERMISOS_TOGGLEABLES = {campo for campo, _ in PERMISOS_LISTA}


@app.route("/catalogos/permisos-actualizar")
@admin_required
def permisos_actualizar():
    db = get_db()
    filas = db.execute(
        """
        SELECT u.id, u.email,
            coalesce(p.es_admin, false) AS es_admin,
            coalesce(p.puede_actualizar, false) AS puede_actualizar,
            coalesce(p.puede_comisiones, false) AS puede_comisiones,
            coalesce(p.puede_ver_ventas, true) AS puede_ver_ventas,
            coalesce(p.puede_ver_reportes, true) AS puede_ver_reportes,
            coalesce(p.puede_ver_catalogos, false) AS puede_ver_catalogos,
            coalesce(p.puede_ver_crm, false) AS puede_ver_crm,
            coalesce(p.puede_pricing, false) AS puede_pricing
        FROM auth.users u
        LEFT JOIN public.app_user_permissions p ON p.user_id = u.id
        ORDER BY u.email
        """
    ).fetchall()
    db.close()
    return render_template("permisos_actualizar.html", filas=filas, permisos_lista=PERMISOS_LISTA)


@app.route("/catalogos/permisos-actualizar/<uuid:user_id>/toggle", methods=["POST"])
@admin_required
def permisos_actualizar_toggle(user_id):
    campo = request.form.get("campo") or "puede_actualizar"
    if campo not in PERMISOS_TOGGLEABLES:
        flash("Permiso inválido.")
        return redirect(url_for("permisos_actualizar"))

    nuevo_valor = request.form.get("valor") == "1"
    db = get_db()
    db.execute(
        f"""
        INSERT INTO app_user_permissions (user_id, {campo})
        VALUES (%s, %s)
        ON CONFLICT (user_id) DO UPDATE SET {campo} = EXCLUDED.{campo}, updated_at = now()
        """,
        (str(user_id), nuevo_valor),
    )
    db.commit()
    db.close()
    return redirect(url_for("permisos_actualizar"))


@app.route("/catalogos/firmas-cotizacion")
@admin_required
def firmas_cotizacion():
    db = get_db()
    filas = db.execute(
        """
        SELECT u.id, u.email, f.nombre_firma, f.puesto, f.telefono, f.correo
        FROM auth.users u
        LEFT JOIN public.crm_firmas f ON f.user_id = u.id
        ORDER BY u.email
        """
    ).fetchall()
    db.close()
    return render_template("firmas_cotizacion.html", filas=filas)


@app.route("/catalogos/firmas-cotizacion/<uuid:user_id>/guardar", methods=["POST"])
@admin_required
def firmas_cotizacion_guardar(user_id):
    nombre_firma = request.form.get("nombre_firma", "").strip()[:120]
    puesto = request.form.get("puesto", "").strip()[:120]
    telefono = request.form.get("telefono", "").strip()[:40]
    correo = request.form.get("correo", "").strip()[:120]
    db = get_db()
    db.execute(
        """
        INSERT INTO crm_firmas (user_id, nombre_firma, puesto, telefono, correo)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (user_id) DO UPDATE SET
            nombre_firma = EXCLUDED.nombre_firma, puesto = EXCLUDED.puesto,
            telefono = EXCLUDED.telefono, correo = EXCLUDED.correo, updated_at = now()
        """,
        (str(user_id), nombre_firma, puesto, telefono, correo),
    )
    db.commit()
    db.close()
    flash("Firma guardada.")
    return redirect(url_for("firmas_cotizacion"))


@app.route("/catalogos/crm-plazas")
@admin_required
def crm_plazas():
    habilitadas = plazas_con_crm_habilitado()
    filas = [{"plaza": p, "habilitada": p in habilitadas} for p in get_plazas_catalogo()]
    return render_template("crm_plazas.html", filas=filas)


@app.route("/catalogos/crm-plazas/toggle", methods=["POST"])
@admin_required
def crm_plazas_toggle():
    plaza = request.form.get("plaza", "").strip()
    if not plaza:
        flash("Plaza inválida.")
        return redirect(url_for("crm_plazas"))
    db = get_db()
    if request.form.get("valor") == "1":
        db.execute("INSERT INTO crm_plazas_habilitadas (plaza) VALUES (%s) ON CONFLICT (plaza) DO NOTHING", (plaza,))
    else:
        db.execute("DELETE FROM crm_plazas_habilitadas WHERE plaza = %s", (plaza,))
    db.commit()
    db.close()
    return redirect(url_for("crm_plazas"))


def get_plazas_catalogo():
    """Todas las plazas conocidas (unión de catalogo_vendedores y
    catalogo_desarrolladores), para poblar los checkboxes de visibilidad."""
    db = get_db()
    filas_v = db.execute("SELECT DISTINCT plaza FROM catalogo_vendedores").fetchall()
    filas_d = db.execute("SELECT DISTINCT plaza FROM catalogo_desarrolladores").fetchall()
    db.close()
    plazas = {f["plaza"] for f in filas_v if f["plaza"]} | {f["plaza"] for f in filas_d if f["plaza"]}
    return sorted(plazas)


@app.route("/catalogos/visibilidad-plazas")
@admin_required
def visibilidad_plazas():
    db = get_db()
    usuarios_filas = db.execute(
        """
        SELECT u.id, u.email,
            coalesce(p.es_admin, false) AS es_admin,
            coalesce(p.todas_las_plazas, false) AS todas_las_plazas
        FROM auth.users u
        LEFT JOIN public.app_user_permissions p ON p.user_id = u.id
        ORDER BY u.email
        """
    ).fetchall()
    plazas_por_usuario = {}
    for r in db.execute("SELECT user_id, plaza FROM app_user_plazas ORDER BY plaza"):
        plazas_por_usuario.setdefault(str(r["user_id"]), []).append(r["plaza"])
    db.close()

    filas = []
    for u in usuarios_filas:
        filas.append({
            "id": u["id"],
            "email": u["email"],
            "es_admin": u["es_admin"],
            "todas_las_plazas": u["todas_las_plazas"],
            "plazas": plazas_por_usuario.get(str(u["id"]), []),
        })
    return render_template("visibilidad_plazas.html", filas=filas)


@app.route("/catalogos/visibilidad-plazas/<uuid:user_id>/editar", methods=["GET", "POST"])
@admin_required
def visibilidad_plazas_editar(user_id):
    db = get_db()
    usuario = db.execute("SELECT id, email FROM auth.users WHERE id = %s", (str(user_id),)).fetchone()
    if usuario is None:
        db.close()
        return "No encontrado", 404

    if request.method == "POST":
        todas_las_plazas = request.form.get("todas_las_plazas") == "on"
        plazas_seleccionadas = request.form.getlist("plazas")
        if not todas_las_plazas and not plazas_seleccionadas:
            flash('Selecciona al menos una plaza, o marca "Todas las plazas".')
        else:
            db.execute(
                """
                INSERT INTO app_user_permissions (user_id, todas_las_plazas)
                VALUES (%s, %s)
                ON CONFLICT (user_id) DO UPDATE SET todas_las_plazas = EXCLUDED.todas_las_plazas, updated_at = now()
                """,
                (str(user_id), todas_las_plazas),
            )
            db.execute("DELETE FROM app_user_plazas WHERE user_id = %s", (str(user_id),))
            if not todas_las_plazas:
                for plaza in plazas_seleccionadas:
                    db.execute(
                        "INSERT INTO app_user_plazas (user_id, plaza) VALUES (%s, %s)",
                        (str(user_id), plaza),
                    )
            db.commit()
            db.close()
            return redirect(url_for("visibilidad_plazas"))

    plazas_actuales = {r["plaza"] for r in db.execute("SELECT plaza FROM app_user_plazas WHERE user_id = %s", (str(user_id),))}
    todas_las_plazas_actual = db.execute(
        "SELECT coalesce(todas_las_plazas, false) AS v FROM app_user_permissions WHERE user_id = %s", (str(user_id),)
    ).fetchone()
    db.close()
    return render_template(
        "visibilidad_plazas_editar.html",
        usuario=usuario,
        plazas_catalogo=get_plazas_catalogo(),
        plazas_actuales=plazas_actuales,
        sin_restriccion=bool(todas_las_plazas_actual and todas_las_plazas_actual["v"]),
    )


@app.route("/catalogos/actividad-usuarios")
@admin_required
def actividad_usuarios():
    db = get_db()
    filas = db.execute(
        "SELECT usuario, fecha_hora FROM registro_ingresos ORDER BY fecha_hora DESC LIMIT 500"
    ).fetchall()
    db.close()
    ingresos = [
        {"usuario": f["usuario"], "fecha_hora": f["fecha_hora"].astimezone(TZ_LOCAL).strftime("%d/%m/%Y %H:%M")}
        for f in filas
    ]
    return render_template("actividad_usuarios.html", ingresos=ingresos)


SUPABASE_DB_LIMIT_MB = float(os.environ.get("SUPABASE_DB_LIMIT_MB", "500"))


@app.route("/catalogos/almacenamiento")
@admin_required
def almacenamiento_bd():
    db = get_db()
    total = db.execute(
        "SELECT pg_database_size(current_database()) AS bytes, "
        "pg_size_pretty(pg_database_size(current_database())) AS legible"
    ).fetchone()
    tablas_filas = db.execute("""
        SELECT relname AS tabla,
               pg_total_relation_size(relid) AS bytes,
               pg_size_pretty(pg_total_relation_size(relid)) AS legible
        FROM pg_catalog.pg_statio_user_tables
        WHERE schemaname = 'public'
        ORDER BY bytes DESC
    """).fetchall()
    db.close()

    total_bytes = total["bytes"]
    limite_bytes = SUPABASE_DB_LIMIT_MB * 1024 * 1024
    porcentaje = round(total_bytes / limite_bytes * 100, 1) if limite_bytes else 0
    tablas = [
        {
            "tabla": t["tabla"],
            "legible": t["legible"],
            "porcentaje": round(t["bytes"] / total_bytes * 100, 1) if total_bytes else 0,
        }
        for t in tablas_filas
    ]
    return render_template(
        "almacenamiento_bd.html",
        total_legible=total["legible"],
        limite_mb=SUPABASE_DB_LIMIT_MB,
        porcentaje=porcentaje,
        tablas=tablas,
    )


@app.route("/catalogos/vendedores", methods=["GET", "POST"])
@admin_required
def catalogo_vendedores():
    db = get_db()
    if request.method == "POST":
        vendedor = request.form.get("vendedor", "").strip()
        plaza = request.form.get("plaza", "").strip()
        ocultar_detalle = request.form.get("ocultar_detalle") == "on"
        if not vendedor or not plaza:
            flash("Vendedor y Plaza son obligatorios.")
        else:
            try:
                db.execute(
                    "INSERT INTO catalogo_vendedores (vendedor, plaza, ocultar_detalle) VALUES (%s, %s, %s)",
                    (vendedor, plaza, ocultar_detalle),
                )
                db.commit()
            except psycopg.errors.UniqueViolation:
                db.rollback()
                flash("Ese Vendedor ya existe en el catálogo.")
        db.close()
        return redirect(url_for("catalogo_vendedores"))

    filas = db.execute("SELECT * FROM catalogo_vendedores ORDER BY plaza, vendedor").fetchall()
    db.close()
    return render_template("catalogo_vendedores.html", filas=filas)


@app.route("/catalogos/vendedores/<int:fila_id>/editar", methods=["GET", "POST"])
@admin_required
def catalogo_vendedores_editar(fila_id):
    db = get_db()
    if request.method == "POST":
        vendedor = request.form.get("vendedor", "").strip()
        plaza = request.form.get("plaza", "").strip()
        ocultar_detalle = request.form.get("ocultar_detalle") == "on"
        if not vendedor or not plaza:
            flash("Vendedor y Plaza son obligatorios.")
        else:
            try:
                db.execute(
                    "UPDATE catalogo_vendedores SET vendedor = %s, plaza = %s, ocultar_detalle = %s WHERE id = %s",
                    (vendedor, plaza, ocultar_detalle, fila_id),
                )
                db.commit()
            except psycopg.errors.UniqueViolation:
                db.rollback()
                flash("Ese Vendedor ya existe en el catálogo.")
            db.close()
            return redirect(url_for("catalogo_vendedores"))

    fila = db.execute("SELECT * FROM catalogo_vendedores WHERE id = %s", (fila_id,)).fetchone()
    db.close()
    if fila is None:
        return "No encontrado", 404
    return render_template("catalogo_vendedores_editar.html", fila=fila)


@app.route("/catalogos/vendedores/<int:fila_id>/eliminar", methods=["POST"])
@admin_required
def catalogo_vendedores_eliminar(fila_id):
    db = get_db()
    db.execute("DELETE FROM catalogo_vendedores WHERE id = %s", (fila_id,))
    db.commit()
    db.close()
    return redirect(url_for("catalogo_vendedores"))


@app.route("/catalogos/desarrolladores", methods=["GET", "POST"])
@admin_required
def catalogo_desarrolladores():
    db = get_db()
    if request.method == "POST":
        desarrollador = request.form.get("desarrollador", "").strip()
        plaza = request.form.get("plaza", "").strip()
        if not desarrollador or not plaza:
            flash("Desarrollador y Plaza son obligatorios.")
        else:
            try:
                db.execute("INSERT INTO catalogo_desarrolladores (desarrollador, plaza) VALUES (%s, %s)", (desarrollador, plaza))
                db.commit()
            except psycopg.errors.UniqueViolation:
                db.rollback()
                flash("Ese Desarrollador ya existe en el catálogo.")
        db.close()
        return redirect(url_for("catalogo_desarrolladores"))

    filas = db.execute("SELECT * FROM catalogo_desarrolladores ORDER BY plaza, desarrollador").fetchall()
    db.close()
    return render_template("catalogo_desarrolladores.html", filas=filas)


@app.route("/catalogos/desarrolladores/<int:fila_id>/editar", methods=["GET", "POST"])
@admin_required
def catalogo_desarrolladores_editar(fila_id):
    db = get_db()
    if request.method == "POST":
        desarrollador = request.form.get("desarrollador", "").strip()
        plaza = request.form.get("plaza", "").strip()
        if not desarrollador or not plaza:
            flash("Desarrollador y Plaza son obligatorios.")
        else:
            try:
                db.execute(
                    "UPDATE catalogo_desarrolladores SET desarrollador = %s, plaza = %s WHERE id = %s",
                    (desarrollador, plaza, fila_id),
                )
                db.commit()
            except psycopg.errors.UniqueViolation:
                db.rollback()
                flash("Ese Desarrollador ya existe en el catálogo.")
            db.close()
            return redirect(url_for("catalogo_desarrolladores"))

    fila = db.execute("SELECT * FROM catalogo_desarrolladores WHERE id = %s", (fila_id,)).fetchone()
    db.close()
    if fila is None:
        return "No encontrado", 404
    return render_template("catalogo_desarrolladores_editar.html", fila=fila)


@app.route("/catalogos/desarrolladores/<int:fila_id>/eliminar", methods=["POST"])
@admin_required
def catalogo_desarrolladores_eliminar(fila_id):
    db = get_db()
    db.execute("DELETE FROM catalogo_desarrolladores WHERE id = %s", (fila_id,))
    db.commit()
    db.close()
    return redirect(url_for("catalogo_desarrolladores"))


@app.route("/catalogos/presupuesto", methods=["GET", "POST"])
@admin_required
def catalogo_presupuesto():
    db = get_db()
    if request.method == "POST":
        mes = request.form.get("mes", "").strip()
        vendedor = request.form.get("vendedor", "").strip() or None
        desarrollador = request.form.get("desarrollador", "").strip() or None
        presupuesto_raw = request.form.get("presupuesto", "")
        if not mes or not presupuesto_raw:
            flash("Mes y Presupuesto son obligatorios.")
        elif not vendedor and not desarrollador:
            flash("Debes indicar al menos un Vendedor o un Desarrollador.")
        else:
            try:
                presupuesto = parse_presupuesto(presupuesto_raw)
            except ValueError:
                flash("El presupuesto debe ser un número válido.")
                presupuesto = None
            if presupuesto is not None:
                try:
                    db.execute(
                        "INSERT INTO catalogo_presupuesto (mes, vendedor, desarrollador, presupuesto) VALUES (%s, %s, %s, %s)",
                        (mes, vendedor, desarrollador, presupuesto),
                    )
                    db.commit()
                except psycopg.errors.UniqueViolation:
                    db.rollback()
                    flash("Ya existe un presupuesto para ese Mes, Vendedor y Desarrollador.")
        db.close()
        return redirect(url_for("catalogo_presupuesto"))

    filas = db.execute("SELECT * FROM catalogo_presupuesto ORDER BY mes DESC, vendedor").fetchall()
    vendedores = get_vendedores(db)
    desarrolladores = get_desarrolladores(db)
    db.close()
    return render_template(
        "catalogo_presupuesto.html", filas=filas, opciones_mes=opciones_mes(),
        mes_actual=datetime.now(TZ_LOCAL).strftime("%Y-%m"),
        vendedores=vendedores, desarrolladores=desarrolladores,
    )


@app.route("/catalogos/presupuesto/<int:fila_id>/editar", methods=["GET", "POST"])
@admin_required
def catalogo_presupuesto_editar(fila_id):
    db = get_db()
    if request.method == "POST":
        mes = request.form.get("mes", "").strip()
        vendedor = request.form.get("vendedor", "").strip() or None
        desarrollador = request.form.get("desarrollador", "").strip() or None
        presupuesto_raw = request.form.get("presupuesto", "")
        if not mes or not presupuesto_raw:
            flash("Mes y Presupuesto son obligatorios.")
        elif not vendedor and not desarrollador:
            flash("Debes indicar al menos un Vendedor o un Desarrollador.")
        else:
            try:
                presupuesto = parse_presupuesto(presupuesto_raw)
            except ValueError:
                flash("El presupuesto debe ser un número válido.")
                presupuesto = None
            if presupuesto is not None:
                try:
                    db.execute(
                        "UPDATE catalogo_presupuesto SET mes = %s, vendedor = %s, desarrollador = %s, presupuesto = %s WHERE id = %s",
                        (mes, vendedor, desarrollador, presupuesto, fila_id),
                    )
                    db.commit()
                except psycopg.errors.UniqueViolation:
                    db.rollback()
                    flash("Ya existe un presupuesto para ese Mes, Vendedor y Desarrollador.")
                db.close()
                return redirect(url_for("catalogo_presupuesto"))

    fila = db.execute("SELECT * FROM catalogo_presupuesto WHERE id = %s", (fila_id,)).fetchone()
    vendedores = get_vendedores(db)
    desarrolladores = get_desarrolladores(db)
    db.close()
    if fila is None:
        return "No encontrado", 404
    return render_template(
        "catalogo_presupuesto_editar.html", fila=fila, opciones_mes=opciones_mes(),
        vendedores=vendedores, desarrolladores=desarrolladores,
    )


@app.route("/catalogos/presupuesto/<int:fila_id>/eliminar", methods=["POST"])
@admin_required
def catalogo_presupuesto_eliminar(fila_id):
    db = get_db()
    db.execute("DELETE FROM catalogo_presupuesto WHERE id = %s", (fila_id,))
    db.commit()
    db.close()
    return redirect(url_for("catalogo_presupuesto"))


@app.route("/catalogos/presupuesto/plantilla")
@admin_required
def catalogo_presupuesto_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws.append(PRESUPUESTO_COLUMNAS)
    ws.append(["2026-07", "ARMANDO VILLANUEVA SILVA", "", 20000])
    for col_idx, ancho in enumerate([12, 30, 22, 16], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_presupuesto.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/catalogos/presupuesto/carga_masiva", methods=["GET", "POST"])
@admin_required
def catalogo_presupuesto_carga_masiva():
    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or archivo.filename == "":
            flash("Selecciona un archivo .csv o .xlsx para cargar.")
            return redirect(url_for("catalogo_presupuesto_carga_masiva"))

        nombre = archivo.filename.lower()
        try:
            if nombre.endswith(".csv"):
                filas_crudas = leer_filas_csv(archivo)
            elif nombre.endswith(".xlsx"):
                filas_crudas = leer_filas_xlsx(archivo)
            else:
                flash("Formato no soportado. Usa un archivo .csv o .xlsx.")
                return redirect(url_for("catalogo_presupuesto_carga_masiva"))
        except Exception:
            flash("No se pudo leer el archivo. Verifica que tenga las columnas mes, vendedor, desarrollador, presupuesto.")
            return redirect(url_for("catalogo_presupuesto_carga_masiva"))

        db = get_db()
        agregados = 0
        actualizados = 0
        errores = []
        for numero_fila, fila in enumerate(filas_crudas, start=2):
            mes = str(fila.get("mes") or "").strip()
            vendedor = str(fila.get("vendedor") or "").strip() or None
            desarrollador = str(fila.get("desarrollador") or "").strip() or None
            presupuesto_raw = fila.get("presupuesto")

            if not mes or presupuesto_raw in (None, ""):
                errores.append(f"Fila {numero_fila}: faltan datos obligatorios (mes o presupuesto).")
                continue
            if not vendedor and not desarrollador:
                errores.append(f"Fila {numero_fila}: debe indicar Vendedor o Desarrollador.")
                continue
            if mes not in MESES_VALIDOS:
                errores.append(f"Fila {numero_fila}: mes '{mes}' inválido, usa formato AAAA-MM.")
                continue
            try:
                presupuesto = parse_presupuesto(str(presupuesto_raw))
            except ValueError:
                errores.append(f"Fila {numero_fila}: presupuesto '{presupuesto_raw}' no es numérico.")
                continue

            existente = db.execute(
                "SELECT id FROM catalogo_presupuesto WHERE mes = %s AND coalesce(vendedor, '') = coalesce(%s, '') AND coalesce(desarrollador, '') = coalesce(%s, '')",
                (mes, vendedor, desarrollador),
            ).fetchone()
            if existente:
                db.execute("UPDATE catalogo_presupuesto SET presupuesto = %s WHERE id = %s", (presupuesto, existente["id"]))
                actualizados += 1
            else:
                db.execute(
                    "INSERT INTO catalogo_presupuesto (mes, vendedor, desarrollador, presupuesto) VALUES (%s, %s, %s, %s)",
                    (mes, vendedor, desarrollador, presupuesto),
                )
                agregados += 1

        db.commit()
        db.close()

        mensaje = f"Carga completa: {agregados} agregados, {actualizados} actualizados."
        if errores:
            mensaje += f" {len(errores)} fila(s) con errores."
        session["resultado_carga"] = {"mensaje": mensaje, "ok": not errores, "errores": errores[:50]}
        return redirect(url_for("catalogo_presupuesto_carga_masiva"))

    resultado = session.pop("resultado_carga", None)
    return render_template("catalogo_presupuesto_carga_masiva.html", resultado=resultado)


if __name__ == "__main__":
    app.config["TEMPLATES_AUTO_RELOAD"] = True
    app.run(host=os.environ.get("HOST", "127.0.0.1"), port=int(os.environ.get("PORT", 5050)), debug=False)
